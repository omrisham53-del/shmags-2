"""
Capex pipeline: category-filter -> file-match -> extract -> classify -> average.

Ties together the category table (the master CSV with one row per grant request)
and the individual grant Excel files, to produce per-technology average CapEx
for the techno-economic model.

Steps:
  1. Read the category CSV (CP862 / DOS-Hebrew encoded), filter to the relevant
     consolidated categories (col F) plus הסבה-to-electricity rows.
  2. Walk the round folder; each request lives in its own subfolder whose name
     contains the request ID. Pick the best grant-form Excel per request
     (prefer 'בדיקה' reviewed file, then highest version, then newest).
  3. Extract line items from each matched file (via extract_capex.py).
  4. Classify each line item to one of the 4 model technologies by keyword.
  5. Aggregate per request per technology and compute averages.

Outputs (written next to the grant files):
  - capex_lineitems.csv       every extracted line item, with technology tag
  - capex_by_request.csv      per request per technology: core-equipment sum + full-site total
  - capex_averages.csv        average cost per installation per technology (core-only and full)
  - capex_coverage.txt        ID/file match coverage and validation warnings

Usage:
    python capex_pipeline.py <category_csv> <folder_with_grant_xlsx>
"""

import sys
import os
import re
import csv
from collections import defaultdict

from extract_capex import extract_all_items, extract_any_identity
import openpyxl
import warnings

# --- Category filter (column F = איחוד קטגוריות) ---
RELEVANT_CATEGORIES = {
    "אקלום מבנים",
    "חימום מים",
    "מדחסים",
    "כללי/אחר/משולב",
    "פרויקטים המשלבים מס' טכנולוגיות",
}
CONVERSION_CATEGORY = "הסבה"          # only the "...לחשמל" rows of this category

# CSV column indexes (0-based)
COL_ID = 0          # מספר בקשה
COL_CAT_DETAIL = 4  # קטגוריה
COL_CAT_GROUP = 5   # איחוד קטגוריות
COL_TOTAL_INV = 7   # סה"כ השקעה

CSV_ENCODING = "cp862"

# --- Technology classification keywords (matched against system + component text) ---
# Order matters: first match wins.
TECH_KEYWORDS = [
    ("מערכות קיטור חשמליות", ["מחולל קיטור", "דוד קיטור", "גנרטור קיטור", "קיטור חשמלי", "קיטור"]),
    ("משאבות חום",           ["משאבת חום", "משאבות חום", "משאב"]),
    ("צ'ילרים",              ["צ'ילר", "צ׳ילר", "צ'ילרים", "צ׳ילרים", "מקרר מים"]),
    ("מדחסי VSD",            ["מדחס", "vsd", "מהפך תדר", "וי.אס.די"]),
]

# Supporting / balance-of-system items (not the core technology)
SUPPORT_KEYWORDS = [
    "טיפול במים", "מרכך מים", "ריכוך מים",
    "בקרה", "מנייה", "מנית", "ניטור", "מונה", "מד ",
    "תשתית חשמל", "תשתית", "צנרת", "חשמל", "התקנה", "חיווט", "לוח חשמל",
]


def classify_line_item(system, component):
    """Return (technology, is_core). technology is one of the 4 model techs,
    or 'תומך' (supporting/BoS) or 'לא מסווג' (unclassified)."""
    text = f"{system} {component}".lower()
    for tech, kws in TECH_KEYWORDS:
        for kw in kws:
            if kw.lower() in text:
                return tech, True
    for kw in SUPPORT_KEYWORDS:
        if kw.lower() in text:
            return "תומך", False
    return "לא מסווג", False


def load_selected_requests(csv_path):
    """Return dict {request_id: {'category':..., 'total_inv':...}} for relevant rows."""
    selected = {}
    with open(csv_path, encoding=CSV_ENCODING) as f:
        reader = csv.reader(f)
        next(reader, None)  # header
        for r in reader:
            if len(r) <= COL_CAT_GROUP:
                continue
            req_id = r[COL_ID].strip()
            if not req_id:
                continue
            cat_group = r[COL_CAT_GROUP].strip()
            cat_detail = r[COL_CAT_DETAIL].strip()

            category = None
            if cat_group in RELEVANT_CATEGORIES:
                category = cat_group
            elif cat_group == CONVERSION_CATEGORY and "לחשמל" in cat_detail:
                category = "הסבה-חשמל"

            if category is None:
                continue

            total_inv = None
            if len(r) > COL_TOTAL_INV:
                try:
                    total_inv = float(r[COL_TOTAL_INV])
                except (ValueError, TypeError):
                    total_inv = None

            selected[req_id] = {"category": category, "total_inv": total_inv}
    return selected


REVIEW_MARKER = "בדיקה"   # reviewed/working version, usually the one to read


def file_priority(path):
    """Sort key for choosing the best file in a request folder. Higher = better.
    Priority: (is reviewed 'בדיקה') > (highest numeric version) > (most recent mtime).
    Version tokens look like 0.1 / 1.2 / v2 — request IDs (5-7 digit integers)
    are deliberately excluded so they aren't mistaken for versions."""
    name = os.path.basename(path)
    is_review = 1 if REVIEW_MARKER in name else 0
    # decimal version tokens like 0.1, 1.2 (not whole integers, which could be IDs)
    decimals = re.findall(r"(?<!\d)(\d{1,2}\.\d{1,2})(?!\d)", name)
    # explicit v2 / גרסה 3 style
    vtokens = re.findall(r"(?:v|גרסה|גרסא)\s*(\d{1,2})", name, flags=re.IGNORECASE)
    version = max([float(d) for d in decimals] + [float(v) for v in vtokens],
                  default=-1.0)
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        mtime = 0
    return (is_review, version, mtime)


def is_grant_form(path):
    """True if the workbook is a grant form (2025 site-tab format or 2017 tech-tab format)."""
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        ok = (
            any(s in sheets for s in ["אתר 1", "סיכום ובקשת המענק", "אתר 2", "אתר 3"])
            or "1. פרטים כלליים ועלויות" in sheets
        )
        wb.close()
        return ok
    except Exception:
        return False


def discover_grant_files(root, wanted_ids):
    """Walk the round folder tree. Group .xlsx files by the request ID found in
    their path (folder name or filename), then pick the best grant-form file per
    request. Returns (matches {id: path}, selection_log list)."""
    wanted = set(wanted_ids)
    candidates = defaultdict(list)   # id -> [paths]

    # followlinks=True is required: on SharePoint/OneDrive the request subfolders
    # are reparse points, which os.walk skips by default.
    for dirpath, _dirs, files in os.walk(root, followlinks=True):
        for fn in files:
            if not fn.lower().endswith(".xlsx") or fn.startswith("~$"):
                continue
            full = os.path.join(dirpath, fn)
            hay = dirpath + os.sep + fn
            ids_in_path = set(re.findall(r"\d{5,7}", hay))
            for rid in (wanted & ids_in_path):
                candidates[rid].append(full)

    matches = {}
    selection_log = []
    for rid, paths in candidates.items():
        ranked = sorted(paths, key=file_priority, reverse=True)
        chosen = next((p for p in ranked if is_grant_form(p)), None)
        if chosen is None:
            selection_log.append({
                "request_id": rid, "chosen": "(none valid)",
                "n_candidates": len(paths),
                "others": " | ".join(os.path.basename(p) for p in ranked),
            })
            continue
        matches[rid] = chosen
        others = [os.path.basename(p) for p in ranked if p != chosen]
        selection_log.append({
            "request_id": rid, "chosen": os.path.basename(chosen),
            "n_candidates": len(paths),
            "others": " | ".join(others),
        })
    return matches, selection_log


def main():
    if len(sys.argv) < 3:
        print("Usage: python capex_pipeline.py <category_csv> <folder_with_grant_xlsx>")
        sys.exit(1)

    csv_path = sys.argv[1]
    folder = sys.argv[2]

    if not os.path.isfile(csv_path):
        print(f"Category CSV not found: {csv_path}")
        sys.exit(1)
    if not os.path.isdir(folder):
        print(f"Grant folder not found: {folder}")
        sys.exit(1)

    selected = load_selected_requests(csv_path)
    print(f"Selected {len(selected)} requests from category table.")

    matches, selection_log = discover_grant_files(folder, selected.keys())
    print(f"Matched {len(matches)} request IDs to files "
          f"({len(selected) - len(matches)} unmatched).")

    # Write the file-selection log so the chosen file per request can be audited
    sel_path = os.path.join(folder, "capex_file_selection.csv")
    with open(sel_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["request_id", "chosen", "n_candidates", "others"])
        w.writeheader()
        w.writerows(sorted(selection_log, key=lambda r: r["request_id"]))

    lineitems = []           # every line item with tech tag
    coverage_notes = []
    # per (request, tech): core sum; per request: full total
    core_by_req_tech = defaultdict(float)
    full_by_req = defaultdict(float)
    req_category = {}

    for req_id, fp in sorted(matches.items()):
        info = selected[req_id]
        req_category[req_id] = info["category"]
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            try:
                wb = openpyxl.load_workbook(fp, data_only=True)
            except Exception as e:
                coverage_notes.append(f"{req_id}: failed to open — {e}")
                continue

        company, tax_id = extract_any_identity(wb)

        items, warn = extract_all_items(wb)
        coverage_notes.extend([f"{req_id} / {w.strip()}" for w in warn])
        for it in items:
            tech, is_core = classify_line_item(it["system"], it["component"])
            lineitems.append({
                "request_id": req_id,
                "category": info["category"],
                "tax_id": tax_id,
                "company_name": company,
                "technology": tech,
                "is_core": is_core,
                **it,
            })
            full_by_req[req_id] += it["cost_ils"]
            if is_core:
                core_by_req_tech[(req_id, tech)] += it["cost_ils"]

        # Cross-check against category-table total investment
        inv = info["total_inv"]
        if inv is not None and full_by_req.get(req_id):
            if inv > 0 and abs(full_by_req[req_id] - inv) / inv > 0.5:
                coverage_notes.append(
                    f"{req_id}: extracted line-item total {full_by_req[req_id]:,.0f} "
                    f"differs >50% from table total investment {inv:,.0f} "
                    f"(grant file may include only part of project, or vice versa)"
                )

    # --- Write line items ---
    li_path = os.path.join(folder, "capex_lineitems.csv")
    li_fields = ["request_id", "category", "tax_id", "company_name", "technology",
                 "is_core", "site_sheet", "site_name", "system", "component",
                 "cost_ils", "notes"]
    with open(li_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=li_fields)
        w.writeheader()
        w.writerows(lineitems)

    # --- Per request per technology ---
    br_path = os.path.join(folder, "capex_by_request.csv")
    with open(br_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["request_id", "category", "technology",
                    "core_equipment_cost", "full_site_total"])
        for (req_id, tech), core in sorted(core_by_req_tech.items()):
            w.writerow([req_id, req_category.get(req_id, ""), tech,
                        f"{core:.0f}", f"{full_by_req.get(req_id, 0):.0f}"])

    # --- Averages per technology ---
    # Average is over the set of requests that had a core item for that tech.
    by_tech_costs = defaultdict(list)
    for (req_id, tech), core in core_by_req_tech.items():
        if tech in [t for t, _ in TECH_KEYWORDS]:
            by_tech_costs[tech].append(core)

    avg_path = os.path.join(folder, "capex_averages.csv")
    with open(avg_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["technology", "n_installations",
                    "avg_core_equipment_cost", "median_core_equipment_cost",
                    "min", "max"])
        for tech, _ in TECH_KEYWORDS:
            costs = sorted(by_tech_costs.get(tech, []))
            if not costs:
                w.writerow([tech, 0, "", "", "", ""])
                continue
            n = len(costs)
            avg = sum(costs) / n
            median = costs[n // 2] if n % 2 else (costs[n // 2 - 1] + costs[n // 2]) / 2
            w.writerow([tech, n, f"{avg:.0f}", f"{median:.0f}",
                        f"{costs[0]:.0f}", f"{costs[-1]:.0f}"])

    # --- Coverage / warnings ---
    cov_path = os.path.join(folder, "capex_coverage.txt")
    with open(cov_path, "w", encoding="utf-8") as f:
        f.write(f"Selected requests: {len(selected)}\n")
        f.write(f"Matched to files: {len(matches)}\n")
        f.write(f"Unmatched IDs: {len(selected) - len(matches)}\n\n")
        no_valid = [s["request_id"] for s in selection_log if s["chosen"] == "(none valid)"]
        if no_valid:
            f.write("FOUND FOLDER BUT NO VALID GRANT FORM:\n  " + ", ".join(sorted(no_valid)) + "\n\n")
        unmatched = sorted(set(selected) - set(matches))
        if unmatched:
            f.write("UNMATCHED IDs (no file found):\n  " + ", ".join(unmatched) + "\n\n")
        if coverage_notes:
            f.write("NOTES / VALIDATION:\n")
            for n in coverage_notes:
                f.write(f"  {n}\n")

    print(f"\nDone.")
    print(f"  Line items:     {li_path}  ({len(lineitems)} rows)")
    print(f"  By request/tech:{br_path}")
    print(f"  Averages:       {avg_path}")
    print(f"  File selection: {sel_path}")
    print(f"  Coverage:       {cov_path}")


if __name__ == "__main__":
    main()
