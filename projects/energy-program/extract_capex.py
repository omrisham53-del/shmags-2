"""
Extract line-item Capex from Ministry of Energy grant request Excel files.

For each file, scans all site sheets (אתר 1/2/3) for the cost table header
("עלות ב-₪") to locate the right rows/columns dynamically, then reads every
equipment row until the סה"כ total row. Also pulls company name and ח.פ by
searching for their label cells.

Output: capex_output.csv  (one row per line item)
        capex_warnings.txt (files that had missing/unexpected structure)

Usage:
    python extract_capex.py <folder_with_xlsx_files>
"""

import sys
import os
import csv
import warnings
import openpyxl

SITE_SHEETS = ["אתר 1", "אתר 2", "אתר 3"]
COST_HEADER_MARKER = "עלות ב-₪"
TOTAL_MARKER = 'סה"כ'
COMPANY_NAME_LABEL = "שם רשמי של החברה"
TAX_ID_LABEL = 'מספר ח.פ'

# 2017-era form: costs are in one flat table in the general-details sheet
FORM_2017_COST_SHEET = "1. פרטים כלליים ועלויות"
COMPANY_NAME_LABEL_2017 = "שם היזם"
TAX_ID_LABEL_2017 = "ח.פ"   # 2017 label is "ח.פ/ מספר רישום..." vs 2025 "מספר ח.פ"


def find_cell_by_label(ws, label):
    """Return the first cell whose string value contains `label`."""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and label in cell.value:
                return cell
    return None


def extract_identity(wb):
    """Pull company name and ח.פ from the general details sheet."""
    ws = wb["פרטים כלליים ואסמכתאות"] if "פרטים כלליים ואסמכתאות" in wb.sheetnames else None
    if ws is None:
        return "", ""

    company_name = ""
    tax_id = ""

    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            if COMPANY_NAME_LABEL in cell.value:
                # Value is one column to the right
                neighbor = ws.cell(row=cell.row, column=cell.column + 1)
                if neighbor.value:
                    company_name = str(neighbor.value).strip()
            if TAX_ID_LABEL in cell.value:
                neighbor = ws.cell(row=cell.row, column=cell.column + 1)
                if neighbor.value:
                    tax_id = str(neighbor.value).strip()

    return company_name, tax_id


def extract_site_items(ws, site_name):
    """
    Find the cost table header by searching for COST_HEADER_MARKER,
    then read every line item until the סה"כ row.

    Returns list of dicts and a list of warning strings.
    """
    items = []
    warn = []

    # --- Find header row and cost column ---
    header_cell = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and COST_HEADER_MARKER in cell.value:
                header_cell = cell
                break
        if header_cell:
            break

    if header_cell is None:
        warn.append(f"  {site_name}: cost table header not found, skipping")
        return items, warn

    header_row = header_cell.row
    cost_col = header_cell.column   # column index (1-based)

    # Expect system label 2 cols left, component label 1 col left of cost col
    system_col = cost_col - 2
    component_col = cost_col - 1
    notes_col = cost_col + 1

    # --- Read data rows ---
    site_name_cell = ws.cell(row=6, column=2)  # row 6 col B = site name
    site_display = str(site_name_cell.value).strip() if site_name_cell.value else site_name

    # The row immediately after the header is the template's example row
    # (always "מרכך מים ... מסוג XXX של חברת XXX בהספק של YYY"). Skip it.
    example_row = header_row + 1
    reported_total = None

    for r in range(header_row + 1, ws.max_row + 1):
        b_val = ws.cell(row=r, column=cost_col - 2).value  # system col
        cost_val = ws.cell(row=r, column=cost_col).value

        # Stop at סה"כ row, capturing the reported total for validation
        if isinstance(b_val, str) and TOTAL_MARKER in b_val:
            if isinstance(cost_val, (int, float)):
                reported_total = cost_val
            break

        # Skip the template example row
        if r == example_row:
            continue

        # Skip rows with no numeric cost
        if not isinstance(cost_val, (int, float)):
            continue
        if cost_val == 0:
            continue

        system = str(ws.cell(row=r, column=system_col).value or "").strip()
        component = str(ws.cell(row=r, column=component_col).value or "").strip()
        notes = str(ws.cell(row=r, column=notes_col).value or "").strip()

        items.append({
            "site_sheet": site_name,
            "site_name": site_display,
            "system": system,
            "component": component,
            "cost_ils": cost_val,
            "notes": notes,
        })

    # Validate: extracted line items should sum to the reported סה"כ total
    if reported_total is not None and items:
        extracted_sum = sum(it["cost_ils"] for it in items)
        if abs(extracted_sum - reported_total) > 1:  # allow rounding
            warn.append(
                f"  {site_name}: sum mismatch — extracted {extracted_sum:,.0f} "
                f"vs reported total {reported_total:,.0f}"
            )

    return items, warn


def detect_form_version(wb):
    """Return '2025', '2017', or 'unknown' based on sheet names."""
    sheets = wb.sheetnames
    if any(s in sheets for s in ["אתר 1", "סיכום ובקשת המענק"]):
        return "2025"
    if FORM_2017_COST_SHEET in sheets:
        return "2017"
    return "unknown"


def extract_2017_identity(wb):
    """Pull company name and ח.פ from the 2017-format general/costs sheet."""
    ws = wb[FORM_2017_COST_SHEET] if FORM_2017_COST_SHEET in wb.sheetnames else None
    if ws is None:
        return "", ""

    company_name = ""
    tax_id = ""

    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            v = cell.value
            if not company_name and COMPANY_NAME_LABEL_2017 in v:
                neighbor = ws.cell(row=cell.row, column=cell.column + 1)
                if neighbor.value:
                    company_name = str(neighbor.value).strip()
            if not tax_id and TAX_ID_LABEL_2017 in v:
                neighbor = ws.cell(row=cell.row, column=cell.column + 1)
                if neighbor.value:
                    tax_id = str(neighbor.value).strip()
        if company_name and tax_id:
            break

    return company_name, tax_id


def extract_2017_items(wb):
    """Extract investment line items from a 2017-format grant file.

    Reads the flat equipment table in '1. פרטים כלליים ועלויות':
      col C = equipment description (component)
      col D = technology category (system)
      col E = cost in ILS
      col F = site name
    Header row is located dynamically by searching for COST_HEADER_MARKER.
    Returns (items, warnings) with the same field structure as extract_site_items().
    """
    items = []
    warn = []
    sheet_name = FORM_2017_COST_SHEET
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
    if ws is None:
        warn.append(f"  {sheet_name}: sheet not found")
        return items, warn

    header_cell = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and COST_HEADER_MARKER in cell.value:
                header_cell = cell
                break
        if header_cell:
            break

    if header_cell is None:
        warn.append(f"  {sheet_name}: cost table header not found")
        return items, warn

    header_row = header_cell.row
    cost_col = header_cell.column  # col E (5)
    component_col = cost_col - 2   # col C (3): equipment description
    system_col = cost_col - 1      # col D (4): technology category
    site_col = cost_col + 1        # col F (6): site address

    reported_total = None

    for r in range(header_row + 1, ws.max_row + 1):
        system_val = ws.cell(row=r, column=system_col).value
        cost_val = ws.cell(row=r, column=cost_col).value

        if isinstance(system_val, str) and TOTAL_MARKER in system_val:
            if isinstance(cost_val, (int, float)):
                reported_total = cost_val
            break

        if not isinstance(cost_val, (int, float)) or cost_val == 0:
            continue

        component = str(ws.cell(row=r, column=component_col).value or "").strip()
        system = str(system_val or "").strip()
        site_name = str(ws.cell(row=r, column=site_col).value or "").strip()

        items.append({
            "site_sheet": sheet_name,
            "site_name": site_name,
            "system": system,
            "component": component,
            "cost_ils": cost_val,
            "notes": "",
        })

    if reported_total is not None and items:
        extracted_sum = sum(it["cost_ils"] for it in items)
        if abs(extracted_sum - reported_total) > 1:
            warn.append(
                f"  {sheet_name}: sum mismatch — extracted {extracted_sum:,.0f} "
                f"vs reported total {reported_total:,.0f}"
            )

    return items, warn


def extract_any_identity(wb):
    """Pull company identity from either the 2025 or 2017 form format."""
    if detect_form_version(wb) == "2017":
        return extract_2017_identity(wb)
    return extract_identity(wb)


def extract_all_items(wb):
    """Extract all line items from either 2025 or 2017 format.
    Returns (items, warnings)."""
    if detect_form_version(wb) == "2017":
        return extract_2017_items(wb)
    all_items, all_warn = [], []
    for site in SITE_SHEETS:
        if site not in wb.sheetnames:
            continue
        items, warn = extract_site_items(wb[site], site)
        all_items.extend(items)
        all_warn.extend(warn)
    return all_items, all_warn


def process_file(filepath):
    """Return (rows, warnings) for one Excel file."""
    filename = os.path.basename(filepath)
    rows = []
    file_warnings = []

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            return [], [f"{filename}: failed to open — {e}"]

    company_name, tax_id = extract_identity(wb)

    for site in SITE_SHEETS:
        ws = wb[site] if site in wb.sheetnames else None
        if ws is None:
            continue
        items, warn = extract_site_items(ws, site)
        file_warnings.extend([f"{filename} / {w}" for w in warn])
        for item in items:
            rows.append({
                "file": filename,
                "tax_id": tax_id,
                "company_name": company_name,
                **item,
            })

    if not rows and not file_warnings:
        file_warnings.append(f"{filename}: no line items found (all sites empty or zero)")

    return rows, file_warnings


def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_capex.py <folder_with_xlsx_files>")
        sys.exit(1)

    folder = sys.argv[1]
    if not os.path.isdir(folder):
        print(f"Not a directory: {folder}")
        sys.exit(1)

    files = sorted(
        os.path.join(folder, f)
        for f in os.listdir(folder)
        if f.lower().endswith(".xlsx") and not f.startswith("~")
    )

    if not files:
        print("No .xlsx files found.")
        sys.exit(1)

    print(f"Found {len(files)} files. Processing...")

    all_rows = []
    all_warnings = []

    for i, fp in enumerate(files, 1):
        rows, warns = process_file(fp)
        all_rows.extend(rows)
        all_warnings.extend(warns)
        if i % 20 == 0 or i == len(files):
            print(f"  {i}/{len(files)} done, {len(all_rows)} items so far")

    # Write CSV
    out_csv = os.path.join(folder, "capex_output.csv")
    fieldnames = ["file", "tax_id", "company_name", "site_sheet", "site_name",
                  "system", "component", "cost_ils", "notes"]
    with open(out_csv, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    # Write warnings
    out_warn = os.path.join(folder, "capex_warnings.txt")
    with open(out_warn, "w", encoding="utf-8") as f:
        f.write("\n".join(all_warnings) if all_warnings else "No warnings.")

    print(f"\nDone.")
    print(f"  Output:   {out_csv}  ({len(all_rows)} line items)")
    print(f"  Warnings: {out_warn}  ({len(all_warnings)} issues)")
    if all_warnings:
        print("\nWarnings preview:")
        for w in all_warnings[:10]:
            print(f"  {w}")


if __name__ == "__main__":
    main()
