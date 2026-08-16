#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tax Incentive Model Generator -- v4 (2026-08-16)
National Energy Efficiency Program -- EcoTraders

Rebuilds the live workbook (מודל_פחת_מואץ_0_1.xlsx) with 3 technology
blocks instead of 6 -- one averaged capacity point per technology
(Omri's call, plain midpoint) -- and folds in the "morning" fixes from
2026-08-16-morning-model-fixes.md:

  1. Capacity collapse: heat pump 40/70kW -> 55kW, chillers 100/500RT ->
     300RT, VSD 45/150kW -> 97.5kW. Rates that were already flat across
     the two old capacity points (CapEx/ton, CapEx/kW, hours, VSD
     specific power) carry over unchanged; only capacity and any
     genuinely capacity-linked efficiency figure (chiller kW/ton, heat
     pump COP) were averaged.
  2. Heat pump baseline swapped from mazut/diesel oven to a
     standard-efficiency heat pump (COP 3.3, flat across the old
     capacity points) -- same structure as chillers now (baseline vs.
     efficient tier of the same technology, not a fuel-switching
     comparison). Baseline CapEx not directly sourced; derived the same
     way chiller/VSD baseline CapEx already are in this model (efficient
     CapEx / (1 + premium)), 20% midpoint of the 15-25% illustrative
     range -> 1,050 / 1.20 = 875 ILS/kW, flagged peach ("estimated/
     derived") not green ("real source"). The old fuel-price/combustion
     data (rows 14-18) stays as reference context, no longer feeds the
     live calculation chain.
  3. OPEX-אחר hardcoded to 0 in every block (was a broken #REF! pointing
     at a deleted assumptions-sheet row) -- matches the standing decision
     that the maintenance delta is 0 for all 3 technologies.
  4. Efficient-equipment OPEX degradation sign fixed: (1+degr)^(i-1), not
     (1-degr)^(i-1) -- degrading equipment costs more over time, not less.
  5. Depreciation schedule always sums to exactly 100%: the number of
     accelerated-depreciation years is still ROUND(1/(rate*mult),0), but
     the LAST year absorbs the rounding residual instead of every year
     getting an identical flat share.
  6. Payback-B and payback-C array formulas (LET+SEQUENCE+MMULT,
     interpolated crossover) are generated for every block, every time --
     no more empty payback-C cells in any technology.
  7. TAOZ tariff table: winter-peak (חורף/פסגה) now uses the same 5/7
     weekday factor as the other two seasons (was 7/7), and the
     methodology note's payback-threshold default text now reads 3
     years, matching the live F21 parameter (was stuck at "2.5").

NOT changed (raised with Daniel/Rafi separately, not a formula fix):
  - Asymmetric equipment lifespans within the flat 20yr window, no
    replacement/salvage value (flag #2).
  - VSD savings % from full-load specific power vs. full-load baseline
    consumption -- doesn't represent load-following (flag #6).
  - Electricity tariff is hour-weighted only, not technology-run-time
    weighted (flag #7) -- conservative specifically for chillers.
  - Fiscal cost discounted at the firm's 6% rate, not a separate
    government rate (flag #11).
  - Heat pump hours (5,475) unchanged -- carried over as-is (flag #12).

MWh-saved / tCO2-saved output rows (flag #10) are NOT in this version --
that's afternoon work, scoped separately once this rebuild is confirmed.
Sensitivity data tables (native Excel What-If tables) also are NOT
reproduced here -- they need to be rebuilt in Excel itself (Data > What-If
Analysis > Data Table) against the new single-capacity-point blocks;
faking them as static formulas would misrepresent them as live sensitivity
when they would not be.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

# ─── COLORS (matched to the live workbook's own legend, not invented) ────────
F_INPUT   = PatternFill("solid", fgColor="FFFF00")   # yellow  -- input / pending
F_CONTROL = PatternFill("solid", fgColor="FFC000")   # orange  -- value to verify/update
F_SOURCED = PatternFill("solid", fgColor="C6E0B4")   # green   -- real sourced data
F_ESTIM   = PatternFill("solid", fgColor="FCE4D6")   # peach   -- estimated/derived (with source)
F_HEADING = PatternFill("solid", fgColor="1F4E79")   # dark blue -- block header
F_SUBHEAD = PatternFill("solid", fgColor="BDD7EE")   # l.blue  -- column headers
F_RESULT  = PatternFill("solid", fgColor="538135")   # green   -- Scenario A label
F_SCN_B   = PatternFill("solid", fgColor="2E75B6")   # blue    -- Scenario B label
F_SCN_C   = PatternFill("solid", fgColor="7B6000")   # gold    -- Scenario C label
F_ROW_A   = PatternFill("solid", fgColor="E2EFDA")   # l.green -- Scenario A rows
F_ROW_B   = PatternFill("solid", fgColor="DDEBF7")   # l.blue  -- Scenario B rows
F_ROW_C   = PatternFill("solid", fgColor="FFF2CC")   # cream   -- Scenario C rows
F_POLICY  = PatternFill("solid", fgColor="FF6B6B")   # red     -- key policy row

FONT_H  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
FONT_SH = Font(name="Arial", bold=True, color="000000", size=10)
FONT_N  = Font(name="Arial", size=10)
FONT_SM = Font(name="Arial", italic=True, size=9, color="595959")
FONT_B  = Font(name="Arial", bold=True, size=10)

A_R  = Alignment(horizontal="right",  vertical="center", readingOrder=2)
A_C  = Alignment(horizontal="center", vertical="center", readingOrder=2)
A_RW = Alignment(horizontal="right",  vertical="center", readingOrder=2, wrap_text=True)

FMT_NIS   = '#,##0" ₪"'
FMT_PCT   = '0.0%'
FMT_NUM   = '#,##0'
FMT_YR    = '0.0'
FMT_X     = '0.0"×"'
FMT_KW    = '#,##0" kW"'
FMT_RT    = '#,##0" RT"'
FMT_KWTON = '0.00" kW/ton"'
FMT_KWCFM = '0.0" kW/100cfm"'
FMT_COP   = '0.00'
FMT_PRICE3 = '#,##0.000'

GLOBAL_SHEET   = "נתונים והנחות"
ANALYSIS_SHEET = "ניתוח"

TAX_REF    = f"'{GLOBAL_SHEET}'!$F$10"
DISC_REF   = f"'{GLOBAL_SHEET}'!$F$11"
ELEC_REF   = f"'{GLOBAL_SHEET}'!$F$12"
MULT_REF   = f"'{GLOBAL_SHEET}'!$F$20"
THRESH_REF = f"'{GLOBAL_SHEET}'!$F$21"

# ─── Global-sheet tech-table row numbers (3-column layout) ────────────────────
R_CAP   = 29
R_HRS   = 30
R_COPE  = 31   # heat pump only: COP, efficient
R_COPB  = 32   # heat pump only: COP, baseline (standard-efficiency HP) -- NEW, replaces combustion-eff/fuel-type
R_EFFB  = 34   # chiller/VSD only: native-unit efficiency, baseline
R_EFFE  = 35   # chiller/VSD only: native-unit efficiency, efficient
R_SAV   = 36   # computed savings %
R_CONSB = 37   # computed annual consumption, baseline (always kWh now -- heat pump baseline is electric too)
R_CONSE = 38   # computed annual consumption, efficient (always kWh)
R_CAPXB = 40
R_CAPXE = 41
R_DCAPX = 42
R_LIFEB = 43
R_LIFEE = 44
R_DEGR  = 45
R_OPXEB = 46   # computed OPEX year 1, baseline
R_OPXEE = 47   # computed OPEX year 1, efficient

T_START_COL = 6   # column F = first tech column

TECHS = [
    dict(kind="heat_pump", num="3.1", name_eff="משאבות חום — 55 kW",
         name_base="משאבת חום סטנדרטית — 55 kW",
         cap=55, cap_fmt=FMT_KW, hours=5475,
         cop_e=3.68, cop_b=3.3,
         capex_b=48125, capex_e=57750, capex_b_fill=F_ESTIM, capex_e_fill=F_ESTIM,
         life_b=15, life_e=10, degr=0.005),
    dict(kind="chiller", num="3.2", name_eff="צ'ילרים — 300 RT",
         name_base="מערכת קירור קונבנציונלית — 300 RT",
         cap=300, cap_fmt=FMT_RT, hours=3000,
         eff_b=0.775, eff_e=0.64, eff_fmt=FMT_KWTON,
         capex_b=1068600, capex_e=1255800, capex_b_fill=F_ESTIM, capex_e_fill=F_ESTIM,
         life_b=15, life_e=17, degr=0.005),
    dict(kind="vsd", num="3.3", name_eff="מדחסי VSD — 97.5 kW",
         name_base="מדחס מהירות קבועה — 97.5 kW",
         cap=97.5, cap_fmt=FMT_KW, hours=5000,
         eff_b=21.5, eff_e=16.5, eff_fmt=FMT_KWCFM,
         capex_b=119340, capex_e=146250, capex_b_fill=F_ESTIM, capex_e_fill=F_ESTIM,
         life_b=12, life_e=12, degr=0),
]

STD_DEPR_PCT = 0.10
MAX_YRS = 20
YR0_COL = 6
MAX_COL_IDX = YR0_COL + MAX_YRS
MAX_COL_LTR = get_column_letter(MAX_COL_IDX)
YR0_LTR = get_column_letter(YR0_COL)


def sc(ws, row, col, value=None, fill=None, font=None, align=None, fmt=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if fill:  c.fill = fill
    if font:  c.font = font
    if align: c.alignment = align
    if fmt:   c.number_format = fmt
    return c


def section_hdr(ws, row, label, number=None, col_start=2, col_end=8):
    text = f"{number}. {label}" if number else label
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start)
    c.value, c.fill, c.font, c.alignment = text, F_HEADING, FONT_H, A_R


def year_header_row(ws, row, yr0_label="שנה 0"):
    for col, val in [(2, "פרמטר"), (3, "יחידות"), (4, "מקור"), (5, "הערות"), (6, yr0_label)]:
        sc(ws, row, col, val, fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    for i in range(1, MAX_YRS + 1):
        sc(ws, row, 6 + i, f"שנה {i}", fill=F_SUBHEAD, font=FONT_SH, align=A_C)


def scenario_label_row(ws, row, label, fill, col_end=None):
    end = col_end or (6 + MAX_YRS)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=end)
    c = ws.cell(row=row, column=2)
    c.value, c.fill = label, fill
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.alignment = A_R


def set_col_widths(ws, include_year_cols=False, n_tech=0):
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 46
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 34
    ws.column_dimensions['F'].width = 17
    for i in range(1, n_tech + 1):
        ws.column_dimensions[get_column_letter(6 + i)].width = 17
    if include_year_cols:
        for i in range(1, MAX_YRS + 1):
            ws.column_dimensions[get_column_letter(6 + i)].width = 12


def color_legend(ws, row=1):
    ws.cell(row=row, column=2).value = "מקרא צבעים"
    ws.cell(row=row, column=2).font = FONT_SH
    items = [
        ("ערך להזנה / ממתין לנתונים", F_INPUT),
        ("נתון לעדכון / בקרה", F_CONTROL),
        ("נתון ממקור אמיתי", F_SOURCED),
        ("נתון מוערך/נגזר (עם מקור)", F_ESTIM),
    ]
    for i, (label, fill) in enumerate(items):
        sc(ws, row + 1 + i, 2, label, fill=fill, font=FONT_N, align=A_R)


def col_of(i):
    return get_column_letter(T_START_COL + i)


def payback_array_formula(cell_ref, disc_row_scenario, disc_row_baseline):
    """LET+SEQUENCE+MMULT interpolated payback formula, same pattern as the
    live workbook (ported verbatim, just re-pointed at this block's rows)."""
    diffs = f"F{disc_row_scenario}:Z{disc_row_scenario}-F{disc_row_baseline}:Z{disc_row_baseline}"
    formula = (
        f"=_xlfn.LET(_xlpm.diffs, {diffs}, "
        f"_xlpm.n, COLUMNS(_xlpm.diffs), "
        f"_xlpm.cum, MMULT(_xlpm.diffs, N(_xlfn.SEQUENCE(_xlpm.n,1)<=_xlfn.SEQUENCE(1,_xlpm.n))), "
        f"_xlpm.idx, MATCH(TRUE, _xlpm.cum>=0, 0), "
        f'IF(ISNA(_xlpm.idx), "לא מוחזר בטווח", '
        f"IF(_xlpm.idx=1, 0, (_xlpm.idx-2) + (-INDEX(_xlpm.cum,_xlpm.idx-1)/INDEX(_xlpm.diffs,_xlpm.idx)))))"
    )
    return ArrayFormula(cell_ref, formula)


# ─── SHEET 1: GLOBAL ASSUMPTIONS ──────────────────────────────────────────────

def build_global_sheet(wb):
    ws = wb.active
    ws.title = GLOBAL_SHEET
    ws.sheet_view.rightToLeft = True
    set_col_widths(ws, n_tech=len(TECHS))
    color_legend(ws, row=1)

    section_hdr(ws, 9, "פרמטרים פיננסיים", number=1, col_end=6 + len(TECHS))
    sc(ws, 10, 2, "שיעור מס חברות", fill=F_CONTROL, font=FONT_N, align=A_R)
    sc(ws, 10, 3, "%", align=A_R)
    sc(ws, 10, 4, "רשות המסים", align=A_R)
    sc(ws, 10, 6, 0.23, fill=F_CONTROL, fmt=FMT_PCT, align=A_C)

    sc(ws, 11, 2, "שיעור היוון (יזמי / פרטי)", fill=F_SOURCED, font=FONT_N, align=A_R)
    sc(ws, 11, 3, "%", align=A_R)
    sc(ws, 11, 4, "סוכם עם דניאל", align=A_R)
    sc(ws, 11, 6, 0.06, fill=F_SOURCED, fmt=FMT_PCT, align=A_C)

    sc(ws, 12, 2, 'מחיר חשמל ממוצע לתעשייה', fill=F_SOURCED, font=FONT_N, align=A_R)
    sc(ws, 12, 3, 'אג\'/קוט"ש', align=A_R)
    sc(ws, 12, 4, 'חברת החשמל - מחירי התעו"ז במתח גבוה', align=A_R)
    sc(ws, 12, 5, 'ממוצע משוקלל לפי שעות תעו"ז; בתוקף מ-01.07.2026', font=FONT_SM, align=A_RW)
    # F12's formula is set later, once the TAOZ section's real row (R_AVG) is known.

    section_hdr(ws, 13, 'מחירי דלקים ומקדמי המרה (הקשר בלבד -- לא בשימוש בחישוב הפעיל; הבסיס למשאבות חום הוא כעת חשמלי, ראו סעיף 4)',
                number="1ב", col_end=6 + len(TECHS))
    fuel = [
        (14, "מחיר סולר בתעשייה (בתוספת מרווח שיווק)", "₪/ליטר",
             "משרד האנרגיה — מחירים תיאורטיים 2024", 2.59368, FMT_PRICE3),
        (15, "מחיר מזוט בתעשייה (בתוספת מרווח שיווק)", "₪/טון",
             "משרד האנרגיה — מחירים תיאורטיים 2024", 2344.72, FMT_NIS),
        (16, "יחס קלורי — סולר", "טון דלק/MWh", "MRV", 0.085, FMT_PRICE3),
        (17, "יחס קלורי — מזוט", "טון דלק/MWh", "MRV", 0.088, FMT_PRICE3),
        (18, "צפיפות סולר", 'ק"ג/ליטר', "MRV", 0.82, FMT_PRICE3),
    ]
    for r, lbl, u, src, val, fmt in fuel:
        sc(ws, r, 2, lbl, font=FONT_SM, align=A_R)
        sc(ws, r, 3, u, align=A_R)
        sc(ws, r, 4, src, align=A_RW)
        sc(ws, r, 6, val, fmt=fmt, align=A_C)

    section_hdr(ws, 19, "פרמטר תמריץ — מכפיל פחת (פחת מואץ)", number=2, col_end=6 + len(TECHS))
    sc(ws, 20, 2, "מכפיל שיעור הפחת  ←  פרמטר המדיניות המרכזי", fill=F_POLICY,
       font=Font(name="Arial", bold=True, size=11, color="FFFFFF"), align=A_R)
    sc(ws, 20, 3, "מכפיל", fill=F_POLICY, font=Font(name="Arial", bold=True, size=11, color="FFFFFF"), align=A_C)
    sc(ws, 20, 5, "1.0 = ללא תמריץ | 2.0 = כפול, 5 שנים | 5.0 = 2 שנים",
       fill=F_POLICY, font=Font(name="Arial", italic=True, size=9, color="FFFFFF"), align=A_R)
    sc(ws, 20, 6, 2, fill=F_INPUT, font=Font(name="Arial", bold=True, size=14, color="C00000"), fmt=FMT_X, align=A_C)

    sc(ws, 21, 2, "סף החזר השקעה (שנים) — קריטריון אימוץ", fill=F_INPUT, font=FONT_N, align=A_R)
    sc(ws, 21, 3, "שנים", align=A_R)
    sc(ws, 21, 4, "פרמטר ניתן לשינוי", align=A_R)
    sc(ws, 21, 5, "סף אימוץ עסקי — תקופת החזר על תזרים מהוון עודף (יעיל מול בסיסי)", font=FONT_SM, align=A_RW)
    sc(ws, 21, 6, 3, fill=F_INPUT, fmt=FMT_YR, align=A_C)

    section_hdr(ws, 22, "שיעורי פחת סטנדרטיים — תקנות מס הכנסה", number=3, col_end=6 + len(TECHS))
    for r, lbl in [(23, "משאבות חום"), (24, "צ'ילרים"), (25, "מדחסי VSD")]:
        sc(ws, r, 2, lbl, font=FONT_N, align=A_R)
        sc(ws, r, 3, "% לשנה", align=A_R)
        sc(ws, r, 4, "תקנות פחת — ציוד מכני", align=A_R)
        sc(ws, r, 6, 0.10, fmt=FMT_PCT, align=A_C)

    T_END_COL = T_START_COL + len(TECHS) - 1
    section_hdr(ws, 26, "הנחות לפי טכנולוגיה — נקודת קיבולת אחת ממוצעת לכל טכנולוגיה (עודכן 2026-08-16, ראו decisions/log.md)",
                number=4, col_end=T_END_COL)

    for col, lbl in [(2, "פרמטר"), (3, "יחידות"), (4, "מקור"), (5, "הערות")]:
        sc(ws, 27, col, lbl, fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    for i, t in enumerate(TECHS):
        sc(ws, 27, T_START_COL + i, t['name_eff'],
           fill=PatternFill("solid", fgColor="1F4E79"),
           font=Font(name="Arial", bold=True, color="FFFFFF", size=9), align=A_C)
        sc(ws, 28, T_START_COL + i, f"vs. {t['name_base']}", font=FONT_SM, align=A_C)

    rowmeta = [
        (R_CAP,  "קיבולת מותקנת (ממוצע משתי נקודות הקיבולת הקודמות)", "kW / RT",
         "2026-08-16-morning-model-fixes.md", ""),
        (R_HRS,  "שעות פעילות שנתיות", "שעות/שנה", "ראו הערות לפי טכנולוגיה", ""),
        (R_COPE, "COP — משאבת חום (יעיל)", "יחס", "Sprsun — ממוצע 2 המוצרים", "משאבות חום בלבד"),
        (R_COPB, "COP — משאבת חום (בסיסי, תקן מינימלי)", "יחס",
         "DOE FEMP / ASHRAE 90.1-2019", "משאבות חום בלבד — עודכן 2026-08-16, הבסיס כעת משאבת חום סטנדרטית, לא תנור"),
        (R_EFFB, "יעילות אנרגטית — בסיסי", "kW/ton | kW/100cfm",
         "צ'ילר: ASHRAE 90.1 | VSD: CAGI", "צ'ילרים/VSD בלבד"),
        (R_EFFE, "יעילות אנרגטית — יעיל", "kW/ton | kW/100cfm",
         "צ'ילר: DOE FEMP | VSD: CAGI", "צ'ילרים/VSD בלבד"),
        (R_SAV,  "חיסכון אנרגטי (מחושב)", "% מצריכה", "חישוב", "כל הטכנולוגיות"),
        (R_CONSB, "צריכת חשמל שנתית — בסיסי", 'קוט"ש/שנה', "חישוב", "הצד הבסיסי כעת תמיד חשמלי"),
        (R_CONSE, "צריכת חשמל שנתית — יעיל", 'קוט"ש/שנה', "חישוב", "הצד היעיל תמיד חשמלי"),
        (R_CAPXB, "CapEx — ציוד בסיסי", "₪", "ראו הערות לפי טכנולוגיה", ""),
        (R_CAPXE, "CapEx — ציוד יעיל", "₪", "ראו הערות לפי טכנולוגיה", ""),
        (R_DCAPX, "ΔCapEx = יעיל − בסיסי", "₪", "חישוב", ""),
        (R_LIFEB, "אורך חיים — ציוד בסיסי", "שנים", "ממצאי הנדסה / הנחת עבודה", ""),
        (R_LIFEE, "אורך חיים — ציוד יעיל", "שנים", "ממצאי הנדסה / הנחת עבודה", ""),
        (R_DEGR,  "גורם שחיקת ביצועים", "%/שנה", "רפי", ""),
        (R_OPXEB, "OPEX אנרגיה שנה 1 — בסיסי", "₪/שנה", "חישוב", ""),
        (R_OPXEE, "OPEX אנרגיה שנה 1 — יעיל", "₪/שנה", "חישוב", ""),
    ]
    for row, param, units, src, note in rowmeta:
        sc(ws, row, 2, param, font=FONT_N, align=A_R)
        sc(ws, row, 3, units, align=A_R)
        sc(ws, row, 4, src, align=A_RW)
        if note:
            sc(ws, row, 5, note, font=FONT_SM, align=A_RW)

    sc(ws, R_HRS, 5, 'משאבות חום: 5,475 (רפי) | צ\'ילרים: 3,000 (מספר עבודה) | VSD: 5,000 (רפי)',
       font=FONT_SM, align=A_RW)

    for i, t in enumerate(TECHS):
        X = col_of(i)
        col = T_START_COL + i

        sc(ws, R_CAP, col, t['cap'], fill=F_CONTROL, fmt=t['cap_fmt'], align=A_C)
        sc(ws, R_HRS, col, t['hours'], fill=F_CONTROL, fmt=FMT_NUM, align=A_C)

        if t['kind'] == "heat_pump":
            sc(ws, R_COPE, col, t['cop_e'], fill=F_CONTROL, fmt=FMT_COP, align=A_C)
            sc(ws, R_COPB, col, t['cop_b'], fill=F_SOURCED, fmt=FMT_COP, align=A_C)
            sc(ws, R_SAV, col,
               f'=IF(AND(ISNUMBER({X}{R_COPB}),ISNUMBER({X}{R_COPE})),'
               f'({X}{R_COPE}-{X}{R_COPB})/{X}{R_COPE},"PENDING")', fmt=FMT_PCT, align=A_C)
            sc(ws, R_CONSB, col,
               f'=IF(AND(ISNUMBER({X}{R_CAP}),ISNUMBER({X}{R_HRS}),ISNUMBER({X}{R_COPB})),'
               f'{X}{R_CAP}*{X}{R_HRS}/{X}{R_COPB},"PENDING")', fmt=FMT_NUM, align=A_C)
            sc(ws, R_CONSE, col,
               f'=IF(AND(ISNUMBER({X}{R_CAP}),ISNUMBER({X}{R_HRS}),ISNUMBER({X}{R_COPE})),'
               f'{X}{R_CAP}*{X}{R_HRS}/{X}{R_COPE},"PENDING")', fmt=FMT_NUM, align=A_C)
        else:
            sc(ws, R_EFFB, col, t['eff_b'], fill=F_CONTROL, fmt=t['eff_fmt'], align=A_C)
            sc(ws, R_EFFE, col, t['eff_e'], fill=F_CONTROL, fmt=t['eff_fmt'], align=A_C)
            sc(ws, R_SAV, col,
               f'=IF(AND(ISNUMBER({X}{R_EFFB}),ISNUMBER({X}{R_EFFE})),'
               f'({X}{R_EFFB}-{X}{R_EFFE})/{X}{R_EFFB},"PENDING")', fmt=FMT_PCT, align=A_C)
            if t['kind'] == "chiller":
                sc(ws, R_CONSB, col,
                   f'=IF(AND(ISNUMBER({X}{R_CAP}),ISNUMBER({X}{R_HRS}),ISNUMBER({X}{R_EFFB})),'
                   f'{X}{R_CAP}*{X}{R_EFFB}*{X}{R_HRS},"PENDING")', fmt=FMT_NUM, align=A_C)
            else:
                sc(ws, R_CONSB, col,
                   f'=IF(AND(ISNUMBER({X}{R_CAP}),ISNUMBER({X}{R_HRS})),'
                   f'{X}{R_CAP}*{X}{R_HRS},"PENDING")', fmt=FMT_NUM, align=A_C)
            sc(ws, R_CONSE, col,
               f'=IF(AND(ISNUMBER({X}{R_CONSB}),ISNUMBER({X}{R_SAV})),'
               f'{X}{R_CONSB}*(1-{X}{R_SAV}),"PENDING")', fmt=FMT_NUM, align=A_C)

        sc(ws, R_OPXEB, col, f'=IF(ISNUMBER({X}{R_CONSB}),{X}{R_CONSB}*{ELEC_REF}/100,"PENDING")',
           fmt=FMT_NIS, align=A_C)
        sc(ws, R_OPXEE, col, f'=IF(ISNUMBER({X}{R_CONSE}),{X}{R_CONSE}*{ELEC_REF}/100,"PENDING")',
           fmt=FMT_NIS, align=A_C)

        sc(ws, R_CAPXB, col, t['capex_b'], fill=t['capex_b_fill'], fmt=FMT_NIS, align=A_C)
        sc(ws, R_CAPXE, col, t['capex_e'], fill=t['capex_e_fill'], fmt=FMT_NIS, align=A_C)
        sc(ws, R_DCAPX, col,
           f'=IF(ISNUMBER({X}{R_CAPXE})*ISNUMBER({X}{R_CAPXB}),{X}{R_CAPXE}-{X}{R_CAPXB},"PENDING")',
           fmt=FMT_NIS, align=A_C)

        sc(ws, R_LIFEB, col, t['life_b'], fill=F_CONTROL, fmt=FMT_YR, align=A_C)
        sc(ws, R_LIFEE, col, t['life_e'], fill=F_CONTROL, fmt=FMT_YR, align=A_C)
        sc(ws, R_DEGR, col, t['degr'], fill=F_SOURCED if t['degr'] == 0 else F_INPUT, fmt=FMT_PCT, align=A_C)

    sc(ws, R_CAPXE, 5,
       "משאבות חום: מענקים ₪1,050/kW | צ'ילר: מענקים חציון ₪4,186/טון | VSD: מענקים חציון ₪1,500/kW "
       "(כל השיעורים כבר היו קבועים על פני שתי נקודות הקיבולת הקודמות — לא השתנו בקונסולידציה)",
       font=FONT_SM, align=A_RW)
    sc(ws, R_CAPXB, 5,
       "צ'ילר: יעיל ÷ 1.175 (פרמיית יעילות 10-25%) | VSD: יעיל ÷ 1.225 (פרמיית VSD 15-30%) | "
       "משאבות חום: יעיל ÷ 1.20 (פרמיה מוערכת 15-25%, נגזרת באנלוגיה לצ'ילר/VSD -- אינה מקור ייעודי "
       "למשאבות חום, ממתינה לאישור/מספר טוב יותר מרפי)",
       font=FONT_SM, align=A_RW)

    R = 48
    section_hdr(ws, R, "מתודולוגיה — גזירות והבהרות", number=5, col_end=6 + len(TECHS))
    R += 1
    notes = [
        "קונסולידציה לנקודת קיבולת אחת (2026-08-16) — כל טכנולוגיה עברה משתי נקודות קיבולת לנקודה "
        "ממוצעת אחת (ממוצע חשבוני פשוט, ללא משקלול). שיעורים שכבר היו קבועים על פני שתי הנקודות "
        "(CapEx ל-₪/ton או ₪/kW, שעות פעילות, הספק הסגולי של VSD) נשארו כפי שהם; רק ערך הקיבולת עצמו "
        "וכל מדד יעילות שכן היה תלוי-קיבולת (kW/ton של צ'ילר, COP של משאבת חום) מוצעו. תופעת לוואי "
        "מקובלת: זה מטשטש את ההבדל האמיתי בין מדחס בוכני למרכזי בצ'ילרים בין שתי נקודות הקיבולת "
        "הישנות. ראו decisions/log.md 2026-08-16.",
        "בסיס משאבות חום שונה ממנוע בעירה (תנור מזוט/סולר) למשאבת חום סטנדרטית (2026-08-05, אושר ע\"י "
        "דניאל) — אותו מבנה כמו צ'ילרים (בסיסי מול יעיל של אותה טכנולוגיה), לא השוואת מעבר דלק. "
        "COP בסיסי = 3.3 (מינימום תקן DOE FEMP / ASHRAE 90.1-2019 ל-47°F). נתוני הדלק/הבעירה בסעיף 1ב "
        "נשארים כהקשר בלבד, אינם מוזנים עוד לחישוב הפעיל.",
        "CapEx צ'ילר בסיסי — נגזר מהיעיל: נתוני המענקים כוללים רק את הציוד היעיל שמומן, ולכן אין מחיר "
        "לציוד בסיסי. הבסיסי נאמד כ- ₪4,186 ÷ 1.175 = ₪3,562/טון, לפי פרמיית יעילות של 10-25% "
        "(DOE FEMP + מקורות שוק, אמצע 17.5%). הערכה, לא ציטוט ישיר.",
        "CapEx VSD בסיסי — נגזר בדומה: ₪1,500 ÷ 1.225 = ₪1,224/kW, לפי פרמיית VSD מול מהירות קבועה "
        "של 15-30% (אמצע 22.5%).",
        "CapEx משאבת חום בסיסי — נגזר באנלוגיה לצ'ילר/VSD (2026-08-16): ₪1,050 ÷ 1.20 = ₪875/kW, "
        "לפי פרמיית יעילות מוערכת 15-25% (אמצע 20%). זו אנלוגיה מטכנולוגיה אחרת, לא מקור ייעודי "
        "למשאבות חום — סומן מוערך/נגזר, לא מקור אמיתי. ממתין למספר טוב יותר מרפי/דניאל.",
        "OPEX אחר (תחזוקה) — נתוני רפי: ההפרש מוגדר 0 לכל שלוש הטכנולוגיות (צ'ילר יעיל=רגיל, "
        "VSD יעיל=רגיל, משאבת חום זולה יותר מתנור/יעיל). מוזן ישירות כ-0 בגיליון הניתוח, לא כתא נפרד "
        "כאן — התא הישן שאליו הפנו הנוסחאות בגיליון הניתוח נמחק בעבר וגרם ל-#REF!, תוקן 2026-08-16.",
        "שחיקת ביצועים — רפי: 0.5%/שנה למשאבות חום וצ'ילרים, 0% למדחסי VSD בורגיים (נצילות נשארת "
        "קבועה עד תקלה). הכיוון בנוסחה תוקן 2026-08-16: העלות עולה עם השחיקה (1+שיעור)^שנה, לא יורדת.",
        "תקופת החזר השקעה — מחושבת על תזרים מזומנים מהוון עודף (תרחיש יעיל מול בסיסי), עם אינטרפולציה "
        "לינארית לשנה חלקית. סף האימוץ (F21, ברירת מחדל 3 שנים) משקף משוכות אישור השקעה קצרות־טווח "
        "כפי שפירמות מפעילות בפועל, ולא את תקן ה-NPV ל-20 שנה.",
        "פחת מואץ — לוח הפחת מחולק כך שסך הפחת המואץ תמיד שווה בדיוק ל-100% מה-CapEx: השנה האחרונה "
        "בטווח סופגת את שארית העיגול, במקום שכל שנה תקבל חלק שטוח זהה (עלול היה לסכם ל-90%-105% "
        "בהתאם למכפיל). תוקן 2026-08-16.",
    ]
    for txt in notes:
        ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=6 + len(TECHS))
        sc(ws, R, 2, "• " + txt, font=FONT_SM, align=A_RW)
        ws.row_dimensions[R].height = 42
        R += 1

    # ── TAOZ tariff calculation (ported from the live file, one fix: G151) ──
    R += 1
    sc(ws, R, 2, 'חישוב תעריף חשמל ממוצע — תעו"ז מתח גבוה (כלל צרכנות חח"י)', font=FONT_B, align=A_R)
    R += 1
    sc(ws, R, 2, "לוח תעריפים: נספח ה' — לוחות תעריפים החל מיום ה-1.1.2026, לוח 5.2-1 \"תעו\"ז לפי רמות מתח\", רשות החשמל",
       font=FONT_SM, align=A_R)
    R += 2
    R_TAOZ_HDR = R
    for col, lbl in [(2, "עונה"), (3, 'מש"ב'), (4, 'תעריף\n(אג\' לקוט"ש)'), (5, "ימים בעונה"),
                     (6, 'שעות פסגה\nליום רלוונטי'), (7, "מכפל ימים\nבשבוע"), (8, 'סה"כ שעות\nבשנה')]:
        sc(ws, R, col, lbl, fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    R += 1
    R_TAOZ_START = R
    taoz_rows = [
        ("חורף (דצמבר–פברואר)", "שפל", 35.08, 90, None, None),
        ("חורף (דצמבר–פברואר)", "פסגה", 101.21, None, 5, "=5/7"),
        ("מעבר (מרץ–יוני, אוקטובר–נובמבר)", "שפל", 34.26, 183, None, None),
        ("מעבר (מרץ–יוני, אוקטובר–נובמבר)", "פסגה", 37.96, None, 5, "=5/7"),
        ("קיץ (יולי–ספטמבר)", "שפל", 38.77, 92, None, None),
        ("קיץ (יולי–ספטמבר)", "פסגה", 157.09, None, 6, "=5/7"),
    ]
    for idx, (season, band, rate, days, peak_hrs, day_mult) in enumerate(taoz_rows):
        rr = R + idx
        sc(ws, rr, 2, season, align=A_R)
        sc(ws, rr, 3, band, align=A_C)
        sc(ws, rr, 4, rate, fmt='0.00', align=A_C)
        if days is not None:
            sc(ws, rr, 5, days, fmt=FMT_NUM, align=A_C)
            sc(ws, rr, 8, f"=E{rr}*24-H{rr+1}", fmt='0.00', align=A_C)
        else:
            sc(ws, rr, 5, f"=E{rr-1}", fmt=FMT_NUM, align=A_C)
            sc(ws, rr, 6, peak_hrs, fmt=FMT_NUM, align=A_C)
            sc(ws, rr, 7, day_mult, fmt='0.000', align=A_C)
            sc(ws, rr, 8, f"=E{rr}*G{rr}*F{rr}", fmt='0.00', align=A_C)
    R_TAOZ_END = R + len(taoz_rows) - 1
    R_TOT = R_TAOZ_END + 1
    sc(ws, R_TOT, 2, 'סה"כ שעות בשנה', font=FONT_B, align=A_R)
    sc(ws, R_TOT, 8, f"=SUM(H{R_TAOZ_START}:H{R_TAOZ_END})", fmt=FMT_NUM, align=A_C)
    R_AVG = R_TOT + 2
    sc(ws, R_AVG, 2, 'תעריף ממוצע משוקלל לפי שעות', font=FONT_B, align=A_R)
    sc(ws, R_AVG, 3, 'אג\' לקוט"ש', align=A_R)
    sc(ws, R_AVG, 4, 'מחושב: SUMPRODUCT(תעריף × שעות) / סה"כ שעות', align=A_R)
    sc(ws, R_AVG, 6, f"=SUMPRODUCT(D{R_TAOZ_START}:D{R_TAOZ_END},H{R_TAOZ_START}:H{R_TAOZ_END})/H{R_TOT}",
       fmt='#,##0.00', align=A_C, fill=F_SOURCED)

    sc(ws, 12, 6, f"=F{R_AVG}", fill=F_SOURCED, fmt='#,##0.00', align=A_C)

    tech_refs = []
    for i, t in enumerate(TECHS):
        X = col_of(i)
        s = GLOBAL_SHEET
        tech_refs.append(dict(
            capex_b=f"'{s}'!${X}${R_CAPXB}", capex_e=f"'{s}'!${X}${R_CAPXE}",
            opex_e_b=f"'{s}'!${X}${R_OPXEB}", opex_e_e=f"'{s}'!${X}${R_OPXEE}",
            degr=t['degr'],
        ))
    return tech_refs


# ─── SHEET 2: ANALYSIS ────────────────────────────────────────────────────────

def build_analysis_sheet(wb, tech_refs):
    ws = wb.create_sheet(ANALYSIS_SHEET)
    ws.sheet_view.rightToLeft = True
    set_col_widths(ws, include_year_cols=True)
    color_legend(ws, row=1)

    tech_end_col = 6 + MAX_YRS
    ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=tech_end_col)
    sc(ws, 9, 2, "ניתוח כלכלי — פחת מואץ לפי טכנולוגיה | נקודת קיבולת ממוצעת אחת לטכנולוגיה",
       fill=F_HEADING, font=Font(name="Arial", bold=True, color="FFFFFF", size=13), align=A_R)

    section_hdr(ws, 11, "פרמטרים פיננסיים", number=1)
    for r, lbl, u, ref, fmt in [
        (12, "שיעור מס חברות", "%", f"={TAX_REF}", FMT_PCT),
        (13, "שיעור היוון", "%", f"={DISC_REF}", FMT_PCT),
        (14, 'מחיר חשמל (₪/קוט"ש)', '₪/קוט"ש', f"={ELEC_REF}/100", '#,##0.000'),
    ]:
        sc(ws, r, 2, lbl, font=FONT_N, align=A_R)
        sc(ws, r, 3, u, align=A_R)
        sc(ws, r, 4, GLOBAL_SHEET, align=A_R)
        sc(ws, r, 6, ref, fmt=fmt, align=A_C)

    section_hdr(ws, 16, "פרמטר תמריץ — מכפיל פחת", number=2)
    sc(ws, 17, 2, "מכפיל שיעור הפחת  ←  שנה ב'נתונים והנחות'!F20", fill=F_POLICY,
       font=Font(name="Arial", bold=True, size=11, color="FFFFFF"), align=A_R)
    sc(ws, 17, 6, f"={MULT_REF}", fill=F_INPUT,
       font=Font(name="Arial", bold=True, size=13, color="C00000"), fmt=FMT_X, align=A_C)

    section_hdr(ws, 19, "ניתוחים לפי טכנולוגיה", number=3, col_end=tech_end_col)

    R = 21
    tech_results = []
    for i, t in enumerate(TECHS):
        R, rmap = _tech_block(ws, t, R, tech_refs[i])
        tech_results.append((t, rmap))
        R += 2

    _summary_block(ws, tech_results, R)


def _tech_block(ws, t, R_start, refs):
    R = R_start
    tech_end_col = 6 + MAX_YRS
    life_e, life_b = t['life_e'], t['life_b']
    std_r = STD_DEPR_PCT
    degr_v = refs['degr']

    capex_b, capex_e = refs['capex_b'], refs['capex_e']
    opex_e_b, opex_e_e = refs['opex_e_b'], refs['opex_e_e']

    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, f"{t['num']}  {t['name_eff']}", fill=F_HEADING,
       font=Font(name="Arial", bold=True, color="FFFFFF", size=11), align=A_R)
    R += 1
    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, f"* מול {t['name_base']} | נתונים: גיליון '{GLOBAL_SHEET}' סעיף 4", font=FONT_SM, align=A_R)
    R += 1

    # ── ב: Depreciation ──
    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, "ב. לוחות פחת (ציוד בסיסי vs. יעיל)", fill=F_SUBHEAD, font=FONT_SH, align=A_R)
    R += 1
    year_header_row(ws, R, yr0_label="שנה 0 (השקעה)")
    R += 1

    R_SD_B = R; R += 1
    R_ST_B = R; R += 1
    R_SD_E = R; R += 1
    R_ST_E = R; R += 1
    R_AD_E = R; R += 1
    R_AT_E = R; R += 1

    for rr, lbl, fill in [
        (R_SD_B, f"פחת סטנדרטי — {t['name_base']}", F_ROW_A),
        (R_ST_B, f"מגן מס — {t['name_base']}", F_ROW_A),
        (R_SD_E, f"פחת סטנדרטי — {t['name_eff']}", F_ROW_B),
        (R_ST_E, f"מגן מס סטנדרטי — {t['name_eff']}", F_ROW_B),
        (R_AD_E, f"פחת מואץ — {t['name_eff']}", F_ROW_C),
        (R_AT_E, f"מגן מס מואץ — {t['name_eff']}", F_ROW_C),
    ]:
        sc(ws, rr, 2, lbl, fill=fill, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪", align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)
        sc(ws, rr, 6, 0, fill=fill, fmt=FMT_NIS, align=A_C)

    n_years_expr = f"ROUND(1/({std_r}*{MULT_REF}),0)"
    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl = get_column_letter(col)
        sd_b = f"=IF(ISNUMBER({capex_b}),{capex_b}*{std_r},0)" if i <= 10 else "=0"
        sc(ws, R_SD_B, col, sd_b, fill=F_ROW_A, fmt=FMT_NIS, align=A_C)
        sc(ws, R_ST_B, col, f"={cl}{R_SD_B}*{TAX_REF}", fill=F_ROW_A, fmt=FMT_NIS, align=A_C)
        sd_e = f"=IF(ISNUMBER({capex_e}),{capex_e}*{std_r},0)" if i <= 10 else "=0"
        sc(ws, R_SD_E, col, sd_e, fill=F_ROW_B, fmt=FMT_NIS, align=A_C)
        sc(ws, R_ST_E, col, f"={cl}{R_SD_E}*{TAX_REF}", fill=F_ROW_B, fmt=FMT_NIS, align=A_C)
        # Accelerated depreciation: last year in the schedule absorbs the rounding
        # residual so the schedule always sums to exactly 100% of CapEx (flag #5).
        ad_e = (
            f"=IF(ISNUMBER({capex_e}),"
            f"IF({i}<{n_years_expr},{capex_e}*{std_r}*{MULT_REF},"
            f"IF({i}={n_years_expr},{capex_e}-({capex_e}*{std_r}*{MULT_REF})*({n_years_expr}-1),0)),0)"
        )
        sc(ws, R_AD_E, col, ad_e, fill=F_ROW_C, fmt=FMT_NIS, align=A_C)
        sc(ws, R_AT_E, col, f"={cl}{R_AD_E}*{TAX_REF}", fill=F_ROW_C, fmt=FMT_NIS, align=A_C)
    R += 1

    # ── ג: Cash flow ──
    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, "ג. ניתוח תזרים מזומנים — שלושה תרחישים", fill=F_SUBHEAD, font=FONT_SH, align=A_R)
    R += 1
    year_header_row(ws, R, yr0_label="שנה 0 (השקעה)")
    R += 1

    scenario_label_row(ws, R, f"תרחיש A — רכישת {t['name_base']} (ציוד קיים)", F_RESULT)
    R += 1
    R_A_INV = R; R += 1
    R_A_EOPC = R; R += 1
    R_A_OOPC = R; R += 1
    R_A_NET = R; R += 1
    R_A_DISC = R; R += 1
    for rr, lbl in [(R_A_INV, "השקעה ראשונית"), (R_A_EOPC, "OPEX אנרגיה"),
                    (R_A_OOPC, "OPEX אחר"), (R_A_NET, "תזרים נקי — A"),
                    (R_A_DISC, "תזרים מהוון — A")]:
        sc(ws, rr, 2, lbl, fill=F_ROW_A, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪", align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)
    sc(ws, R_A_INV, 6, f'=IF(ISNUMBER({capex_b}),-{capex_b},"PENDING")', fill=F_ROW_A, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_EOPC, 6, 0, fill=F_ROW_A, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_OOPC, 6, 0, fill=F_ROW_A, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_NET, 6, f"=F{R_A_INV}", fill=F_ROW_A, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_DISC, 6, f"=F{R_A_NET}", fill=F_ROW_A, fmt=FMT_NIS, align=A_C)
    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl = get_column_letter(col)
        eopc = f"=IF(ISNUMBER({opex_e_b}),-{opex_e_b},0)" if i <= life_b else "=0"
        sc(ws, R_A_EOPC, col, eopc, fill=F_ROW_A, fmt=FMT_NIS, align=A_C)
        # OPEX-אחר (maintenance) confirmed 0 for all 3 technologies (Rafi, 2026-07-26) --
        # hardcoded, not a broken cross-sheet reference (flag #9).
        sc(ws, R_A_OOPC, col, 0, fill=F_ROW_A, fmt=FMT_NIS, align=A_C)
        sc(ws, R_A_NET, col, f"={cl}{R_A_EOPC}+{cl}{R_A_OOPC}+{cl}{R_ST_B}", fill=F_ROW_A, fmt=FMT_NIS, align=A_C)
        sc(ws, R_A_DISC, col, f"={cl}{R_A_NET}/(1+{DISC_REF})^{i}", fill=F_ROW_A, fmt=FMT_NIS, align=A_C)
    R += 1

    scenario_label_row(ws, R, f"תרחיש B — רכישת {t['name_eff']} | פחת סטנדרטי (ללא תמריץ)", F_SCN_B)
    R += 1
    R_B_INV = R; R += 1
    R_B_EOPC = R; R += 1
    R_B_OOPC = R; R += 1
    R_B_NET = R; R += 1
    R_B_DISC = R; R += 1
    for rr, lbl in [(R_B_INV, "השקעה ראשונית"), (R_B_EOPC, "OPEX אנרגיה (עם שחיקה)"),
                    (R_B_OOPC, "OPEX אחר"), (R_B_NET, "תזרים נקי — B"),
                    (R_B_DISC, "תזרים מהוון — B")]:
        sc(ws, rr, 2, lbl, fill=F_ROW_B, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪", align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)
    sc(ws, R_B_INV, 6, f'=IF(ISNUMBER({capex_e}),-{capex_e},"PENDING")', fill=F_ROW_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_EOPC, 6, 0, fill=F_ROW_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_OOPC, 6, 0, fill=F_ROW_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_NET, 6, f"=F{R_B_INV}", fill=F_ROW_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_DISC, 6, f"=F{R_B_NET}", fill=F_ROW_B, fmt=FMT_NIS, align=A_C)
    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl = get_column_letter(col)
        # Degradation sign fixed (flag #1): degrading equipment costs MORE over
        # time, so (1+degr)^(i-1), not (1-degr)^(i-1).
        eopc = (f"=IF(ISNUMBER({opex_e_e}),-{opex_e_e}*(1+{degr_v})^{i-1},0)" if i <= life_e else "=0")
        sc(ws, R_B_EOPC, col, eopc, fill=F_ROW_B, fmt=FMT_NIS, align=A_C)
        sc(ws, R_B_OOPC, col, 0, fill=F_ROW_B, fmt=FMT_NIS, align=A_C)
        sc(ws, R_B_NET, col, f"={cl}{R_B_EOPC}+{cl}{R_B_OOPC}+{cl}{R_ST_E}", fill=F_ROW_B, fmt=FMT_NIS, align=A_C)
        sc(ws, R_B_DISC, col, f"={cl}{R_B_NET}/(1+{DISC_REF})^{i}", fill=F_ROW_B, fmt=FMT_NIS, align=A_C)
    R += 1

    scenario_label_row(ws, R, f"תרחיש C — רכישת {t['name_eff']} | פחת מואץ (עם תמריץ)", F_SCN_C)
    R += 1
    sc(ws, R, 2, "* OPEX זהה ל-B — ההבדל הוא מגן המס המואץ בלבד", font=FONT_SM, align=A_R)
    R += 1
    R_C_NET = R; R += 1
    R_C_DISC = R; R += 1
    for rr, lbl in [(R_C_NET, "תזרים נקי — C"), (R_C_DISC, "תזרים מהוון — C")]:
        sc(ws, rr, 2, lbl, fill=F_ROW_C, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪", align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)
    sc(ws, R_C_NET, 6, f"=F{R_B_INV}", fill=F_ROW_C, fmt=FMT_NIS, align=A_C)
    sc(ws, R_C_DISC, 6, f"=F{R_C_NET}", fill=F_ROW_C, fmt=FMT_NIS, align=A_C)
    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl = get_column_letter(col)
        sc(ws, R_C_NET, col, f"={cl}{R_B_EOPC}+{cl}{R_B_OOPC}+{cl}{R_AT_E}", fill=F_ROW_C, fmt=FMT_NIS, align=A_C)
        sc(ws, R_C_DISC, col, f"={cl}{R_C_NET}/(1+{DISC_REF})^{i}", fill=F_ROW_C, fmt=FMT_NIS, align=A_C)
    R += 1

    # ── ד: Results ──
    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=8)
    sc(ws, R, 2, "ד. תוצאות — NPV, החזר השקעה, עלות פיסקלית", fill=F_SUBHEAD, font=FONT_SH, align=A_R)
    R += 1

    def result_row(row, label, formula, fill, fmt=FMT_NIS, unit='ש"ח'):
        sc(ws, row, 2, label, fill=fill, font=Font(name="Arial", bold=True, size=10), align=A_R)
        sc(ws, row, 3, unit, align=A_R)
        sc(ws, row, 6, formula, fill=fill, fmt=fmt, align=A_C)

    R_NPV_A = R; R += 1
    R_NPV_B = R; R += 1
    R_NPV_C = R; R += 1
    R += 1
    R_DNPV = R; R += 1
    R_INCV = R; R += 1
    R_PB_B = R; R += 1
    R_PB_C = R; R += 1
    R_VERDICT = R; R += 1
    R_CA = R; R += 1
    R_FISCAL = R; R += 1

    YR0 = YR0_LTR
    result_row(R_NPV_A, f"NPV — תרחיש A ({t['name_base']})",
               f"=SUM({YR0}{R_A_DISC}:{MAX_COL_LTR}{R_A_DISC})", F_ROW_A)
    result_row(R_NPV_B, f"NPV — תרחיש B ({t['name_eff']}, ללא תמריץ)",
               f"=SUM({YR0}{R_B_DISC}:{MAX_COL_LTR}{R_B_DISC})", F_ROW_B)
    result_row(R_NPV_C, f"NPV — תרחיש C ({t['name_eff']}, עם תמריץ)",
               f"=SUM({YR0}{R_C_DISC}:{MAX_COL_LTR}{R_C_DISC})", F_ROW_C)
    result_row(R_DNPV, "NPV מצטבר — הצדקת השדרוג (B−A)", f"=F{R_NPV_B}-F{R_NPV_A}", F_INPUT)
    result_row(R_INCV, "ערך התמריץ — תועלת הפחת המואץ (C−B)", f"=F{R_NPV_C}-F{R_NPV_B}", F_POLICY)

    sc(ws, R_PB_B, 2, "תקופת החזר מהוון — B (ללא תמריץ)", fill=F_ROW_B, font=FONT_N, align=A_R)
    sc(ws, R_PB_B, 3, "שנים", align=A_R)
    sc(ws, R_PB_B, 6, payback_array_formula(f"F{R_PB_B}", R_B_DISC, R_A_DISC), fill=F_ROW_B, fmt=FMT_YR, align=A_C)

    sc(ws, R_PB_C, 2, "תקופת החזר מהוון — C (עם תמריץ)", fill=F_ROW_C, font=FONT_N, align=A_R)
    sc(ws, R_PB_C, 3, "שנים", align=A_R)
    sc(ws, R_PB_C, 6, payback_array_formula(f"F{R_PB_C}", R_C_DISC, R_A_DISC), fill=F_ROW_C, fmt=FMT_YR, align=A_C)

    sc(ws, R_VERDICT, 2, "הכרעה — מבחן סף החזר ההשקעה", font=FONT_N, align=A_R)
    sc(ws, R_VERDICT, 3, "הערכה", align=A_R)
    sc(ws, R_VERDICT, 4, "חישוב", align=A_R)
    sc(ws, R_VERDICT, 6,
       f'=IF(AND(ISNUMBER(F{R_PB_B}),F{R_PB_B}<={THRESH_REF}),"כדאי גם ללא תמריץ",'
       f'IF(AND(ISNUMBER(F{R_PB_C}),F{R_PB_C}<={THRESH_REF}),"התמריץ הפך את ההחלטה",'
       f'"לא כדאי גם עם תמריץ"))', align=A_C)

    result_row(R_CA, "NPV מצטבר — ערך ההשקעה המתומרצת מול הנוהג הקיים (C−A)",
               f"=F{R_NPV_C}-F{R_NPV_A}", F_INPUT)
    sc(ws, R_FISCAL, 2, "עלות פיסקלית של התמריץ למדינה (NPV)", fill=F_POLICY, font=FONT_N, align=A_R)
    sc(ws, R_FISCAL, 3, 'ש"ח', align=A_R)
    sc(ws, R_FISCAL, 4, "חישוב", align=A_R)
    sc(ws, R_FISCAL, 6, f"=F{R_INCV}", fill=F_POLICY, fmt=FMT_NIS, align=A_C)

    return R + 1, dict(npv_a=R_NPV_A, npv_b=R_NPV_B, npv_c=R_NPV_C, d_npv=R_DNPV, inc=R_INCV,
                        pb_b=R_PB_B, pb_c=R_PB_C, verdict=R_VERDICT, ca=R_CA, fiscal=R_FISCAL)


def _summary_block(ws, tech_results, R_start):
    R = R_start
    section_hdr(ws, R, "סיכום השוואתי לפי טכנולוגיה", number=4, col_end=10)
    R += 1
    hdrs = [(2, "טכנולוגיה"), (3, "NPV A"), (4, "NPV B"), (5, "NPV C"), (6, "ΔNPV B−A"),
            (7, "ערך תמריץ C−B"), (8, "החזר B"), (9, "החזר C"), (10, "הכרעה")]
    for col, lbl in hdrs:
        sc(ws, R, col, lbl, fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    R += 1
    first_data_row = R
    for t, rmap in tech_results:
        sc(ws, R, 2, t['name_eff'], font=FONT_N, align=A_R)
        for col, key, fmt in [(3, 'npv_a', FMT_NIS), (4, 'npv_b', FMT_NIS), (5, 'npv_c', FMT_NIS),
                               (6, 'd_npv', FMT_NIS), (7, 'inc', FMT_NIS),
                               (8, 'pb_b', FMT_YR), (9, 'pb_c', FMT_YR)]:
            sc(ws, R, col, f"=$F${rmap[key]}", fmt=fmt, align=A_C)
        sc(ws, R, 10, f"=$F${rmap['verdict']}", align=A_C)
        R += 1
    last_data_row = R - 1
    R += 1
    sc(ws, R, 2, "סה\"כ עלות פיסקלית (סכום C−B, שלוש הטכנולוגיות)", fill=F_POLICY,
       font=Font(name="Arial", bold=True, size=10, color="FFFFFF"), align=A_R)
    total_refs = "+".join(f"$F${rmap['fiscal']}" for _, rmap in tech_results)
    sc(ws, R, 7, f"={total_refs}", fill=F_POLICY, fmt=FMT_NIS, align=A_C)
    R += 2
    sc(ws, R, 2,
       "** יעד רגישות (טבלאות Data Table לשעות פעילות ולמכפיל הפחת) לא שוכפל בגרסה זו — "
       "יש לבנות מחדש ב-Excel (Data > What-If Analysis > Data Table) מול נקודות הקיבולת הממוצעות החדשות.",
       font=Font(name="Arial", italic=True, color="FF0000", size=9), align=A_RW)


def main():
    wb = openpyxl.Workbook()
    tech_refs = build_global_sheet(wb)
    build_analysis_sheet(wb, tech_refs)
    out = "/home/user/shmags-2/projects/energy-program/tax_incentive_model_v4.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print("Sheets:", [ws.title for ws in wb.worksheets])


if __name__ == "__main__":
    main()
