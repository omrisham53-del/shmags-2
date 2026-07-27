#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tax Incentive Model Generator — v4 (2026-07-13)
National Energy Efficiency Program — EcoTraders

Changes from v3 (generate_tax_model.py):
  - Electric steam dropped (weak primary-energy case vs. fuel-oil baseline)
  - 3 technologies: heat pumps, chillers, VSD compressors
  - Every assumption now shows its derivation (capacity x efficiency x hours
    -> annual consumption) instead of a blind "PENDING" input, and every row
    carries a source in the מקור column
  - Heat pump baseline is a real fuel-combustion calc (mazut/diesel oven),
    not electricity x price -- this was the structural bug in v3
  - Real CapEx: chillers from grant-program data (median/estimate), heat
    pumps and VSD from grant-program averages (VSD baseline via a sourced
    fixed-speed-vs-VSD premium)
  - Discount rate set to 6%, agreed with Daniel
  - Full "Sources & Methodology" section at the bottom of the assumptions
    sheet

Three-scenario comparison per technology (unchanged from v3):
  A — Buy baseline (inefficient) technology
  B — Buy efficient technology, standard depreciation
  C — Buy efficient technology, accelerated depreciation (incentive)
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─── COLORS ───────────────────────────────────────────────────────────────────
F_INPUT   = PatternFill("solid", fgColor="FFFF00")   # yellow  – user input
F_CONTROL = PatternFill("solid", fgColor="FFC000")   # orange  – verify/update
F_HEADING = PatternFill("solid", fgColor="4472C4")   # blue    – section header
F_SUBHEAD = PatternFill("solid", fgColor="BDD7EE")   # l.blue  – col headers
F_RESULT  = PatternFill("solid", fgColor="E2EFDA")   # l.green – Scenario A results
F_SCN_B   = PatternFill("solid", fgColor="DDEBF7")   # blue-tint – Scenario B (efficient, std)
F_SCN_C   = PatternFill("solid", fgColor="FFF2CC")   # cream   – Scenario C (efficient, accel)
F_POLICY  = PatternFill("solid", fgColor="FF6B6B")   # red     – key policy row
F_SOURCED = PatternFill("solid", fgColor="C6E0B4")   # green   – real sourced data (not an estimate)
F_ESTIM   = PatternFill("solid", fgColor="FCE4D6")   # peach   – derived/estimated (sourced but not a direct quote)
F_NONE    = PatternFill("solid", fgColor="FFFFFF")

FONT_H   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
FONT_SH  = Font(name="Arial", bold=True, color="000000", size=10)
FONT_N   = Font(name="Arial", size=10)
FONT_P   = Font(name="Arial", italic=True, color="FF0000", size=10)
FONT_SM  = Font(name="Arial", italic=True, size=9, color="595959")
FONT_B   = Font(name="Arial", bold=True, size=10)

A_R = Alignment(horizontal="right",  vertical="center", readingOrder=2)
A_C = Alignment(horizontal="center", vertical="center", readingOrder=2)
A_RW = Alignment(horizontal="right", vertical="center", readingOrder=2, wrap_text=True)

FMT_NIS = '#,##0 ₪'
FMT_PCT = '0.0%'
FMT_NUM = '#,##0'
FMT_YR  = '0.0'
FMT_X   = '0.0"×"'
FMT_3   = '#,##0.000'

# ─── SHEET / CELL REFERENCES ──────────────────────────────────────────────────
GLOBAL_SHEET   = "נתונים והנחות"
ANALYSIS_SHEET = "ניתוח"

TAX_REF     = f"'{GLOBAL_SHEET}'!$F$10"
DISC_REF    = f"'{GLOBAL_SHEET}'!$F$11"
ELEC_REF    = f"'{GLOBAL_SHEET}'!$F$12"   # agorot/kWh — divide by 100 for ₪/kWh
DIESEL_PRICE_REF = f"'{GLOBAL_SHEET}'!$F$15"   # ₪/liter
MAZUT_PRICE_REF  = f"'{GLOBAL_SHEET}'!$F$16"   # ₪/ton
DIESEL_CAL_REF   = f"'{GLOBAL_SHEET}'!$F$17"   # ton fuel/MWh
MAZUT_CAL_REF    = f"'{GLOBAL_SHEET}'!$F$18"   # ton fuel/MWh
DIESEL_DENS_REF  = f"'{GLOBAL_SHEET}'!$F$19"   # kg/liter
MULT_REF    = f"'{GLOBAL_SHEET}'!$F$22"   # depreciation multiplier (key policy param)

# ─── TECHNOLOGIES ─────────────────────────────────────────────────────────────
# baseline_type: 'fuel' (heat pump — mazut/diesel oven) or 'electric' (chillers, VSD)
TECHS = [
    dict(
        number="3.1",
        name_efficient="משאבות חום (חימום מים)",
        name_baseline="תנור מזוט/סולר",
        baseline_type="fuel",
        lifetime_efficient=10,
        lifetime_baseline=15,
        std_depr_pct=0.10,
        std_depr_yrs=10,
        notes="החלפת תנור מזוט/סולר במשאבת חום | אורך חיים: 10 vs 15 שנה",
    ),
    dict(
        number="3.2",
        name_efficient="צ'ילרים",
        name_baseline="מערכת קירור קונבנציונלית",
        baseline_type="electric",
        lifetime_efficient=17,
        lifetime_baseline=15,
        std_depr_pct=0.10,
        std_depr_yrs=10,
        notes="שדרוג מערכת קירור תעשייתית | אורך חיים: 17 vs 15 שנה",
    ),
    dict(
        number="3.3",
        name_efficient="מדחסי VSD",
        name_baseline="מדחס מהירות קבועה",
        baseline_type="electric",
        lifetime_efficient=12,
        lifetime_baseline=12,
        std_depr_pct=0.10,
        std_depr_yrs=10,
        notes="מדחסי אוויר עם בקרת מהירות משתנה | אורך חיים: 12 שנה",
    ),
]

MAX_YRS     = 20
YR0_COL     = 6                          # column index of "Year 0"
MAX_COL_IDX = YR0_COL + MAX_YRS
MAX_COL_LTR = get_column_letter(MAX_COL_IDX)
YR0_LTR     = get_column_letter(YR0_COL)  # "F"


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def sc(ws, row, col, value=None, fill=None, font=None, align=None, fmt=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if fill:  c.fill = fill
    if font:  c.font = font
    if align: c.alignment = align
    if fmt:   c.number_format = fmt
    return c


def section_hdr(ws, row, label, number=None, col_start=2, col_end=8):
    text = f"{number}. {label}" if number else label
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start)
    c.value = text
    c.fill  = F_HEADING
    c.font  = FONT_H
    c.alignment = A_R


def year_header_row(ws, row, yr0_label="שנה 0"):
    hdrs = [(2, "פרמטר"), (3, "יחידות"), (4, "מקור"),
            (5, "הערות"), (6, yr0_label)]
    for col, val in hdrs:
        sc(ws, row, col, val, fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    for i in range(1, MAX_YRS + 1):
        sc(ws, row, 6 + i, f"שנה {i}", fill=F_SUBHEAD, font=FONT_SH, align=A_C)


def scenario_label_row(ws, row, label, fill, col_end=None):
    end = col_end or (6 + MAX_YRS)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=end)
    c = ws.cell(row=row, column=2)
    c.value = label
    c.fill  = fill
    c.font  = Font(name="Arial", bold=True, size=10,
                   color="FFFFFF" if fill in (F_HEADING, F_POLICY) else "000000")
    c.alignment = A_R


def set_col_widths(ws, include_year_cols=False, extra_tech_cols=0):
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 46
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 36
    ws.column_dimensions['F'].width = 18
    for i in range(1, extra_tech_cols + 1):
        ws.column_dimensions[get_column_letter(6 + i)].width = 20
    if include_year_cols:
        for i in range(1, MAX_YRS + 1):
            ws.column_dimensions[get_column_letter(6 + i)].width = 13


def color_legend(ws, row=1):
    ws.cell(row=row, column=2).value = "מקרא צבעים"
    ws.cell(row=row, column=2).font = FONT_SH
    items = [
        ("ערך להזנה",                       F_INPUT),
        ("נתון לעדכון / בקרה",              F_CONTROL),
        ("נתון ממקור אמיתי (לא הערכה)",     F_SOURCED),
        ("נתון מוערך/נגזר (עם מקור)",       F_ESTIM),
        ("תרחיש A — טכנולוגיה קיימת",       F_RESULT),
        ("תרחיש B — יעיל, ללא תמריץ",       F_SCN_B),
        ("תרחיש C — יעיל, עם תמריץ",        F_SCN_C),
        ("פרמטר מדיניות מרכזי",              F_POLICY),
    ]
    for i, (label, fill) in enumerate(items):
        r = row + 1 + i
        sc(ws, r, 2, label, fill=fill, font=FONT_N, align=A_R)


def pbk_formula(row_cumul):
    """Payback via cumulative NPV crossover — linear interpolation."""
    rng = f"{YR0_LTR}{row_cumul}:{MAX_COL_LTR}{row_cumul}"
    return (
        f"=IFERROR("
        f"MATCH(1,({rng}>0)*1,0)-2+"
        f"ABS(INDEX({rng},MATCH(1,({rng}>0)*1,0)-1))/"
        f"(INDEX({rng},MATCH(1,({rng}>0)*1,0))-"
        f"INDEX({rng},MATCH(1,({rng}>0)*1,0)-1)),"
        f"\"לא מגיע לפירעון\")"
    )


# ─── SHEET 1: GLOBAL ASSUMPTIONS ──────────────────────────────────────────────

def build_global_sheet(wb):
    ws = wb.active
    ws.title = GLOBAL_SHEET
    ws.sheet_view.rightToLeft = True
    set_col_widths(ws, extra_tech_cols=len(TECHS))
    color_legend(ws, row=1)

    # ── Section 1: Financial Parameters ──
    section_hdr(ws, 9, "פרמטרים פיננסיים", number=1, col_end=6 + len(TECHS))
    fin_rows = [
        (10, "שיעור מס חברות", "%", "רשות המסים", "", 0.23, F_CONTROL, FMT_PCT),
        (11, "שיעור היוון (יזמי / פרטי)", "%", "סוכם עם דניאל",
             "6% — הוחלט עם דניאל 2026-07-13 (לעומת 10% תעשייתי שנשקל קודם)",
             0.06, F_SOURCED, FMT_PCT),
        (12, 'מחיר חשמל ממוצע לתעשייה', 'אג\'/קוט"ש', "חברת החשמל",
             "ממוצע SMP 2024 — נתון כללי, ייתכן שיעודכן לממוצע משוקלל לפי שעות תעו\"ז",
             14.0, F_INPUT, '#,##0.0'),
    ]
    for r, label, units, source, notes, val, fill, fmt in fin_rows:
        sc(ws, r, 2, label,  fill=fill, font=FONT_N, align=A_R)
        sc(ws, r, 3, units,  align=A_R)
        sc(ws, r, 4, source, align=A_R)
        sc(ws, r, 5, notes,  align=A_RW)
        sc(ws, r, 6, val,    fill=fill, fmt=fmt, align=A_C)

    # ── Section 1b: Fuel Prices & Conversion Factors ──
    section_hdr(ws, 14, "מחירי דלקים ומקדמי המרה (עבור בסיס משאבות חום)", number="1ב",
                col_end=6 + len(TECHS))
    fuel_rows = [
        (15, "מחיר סולר בתעשייה", "₪/ליטר", "משרד האנרגיה 2024 [1]",
             'מחירים תיאורטיים של מוצרי דלק שאינם בפיקוח, לפני מע"מ ובלו',
             2.59368, F_SOURCED, FMT_3),
        (16, "מחיר מזוט בתעשייה", "₪/טון", "משרד האנרגיה 2024 [1]",
             'תוקן מהיחידה שהוזנה במקור (₪/ק"ג -> ₪/טון) — הערך המקורי לא עקבי עם טווח המחירים בדוח המקורי',
             2344.72, F_ESTIM, FMT_NIS),
        (17, "יחס קלורי — סולר", "טון דלק/MWh", "אקסל של עמרי (MRV) [2]", "", 0.085, F_SOURCED, FMT_3),
        (18, "יחס קלורי — מזוט", "טון דלק/MWh", "אקסל של עמרי (MRV) [2]", "", 0.088, F_SOURCED, FMT_3),
        (19, "צפיפות סולר", "ק\"ג/ליטר", "אקסל של עמרי (MRV) [2]",
             "להמרת טון סולר לליטר במידה ונבחר סולר כדלק בסיס", 0.82, F_SOURCED, FMT_3),
    ]
    for r, label, units, source, notes, val, fill, fmt in fuel_rows:
        sc(ws, r, 2, label,  fill=fill, font=FONT_N, align=A_R)
        sc(ws, r, 3, units,  align=A_R)
        sc(ws, r, 4, source, align=A_R)
        sc(ws, r, 5, notes,  align=A_RW)
        sc(ws, r, 6, val,    fill=fill, fmt=fmt, align=A_C)

    # ── Section 2: Depreciation Multiplier ──
    section_hdr(ws, 21, "פרמטר תמריץ — מכפיל פחת (פחת מואץ)", number=2,
                col_end=6 + len(TECHS))
    sc(ws, 22, 2, "מכפיל שיעור הפחת  ←  פרמטר המדיניות המרכזי", fill=F_POLICY,
       font=Font(name="Arial", bold=True, size=11, color="FFFFFF"), align=A_R)
    sc(ws, 22, 3, "מכפיל", fill=F_POLICY,
       font=Font(name="Arial", bold=True, size=11, color="FFFFFF"), align=A_C)
    sc(ws, 22, 4, "פרמטר ניתן לשינוי", align=A_R)
    sc(ws, 22, 5, "1.0 = ללא תמריץ | 2.0 = קפריסין (כפול, 5 שנים) | 5.0 = 2 שנים",
       fill=F_POLICY, font=Font(name="Arial", italic=True, size=9, color="FFFFFF"), align=A_R)
    sc(ws, 22, 6, 2.0, fill=F_INPUT,
       font=Font(name="Arial", bold=True, size=14, color="C00000"), fmt=FMT_X, align=A_C)

    for r, txt in [
        (23, ("מכפיל 2.0 + פחת סטנדרטי 10% = שיעור מואץ 20%/שנה למשך 5 שנים. "
              "סה\"כ ניכוי = 100% CapEx — היתרון הוא קדמות הניכוי (Time Value of Money).")),
        (24, ("דוגמאות: 1.0 = 10%/שנה × 10 שנים (סטנדרטי) | "
              "2.0 = 20%/שנה × 5 שנים | 3.33 = 33%/שנה × 3 שנים | "
              "5.0 = 50%/שנה × 2 שנים (מודל קפריסין/אירלנד)")),
    ]:
        ws.cell(row=r, column=2).value = txt
        ws.cell(row=r, column=2).font = FONT_SM
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6 + len(TECHS))

    # ── Section 3: Standard Depreciation Rates ──
    section_hdr(ws, 26, "שיעורי פחת סטנדרטיים — תקנות מס הכנסה", number=3,
                col_end=6 + len(TECHS))
    for r, label, val in [
        (27, "משאבות חום",  0.10),
        (28, "צ'ילרים",      0.10),
        (29, "מדחסי VSD",    0.10),
    ]:
        sc(ws, r, 2, label, font=FONT_N, align=A_R)
        sc(ws, r, 3, "% לשנה",  align=A_R)
        sc(ws, r, 4, "תקנות פחת", align=A_R)
        sc(ws, r, 5, "ציוד מכני", align=A_R)
        sc(ws, r, 6, val, fmt=FMT_PCT, align=A_C)

    sc(ws, 31, 2, "* שיעורי הפחת הם הנחות עבודה — יש לוודא מול יועץ מס לפני הגשה", font=FONT_SM)

    # ── Section 4: Technology Assumptions Table ──
    T_START_COL = 6
    T_END_COL   = T_START_COL + len(TECHS) - 1

    section_hdr(ws, 33, "הנחות לפי טכנולוגיה — כל הנתונים לעדכון בטבלה זו", number=4,
                col_end=T_END_COL)

    sc(ws, 34, 2, "פרמטר",   fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    sc(ws, 34, 3, "יחידות",  fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    sc(ws, 34, 4, "מקור",    fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    sc(ws, 34, 5, "הערות",   fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    for i, tech in enumerate(TECHS):
        sc(ws, 34, T_START_COL + i, tech['name_efficient'],
           fill=PatternFill("solid", fgColor="1F4E79"),
           font=Font(name="Arial", bold=True, color="FFFFFF", size=10), align=A_C)
        sc(ws, 35, T_START_COL + i, f"vs. {tech['name_baseline']}", font=FONT_SM, align=A_C)

    # Row map (returned refs use these row numbers — no longer hardcoded elsewhere,
    # build_analysis_sheet only uses the tech_refs dict)
    R_CAP      = 36
    R_HRS      = 37
    R_COP      = 38   # heat pump only
    R_COMB     = 39   # heat pump only — combustion efficiency, baseline oven
    R_FUELTYPE = 40   # heat pump only — "מזוט" or "סולר"
    R_EFF_B    = 41   # chiller/VSD only — native-unit efficiency, baseline
    R_EFF_E    = 42   # chiller/VSD only — native-unit efficiency, efficient
    R_SVPCT    = 43   # chiller/VSD only — computed savings %
    R_CONS_B   = 44   # computed: tons fuel/yr (heat pump) or kWh/yr (chiller/VSD)
    R_CONS_E   = 45   # computed: kWh/yr, all electric
    R_CAPEX_B  = 47
    R_CAPEX_E  = 48
    R_DCAPEX   = 49
    R_LIFE_B   = 50
    R_LIFE_E   = 51
    R_DEGRAD   = 52
    R_OPEXO_B  = 53
    R_OPEXO_E  = 54
    R_OPEXE_B  = 55
    R_OPEXE_E  = 56

    row_labels = [
        (R_CAP,   "קיבולת מותקנת",                        "ראו הערה",     ""),
        (R_HRS,   "שעות פעילות שנתיות",                    "שעות",        ""),
        (R_COP,   "COP — משאבת חום (יעיל)",                "יחס",         "רלוונטי למשאבות חום בלבד"),
        (R_COMB,  "יעילות בעירה — תנור בסיס",               "%",           "רלוונטי למשאבות חום בלבד"),
        (R_FUELTYPE, "סוג דלק בסיס",                        "טקסט",        'רלוונטי למשאבות חום בלבד — "מזוט" או "סולר"'),
        (R_EFF_B, "יעילות אנרגטית — בסיסי",                 "ראו הערה",     "רלוונטי לצ'ילרים/VSD בלבד"),
        (R_EFF_E, "יעילות אנרגטית — יעיל",                  "ראו הערה",     "רלוונטי לצ'ילרים/VSD בלבד"),
        (R_SVPCT, "חיסכון אנרגטי (מחושב)",                  "% מצריכה",    "רלוונטי לצ'ילרים/VSD בלבד — נגזר משורות 41-42"),
        (R_CONS_B, "צריכת אנרגיה/דלק שנתית — בסיסי",        "ראו הערה",     "טון דלק (משאבות חום) | קוט\"ש (צ'ילרים/VSD)"),
        (R_CONS_E, "צריכת חשמל שנתית — יעיל",               'קוט"ש/שנה',   'הצד היעיל תמיד חשמלי בכל הטכנולוגיות'),
        (R_CAPEX_B, "CapEx — ציוד קיים (בסיסי)",            "₪",           ""),
        (R_CAPEX_E, "CapEx — ציוד יעיל",                    "₪",           ""),
        (R_DCAPEX, "ΔCapEx = יעיל − בסיסי",                 "₪",           ""),
        (R_LIFE_B, "אורך חיים — ציוד בסיסי",                "שנים",        ""),
        (R_LIFE_E, "אורך חיים — ציוד יעיל",                 "שנים",        ""),
        (R_DEGRAD, "גורם שחיקת ביצועים",                    "%/שנה",       ""),
        (R_OPEXO_B, "OPEX אחר — בסיסי (תחזוקה, חלפים...)", "₪/שנה",       "⚠ ממתין לנתוני רפי"),
        (R_OPEXO_E, "OPEX אחר — יעיל (תחזוקה, חלפים...)",  "₪/שנה",       "⚠ ממתין לנתוני רפי"),
        (R_OPEXE_B, "OPEX אנרגיה/דלק שנה 1 — בסיסי",       "₪/שנה",       "חישוב"),
        (R_OPEXE_E, "OPEX אנרגיה שנה 1 — יעיל",             "₪/שנה",       "חישוב"),
    ]
    for row, label, units, note in row_labels:
        sc(ws, row, 2, label, font=FONT_N, align=A_R)
        sc(ws, row, 3, units, align=A_R)
        if note:
            sc(ws, row, 5, note, font=FONT_SM, align=A_RW)

    # Source column per row (shared text, since it's the methodology not the value)
    src_labels = {
        R_CAP:     "baseline-technology-data.md — נקודת קיבולת נבחרת",
        R_HRS:     "ראו הערה בעמודת הטכנולוגיה",
        R_COP:     "Sprsun — עמוד מפרט מוצר [3]",
        R_COMB:    "ASME PTC 4 [4]",
        R_FUELTYPE: "הנחת עבודה — ניתן לשנות",
        R_EFF_B:   "ראו הערה בעמודת הטכנולוגיה",
        R_EFF_E:   "ראו הערה בעמודת הטכנולוגיה",
        R_SVPCT:   "חישוב",
        R_CONS_B:  "חישוב — קיבולת × שעות × יעילות/COP",
        R_CONS_E:  "חישוב",
        R_CAPEX_B: "ראו הערה בעמודת הטכנולוגיה",
        R_CAPEX_E: "ראו הערה בעמודת הטכנולוגיה",
        R_DCAPEX:  "חישוב",
        R_LIFE_B:  "ממצאי הנדסה / הנחת עבודה",
        R_LIFE_E:  "ממצאי הנדסה / הנחת עבודה",
        R_DEGRAD:  "הנחת עבודה (0.5-1%/שנה טיפוסי לציוד מכני)",
        R_OPEXO_B: "נתוני רפי (טרם התקבל)",
        R_OPEXO_E: "נתוני רפי (טרם התקבל)",
        R_OPEXE_B: "חישוב",
        R_OPEXE_E: "חישוב",
    }
    for row, src in src_labels.items():
        sc(ws, row, 4, src, align=A_RW)

    # ── Per-technology values ──
    tech_col = {}
    for i, tech in enumerate(TECHS):
        col = T_START_COL + i
        tech_col[tech['number']] = col
    HP_COL  = tech_col["3.1"]
    CH_COL  = tech_col["3.2"]
    VSD_COL = tech_col["3.3"]

    def cl(col):
        return get_column_letter(col)

    # -- Heat pump (fuel baseline) --
    sc(ws, R_CAP, HP_COL, 40, fill=F_SOURCED, fmt=FMT_NUM, align=A_C)
    sc(ws, R_HRS, HP_COL, "PENDING", fill=F_INPUT, font=FONT_P, fmt=FMT_NUM, align=A_C)
    sc(ws, R_COP, HP_COL, 4.13, fill=F_SOURCED, fmt='0.00', align=A_C)
    sc(ws, R_COMB, HP_COL, 0.835, fill=F_ESTIM, fmt=FMT_PCT, align=A_C)
    sc(ws, R_FUELTYPE, HP_COL, "מזוט", fill=F_INPUT, font=FONT_B, align=A_C)
    sc(ws, R_CAPEX_B, HP_COL, "PENDING", fill=F_INPUT, font=FONT_P, fmt=FMT_NIS, align=A_C)
    sc(ws, R_CAPEX_E, HP_COL, 165177, fill=F_ESTIM, fmt=FMT_NIS, align=A_C)

    # -- Chillers (electric baseline) --
    sc(ws, R_CAP, CH_COL, 500, fill=F_SOURCED, fmt=FMT_NUM, align=A_C)
    sc(ws, R_HRS, CH_COL, 3000, fill=F_ESTIM, fmt=FMT_NUM, align=A_C)
    sc(ws, R_EFF_B, CH_COL, 0.60, fill=F_SOURCED, fmt='0.00', align=A_C)
    sc(ws, R_EFF_E, CH_COL, 0.48, fill=F_SOURCED, fmt='0.00', align=A_C)
    sc(ws, R_CAPEX_B, CH_COL, 1781000, fill=F_ESTIM, fmt=FMT_NIS, align=A_C)
    sc(ws, R_CAPEX_E, CH_COL, 2093000, fill=F_SOURCED, fmt=FMT_NIS, align=A_C)

    # -- VSD (electric baseline) --
    sc(ws, R_CAP, VSD_COL, 45, fill=F_SOURCED, fmt=FMT_NUM, align=A_C)
    sc(ws, R_HRS, VSD_COL, 6400, fill=F_ESTIM, fmt=FMT_NUM, align=A_C)
    sc(ws, R_EFF_B, VSD_COL, 21.5, fill=F_SOURCED, fmt='0.00', align=A_C)
    sc(ws, R_EFF_E, VSD_COL, 16.5, fill=F_SOURCED, fmt='0.00', align=A_C)
    sc(ws, R_CAPEX_B, VSD_COL, 255061, fill=F_ESTIM, fmt=FMT_NIS, align=A_C)
    sc(ws, R_CAPEX_E, VSD_COL, 312450, fill=F_SOURCED, fmt=FMT_NIS, align=A_C)

    # -- Shared rows across all 3 techs (lifetime, degradation, ΔCapEx, other OPEX) --
    for i, tech in enumerate(TECHS):
        col = T_START_COL + i
        c = cl(col)
        sc(ws, R_LIFE_B, col, tech['lifetime_baseline'], fill=F_CONTROL, fmt=FMT_YR, align=A_C)
        sc(ws, R_LIFE_E, col, tech['lifetime_efficient'], fill=F_CONTROL, fmt=FMT_YR, align=A_C)
        sc(ws, R_DEGRAD, col, 0.005, fill=F_INPUT, fmt=FMT_PCT, align=A_C)
        sc(ws, R_OPEXO_B, col, "PENDING", fill=F_INPUT, font=FONT_P, fmt=FMT_NIS, align=A_C)
        sc(ws, R_OPEXO_E, col, "PENDING", fill=F_INPUT, font=FONT_P, fmt=FMT_NIS, align=A_C)
        sc(ws, R_DCAPEX, col,
           f"=IF(ISNUMBER({c}{R_CAPEX_E})*ISNUMBER({c}{R_CAPEX_B}),{c}{R_CAPEX_E}-{c}{R_CAPEX_B},\"PENDING\")",
           fmt=FMT_NIS, align=A_C)

    # -- Computed rows: savings %, consumption, OPEX energy --
    for i, tech in enumerate(TECHS):
        col = T_START_COL + i
        c = cl(col)
        if tech['baseline_type'] == 'fuel':
            # Heat pump: fuel tons/yr baseline, kWh/yr efficient
            sc(ws, R_SVPCT, col, "-", align=A_C)
            sc(ws, R_CONS_B, col,
               f"=IF(AND(ISNUMBER({c}{R_CAP}),ISNUMBER({c}{R_HRS}),ISNUMBER({c}{R_COMB})),"
               f"({c}{R_CAP}*{c}{R_HRS}/1000/{c}{R_COMB})*"
               f"IF({c}{R_FUELTYPE}=\"מזוט\",{MAZUT_CAL_REF},{DIESEL_CAL_REF}),\"PENDING\")",
               fmt='#,##0.00', align=A_C)
            sc(ws, R_CONS_E, col,
               f"=IF(AND(ISNUMBER({c}{R_CAP}),ISNUMBER({c}{R_HRS}),ISNUMBER({c}{R_COP})),"
               f"{c}{R_CAP}*{c}{R_HRS}/{c}{R_COP},\"PENDING\")",
               fmt=FMT_NUM, align=A_C)
            sc(ws, R_OPEXE_B, col,
               f"=IF(ISNUMBER({c}{R_CONS_B}),"
               f"IF({c}{R_FUELTYPE}=\"מזוט\",{c}{R_CONS_B}*{MAZUT_PRICE_REF},"
               f"{c}{R_CONS_B}*1000/{DIESEL_DENS_REF}*{DIESEL_PRICE_REF}),\"PENDING\")",
               fmt=FMT_NIS, align=A_C)
        else:
            # Chillers / VSD: kWh/yr both sides, via native-unit efficiency
            sc(ws, R_SVPCT, col,
               f"=IF(AND(ISNUMBER({c}{R_EFF_B}),ISNUMBER({c}{R_EFF_E})),"
               f"({c}{R_EFF_B}-{c}{R_EFF_E})/{c}{R_EFF_B},\"PENDING\")",
               fmt=FMT_PCT, align=A_C)
            if tech['number'] == "3.2":  # chillers: kWh = RT x kW/ton x hours
                sc(ws, R_CONS_B, col,
                   f"=IF(AND(ISNUMBER({c}{R_CAP}),ISNUMBER({c}{R_HRS}),ISNUMBER({c}{R_EFF_B})),"
                   f"{c}{R_CAP}*{c}{R_EFF_B}*{c}{R_HRS},\"PENDING\")",
                   fmt=FMT_NUM, align=A_C)
            else:  # VSD: kWh = nameplate kW x hours
                sc(ws, R_CONS_B, col,
                   f"=IF(AND(ISNUMBER({c}{R_CAP}),ISNUMBER({c}{R_HRS})),"
                   f"{c}{R_CAP}*{c}{R_HRS},\"PENDING\")",
                   fmt=FMT_NUM, align=A_C)
            sc(ws, R_CONS_E, col,
               f"=IF(AND(ISNUMBER({c}{R_CONS_B}),ISNUMBER({c}{R_SVPCT})),"
               f"{c}{R_CONS_B}*(1-{c}{R_SVPCT}),\"PENDING\")",
               fmt=FMT_NUM, align=A_C)
            sc(ws, R_OPEXE_B, col,
               f"=IF(ISNUMBER({c}{R_CONS_B}),{c}{R_CONS_B}*{ELEC_REF}/100,\"PENDING\")",
               fmt=FMT_NIS, align=A_C)
        # Efficient-side OPEX is always electric, all 3 techs
        sc(ws, R_OPEXE_E, col,
           f"=IF(ISNUMBER({c}{R_CONS_E}),{c}{R_CONS_E}*{ELEC_REF}/100,\"PENDING\")",
           fmt=FMT_NIS, align=A_C)

    # Heat pump hours note (fixed placement now that column is known)
    sc(ws, R_HRS, 5,
       'ממתין לייעוץ מהנדס EcoTraders (משאבות חום) | צ\'ילרים: 3,000 שעות (מספר עבודה) | VSD: 6,400 שעות (אמצע 6,000-6,800)',
       font=FONT_SM, align=A_RW)

    sc(ws, 58, 2,
       "* OPEX אחר ממתין לנתוני רפי. שאר הנתונים מקורם ב-baseline-technology-data.md ומספרים בהערות/מקורות למטה.",
       font=FONT_SM)

    # ── Sources & Methodology ──
    build_sources_section(ws, start_row=61, col_end=T_END_COL)

    # Build and return cell reference dict for each tech (used by analysis sheet)
    tech_refs = []
    for i, tech in enumerate(TECHS):
        c = cl(T_START_COL + i)
        s = GLOBAL_SHEET
        tech_refs.append({
            'capex_b':  f"'{s}'!${c}${R_CAPEX_B}",
            'capex_e':  f"'{s}'!${c}${R_CAPEX_E}",
            'dcapex':   f"'{s}'!${c}${R_DCAPEX}",
            'life_b':   f"'{s}'!${c}${R_LIFE_B}",
            'life_e':   f"'{s}'!${c}${R_LIFE_E}",
            'degrad':   f"'{s}'!${c}${R_DEGRAD}",
            'opex_o_b': f"'{s}'!${c}${R_OPEXO_B}",
            'opex_o_e': f"'{s}'!${c}${R_OPEXO_E}",
            'opex_e_b': f"'{s}'!${c}${R_OPEXE_B}",
            'opex_e_e': f"'{s}'!${c}${R_OPEXE_E}",
        })
    return tech_refs


def tech_by_num(num):
    for t in TECHS:
        if t['number'] == num:
            return t
    return None


def build_sources_section(ws, start_row, col_end):
    R = start_row
    section_hdr(ws, R, "מקורות ומתודולוגיה", number=5, col_end=col_end)
    R += 1
    sources = [
        ("[1]", "משרד האנרגיה, מינהל הדלק והגז — \"מחירים תיאורטיים של מוצרי דלק שאינם בפיקוח\", שנת 2024. "
                "קובץ המקור: orl22024_4.xlsx. מחירי סולר/מזוט בתעשייה ממוצעים לשנת 2024, לפני מע\"מ ובלו."),
        ("[2]", "אקסל של עמרי — מקדמי MRV (Monitoring, Reporting & Verification) הרשמיים של ישראל: "
                "ערכים קלוריים לסולר/מזוט וצפיפות סולר."),
        ("[3]", "Sprsun commercial heat pump product pages — COP 4.13 at 40kW (37-45kW model), "
                "COP 3.23-3.24 at 70kW (42-70kW model). Real capacity ceiling is 70kW, not the "
                "150kW originally assumed in an earlier draft."),
        ("[4]", "ASME PTC 4 (fuel-to-heat combustion efficiency, fire-tube/water-tube units without "
                "heat recovery, 80-88% typical range — 83.5% midpoint used)."),
        ("[5]", "ASHRAE 90.1 Table 6.8.1-3 (chiller code-minimum, 500 RT bracket) + DOE FEMP "
                "designated-efficient level — baseline 0.60 kW/ton, efficient 0.48 kW/ton."),
        ("[6]", "capex_all_rounds_annotated.xlsx, sheet \"עלות לטון קירור\" — 96 real chiller line "
                "items from the 2017-2022 grant program rounds, 27 with usable capacity data. "
                "Median ₪4,186/ton used for efficient CapEx (recommended over the mean in the "
                "source sheet due to flagged outliers). Baseline (₪3,562/ton = ₪4,186 / 1.175) "
                "estimated via a 10-25% efficiency cost premium (DOE FEMP cost-effectiveness data "
                "+ general market commentary, midpoint used) — a derived estimate, not a direct quote."),
        ("[7]", "capex_all_rounds_annotated.xlsx, sheet \"חישוב ממוצע\" — grant program averages: "
                "heat pumps ₪165,177/unit (49 units), VSD ₪312,450/unit (5 units). Blended averages, "
                "not split by capacity or baseline/efficient — applied as the efficient-side CapEx."),
        ("[8]", "US Air Compressor (usaircompressor.com) — real listed prices: 60 HP (45kW) "
                "fixed-speed $18,150-19,800, 200 HP (150kW) fixed-speed $65,999.99. VSD baseline "
                "CapEx (₪255,061) estimated by dividing the grant-average VSD CapEx by a 15-30% "
                "VSD-over-fixed-speed premium (midpoint 22.5%) — a derived estimate."),
        ("[9]", "CAGI specific power ranges (ISO 1217 basis) — fixed-speed 20-23 kW/100cfm, "
                "VSD 15-18 kW/100cfm at 100 psi reference pressure. Midpoints used (21.5 / 16.5)."),
    ]
    for code, text in sources:
        sc(ws, R, 2, code, font=FONT_B, align=A_R)
        ws.merge_cells(start_row=R, start_column=3, end_row=R, end_column=col_end)
        sc(ws, R, 3, text, font=FONT_SM, align=A_RW)
        ws.row_dimensions[R].height = 30
        R += 1
    R += 1
    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=col_end)
    sc(ws, R, 2,
       "מקרא רמת ביטחון: ירוק = נתון אמיתי ממקור ראשוני (לא הערכה) | כתום-אפרסק = נתון נגזר/מוערך עם "
       "מתודולוגיה מוסברת | צהוב = ממתין לנתונים (רפי / ייעוץ מהנדס). ראו projects/energy-program/"
       "baseline-technology-data.md לפירוט המקורות המלא כולל קישורים.",
       font=FONT_SM, align=A_RW)
    ws.row_dimensions[R].height = 30


# ─── SHEET 2: ANALYSIS ────────────────────────────────────────────────────────

def build_analysis_sheet(wb, tech_refs):
    ws = wb.create_sheet(ANALYSIS_SHEET)
    ws.sheet_view.rightToLeft = True
    set_col_widths(ws, include_year_cols=True)
    color_legend(ws, row=1)

    tech_end_col = 6 + MAX_YRS

    ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=tech_end_col)
    sc(ws, 8, 2, "ניתוח כלכלי — פחת מואץ לפי טכנולוגיה | השוואת שלושה תרחישים",
       fill=F_HEADING, font=Font(name="Arial", bold=True, color="FFFFFF", size=13), align=A_R)

    section_hdr(ws, 10, "פרמטרים פיננסיים", number=1)
    for r, label, units, ref, fmt, fill in [
        (11, "שיעור מס חברות",     "%",          f"={TAX_REF}",        FMT_PCT,    F_CONTROL),
        (12, "שיעור היוון",         "%",          f"={DISC_REF}",       FMT_PCT,    F_SOURCED),
        (13, 'מחיר חשמל (₪/קוט"ש)', '₪/קוט"ש',  f"={ELEC_REF}/100",   '#,##0.000', F_INPUT),
    ]:
        sc(ws, r, 2, label,        fill=fill, font=FONT_N, align=A_R)
        sc(ws, r, 3, units,        align=A_R)
        sc(ws, r, 4, GLOBAL_SHEET, align=A_R)
        sc(ws, r, 6, ref,          fill=fill, fmt=fmt, align=A_C)

    section_hdr(ws, 15, "פרמטר תמריץ — מכפיל פחת", number=2)
    sc(ws, 16, 2, "מכפיל שיעור הפחת  ←  שנה ב'נתונים והנחות'!F22", fill=F_POLICY,
       font=Font(name="Arial", bold=True, size=11, color="FFFFFF"), align=A_R)
    sc(ws, 16, 3, "מכפיל", fill=F_POLICY,
       font=Font(name="Arial", bold=True, size=11, color="FFFFFF"), align=A_C)
    sc(ws, 16, 4, GLOBAL_SHEET, align=A_R)
    sc(ws, 16, 6, f"={MULT_REF}", fill=F_INPUT,
       font=Font(name="Arial", bold=True, size=13, color="C00000"), fmt=FMT_X, align=A_C)
    sc(ws, 17, 2,
       "מכפיל × פחת סטנדרטי = שיעור מואץ | סה\"כ ניכוי = 100% CapEx | תקופת פחת = 1 ÷ שיעור מואץ",
       font=FONT_SM, align=A_R)

    section_hdr(ws, 19, "ניתוחים לפי טכנולוגיה", number=3, col_end=tech_end_col)

    R = 21
    tech_results = []
    for idx, tech in enumerate(TECHS):
        R, rmap = _tech_block(ws, tech, R, tech_refs[idx])
        tech_results.append((tech, rmap))
        R += 2

    _summary_block(ws, tech_results, R)


def _tech_block(ws, tech, R_start, refs):
    R = R_start
    tech_end_col = 6 + MAX_YRS

    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, f"{tech['number']}  {tech['name_efficient']}",
       fill=PatternFill("solid", fgColor="1F4E79"),
       font=Font(name="Arial", bold=True, color="FFFFFF", size=11), align=A_R)
    R += 1
    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, f"* {tech['notes']} | כל הנתונים: גיליון '{GLOBAL_SHEET}' — סעיף 4",
       font=FONT_SM, align=A_R)
    R += 1

    std_r   = tech['std_depr_pct']
    std_yrs = tech['std_depr_yrs']
    life_e  = tech['lifetime_efficient']
    life_b  = tech['lifetime_baseline']

    capex_b  = refs['capex_b']
    capex_e  = refs['capex_e']
    opex_e_b = refs['opex_e_b']
    opex_e_e = refs['opex_e_e']
    opex_o_b = refs['opex_o_b']
    opex_o_e = refs['opex_o_e']
    degrad   = refs['degrad']
    dcapex   = refs['dcapex']

    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, "ב. לוחות פחת (ציוד בסיסי vs. יעיל)", fill=F_SUBHEAD, font=FONT_SH, align=A_R)
    R += 1

    year_header_row(ws, R, yr0_label="שנה 0 (השקעה)")
    R += 1

    R_SD_B = R;  R += 1
    R_ST_B = R;  R += 1
    R_SD_E = R;  R += 1
    R_ST_E = R;  R += 1
    R_AD_E = R;  R += 1
    R_AT_E = R;  R += 1

    depr_rows = [
        (R_SD_B, f"פחת סטנדרטי — {tech['name_baseline']}",   F_RESULT),
        (R_ST_B, f"מגן מס — {tech['name_baseline']}",         F_RESULT),
        (R_SD_E, f"פחת סטנדרטי — {tech['name_efficient']}",   F_SCN_B),
        (R_ST_E, f"מגן מס סטנדרטי — {tech['name_efficient']}", F_SCN_B),
        (R_AD_E, f"פחת מואץ — {tech['name_efficient']}",       F_SCN_C),
        (R_AT_E, f"מגן מס מואץ — {tech['name_efficient']}",    F_SCN_C),
    ]
    for rr, lbl, fill in depr_rows:
        sc(ws, rr, 2, lbl, fill=fill, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪",   align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)
        sc(ws, rr, 6, 0, fill=fill, fmt=FMT_NIS, align=A_C)

    sc(ws, R_SD_B, 5, f"קו ישר | {std_r:.0%}/שנה | עד {std_yrs} שנים | CapEx בסיסי",  align=A_R)
    sc(ws, R_ST_B, 5, "= פחת × שיעור מס", align=A_R)
    sc(ws, R_SD_E, 5, f"קו ישר | {std_r:.0%}/שנה | עד {std_yrs} שנים | CapEx יעיל",   align=A_R)
    sc(ws, R_ST_E, 5, "= פחת × שיעור מס", align=A_R)
    sc(ws, R_AD_E, 5, "= CapEx יעיל × (פחת סטנדרטי × מכפיל) | תקופה = 1 ÷ שיעור מואץ", align=A_R)
    sc(ws, R_AT_E, 5, "= פחת מואץ × שיעור מס", align=A_R)

    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl  = get_column_letter(col)

        sd_b = (f"=IF(ISNUMBER({capex_b}),{capex_b}*{std_r},0)" if i <= std_yrs else "=0")
        sc(ws, R_SD_B, col, sd_b, fill=F_RESULT, fmt=FMT_NIS, align=A_C)
        sc(ws, R_ST_B, col, f"={cl}{R_SD_B}*{TAX_REF}", fill=F_RESULT, fmt=FMT_NIS, align=A_C)

        sd_e = (f"=IF(ISNUMBER({capex_e}),{capex_e}*{std_r},0)" if i <= std_yrs else "=0")
        sc(ws, R_SD_E, col, sd_e, fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
        sc(ws, R_ST_E, col, f"={cl}{R_SD_E}*{TAX_REF}", fill=F_SCN_B, fmt=FMT_NIS, align=A_C)

        ad_e = (
            f"=IF(ISNUMBER({capex_e}),"
            f"IF({i}<=ROUND(1/({std_r}*{MULT_REF}),0),"
            f"{capex_e}*{std_r}*{MULT_REF},0),0)"
        )
        sc(ws, R_AD_E, col, ad_e, fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
        sc(ws, R_AT_E, col, f"={cl}{R_AD_E}*{TAX_REF}", fill=F_SCN_C, fmt=FMT_NIS, align=A_C)

    R += 1

    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, "ג. ניתוח תזרים מזומנים — שלושה תרחישים", fill=F_SUBHEAD, font=FONT_SH, align=A_R)
    R += 1

    year_header_row(ws, R, yr0_label="שנה 0 (השקעה)")
    R += 1

    scenario_label_row(ws, R,
       f"תרחיש A — רכישת {tech['name_baseline']} (ציוד קיים / לא יעיל)",
       PatternFill("solid", fgColor="538135"))
    R += 1

    R_A_INV  = R;  R += 1
    R_A_EOPC = R;  R += 1
    R_A_OOPC = R;  R += 1
    R_A_NET  = R;  R += 1
    R_A_DISC = R;  R += 1
    R_A_CUM  = R;  R += 1

    for rr, lbl, fill in [
        (R_A_INV,  "השקעה ראשונית",            F_RESULT),
        (R_A_EOPC, "OPEX אנרגיה",              F_RESULT),
        (R_A_OOPC, "OPEX אחר (תחזוקה וכו')",   F_RESULT),
        (R_A_NET,  "תזרים נקי — תרחיש A",      F_RESULT),
        (R_A_DISC, "תזרים מהוון — תרחיש A",    F_RESULT),
        (R_A_CUM,  "NPV מצטבר — תרחיש A",      F_RESULT),
    ]:
        sc(ws, rr, 2, lbl, fill=fill, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪",  align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)

    sc(ws, R_A_INV,  6, f"=IF(ISNUMBER({capex_b}),-{capex_b},\"PENDING\")",
       fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_EOPC, 6, 0, fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_OOPC, 6, 0, fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_NET,  6, f"=F{R_A_INV}",   fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_DISC, 6, f"=F{R_A_NET}",   fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_CUM,  6, f"=F{R_A_DISC}",  fill=F_RESULT, fmt=FMT_NIS, align=A_C)

    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl  = get_column_letter(col)
        pcl = get_column_letter(col - 1)

        eopc_a = (f"=IF(ISNUMBER({opex_e_b}),-{opex_e_b},0)" if i <= life_b else "=0")
        sc(ws, R_A_EOPC, col, eopc_a, fill=F_RESULT, fmt=FMT_NIS, align=A_C)

        oopc_a = (f"=IF(ISNUMBER({opex_o_b}),-{opex_o_b},0)" if i <= life_b else "=0")
        sc(ws, R_A_OOPC, col, oopc_a, fill=F_RESULT, fmt=FMT_NIS, align=A_C)

        net_a = f"={cl}{R_A_EOPC}+{cl}{R_A_OOPC}+{cl}{R_ST_B}"
        sc(ws, R_A_NET, col, net_a, fill=F_RESULT, fmt=FMT_NIS, align=A_C)

        sc(ws, R_A_DISC, col, f"={cl}{R_A_NET}/(1+{DISC_REF})^{i}",
           fill=F_RESULT, fmt=FMT_NIS, align=A_C)
        sc(ws, R_A_CUM, col, f"={pcl}{R_A_CUM}+{cl}{R_A_DISC}",
           fill=F_RESULT, fmt=FMT_NIS, align=A_C)

    R += 1

    scenario_label_row(ws, R,
       f"תרחיש B — רכישת {tech['name_efficient']} | פחת סטנדרטי (ללא תמריץ)",
       PatternFill("solid", fgColor="2E75B6"))
    R += 1

    R_B_INV  = R;  R += 1
    R_B_EOPC = R;  R += 1
    R_B_OOPC = R;  R += 1
    R_B_NET  = R;  R += 1
    R_B_DISC = R;  R += 1
    R_B_CUM  = R;  R += 1

    for rr, lbl, fill in [
        (R_B_INV,  "השקעה ראשונית",                  F_SCN_B),
        (R_B_EOPC, "OPEX אנרגיה (עם שחיקה שנתית)",   F_SCN_B),
        (R_B_OOPC, "OPEX אחר (תחזוקה וכו')",          F_SCN_B),
        (R_B_NET,  "תזרים נקי — תרחיש B",             F_SCN_B),
        (R_B_DISC, "תזרים מהוון — תרחיש B",           F_SCN_B),
        (R_B_CUM,  "NPV מצטבר — תרחיש B",             F_SCN_B),
    ]:
        sc(ws, rr, 2, lbl, fill=fill, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪", align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)

    sc(ws, R_B_INV,  6, f"=IF(ISNUMBER({capex_e}),-{capex_e},\"PENDING\")",
       fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_EOPC, 6, 0, fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_OOPC, 6, 0, fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_NET,  6, f"=F{R_B_INV}",  fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_DISC, 6, f"=F{R_B_NET}",  fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_CUM,  6, f"=F{R_B_DISC}", fill=F_SCN_B, fmt=FMT_NIS, align=A_C)

    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl  = get_column_letter(col)
        pcl = get_column_letter(col - 1)

        eopc_b = (f"=IF(ISNUMBER({opex_e_e}),-{opex_e_e}*(1-{degrad})^{i-1},0)"
                  if i <= life_e else "=0")
        sc(ws, R_B_EOPC, col, eopc_b, fill=F_SCN_B, fmt=FMT_NIS, align=A_C)

        oopc_b = (f"=IF(ISNUMBER({opex_o_e}),-{opex_o_e},0)" if i <= life_e else "=0")
        sc(ws, R_B_OOPC, col, oopc_b, fill=F_SCN_B, fmt=FMT_NIS, align=A_C)

        sc(ws, R_B_NET, col, f"={cl}{R_B_EOPC}+{cl}{R_B_OOPC}+{cl}{R_ST_E}",
           fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
        sc(ws, R_B_DISC, col, f"={cl}{R_B_NET}/(1+{DISC_REF})^{i}",
           fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
        sc(ws, R_B_CUM, col, f"={pcl}{R_B_CUM}+{cl}{R_B_DISC}",
           fill=F_SCN_B, fmt=FMT_NIS, align=A_C)

    R += 1

    scenario_label_row(ws, R,
       f"תרחיש C — רכישת {tech['name_efficient']} | פחת מואץ (עם תמריץ)",
       PatternFill("solid", fgColor="7B6000"))
    R += 1

    sc(ws, R, 2, "* OPEX זהה לתרחיש B — ההבדל הוא מגן המס המואץ בלבד", font=FONT_SM, align=A_R)
    R += 1

    R_C_NET  = R;  R += 1
    R_C_DISC = R;  R += 1
    R_C_CUM  = R;  R += 1

    for rr, lbl in [
        (R_C_NET,  "תזרים נקי — תרחיש C"),
        (R_C_DISC, "תזרים מהוון — תרחיש C"),
        (R_C_CUM,  "NPV מצטבר — תרחיש C"),
    ]:
        sc(ws, rr, 2, lbl, fill=F_SCN_C, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪", align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)

    sc(ws, R_C_NET,  6, f"=F{R_B_INV}",  fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
    sc(ws, R_C_DISC, 6, f"=F{R_C_NET}",  fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
    sc(ws, R_C_CUM,  6, f"=F{R_C_DISC}", fill=F_SCN_C, fmt=FMT_NIS, align=A_C)

    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl  = get_column_letter(col)
        pcl = get_column_letter(col - 1)

        sc(ws, R_C_NET, col, f"={cl}{R_B_EOPC}+{cl}{R_B_OOPC}+{cl}{R_AT_E}",
           fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
        sc(ws, R_C_DISC, col, f"={cl}{R_C_NET}/(1+{DISC_REF})^{i}",
           fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
        sc(ws, R_C_CUM, col, f"={pcl}{R_C_CUM}+{cl}{R_C_DISC}",
           fill=F_SCN_C, fmt=FMT_NIS, align=A_C)

    R += 1

    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=8)
    sc(ws, R, 2, "ד. תוצאות", fill=F_SUBHEAD, font=FONT_SH, align=A_R)
    R += 1

    def result_row(ws, row, label, formula, fmt, fill, unit=None):
        sc(ws, row, 2, label, fill=fill, font=Font(name="Arial", bold=True, size=10), align=A_R)
        if unit:
            sc(ws, row, 3, unit, align=A_R)
        elif "₪" in fmt:
            sc(ws, row, 3, 'ש"ח', align=A_R)
        elif "%" in fmt:
            sc(ws, row, 3, "%", align=A_R)
        else:
            sc(ws, row, 3, "שנים", align=A_R)
        sc(ws, row, 6, formula, fill=fill, fmt=fmt, align=A_C)

    R_NPV_A  = R;  R += 1
    R_NPV_B  = R;  R += 1
    R_NPV_C  = R;  R += 1
    R += 1
    R_DINPV  = R;  R += 1
    R_DPNPV  = R;  R += 1
    R += 1
    R_ROI_A  = R;  R += 1
    R_ROI_B  = R;  R += 1
    R_ROI_C  = R;  R += 1
    R += 1
    R_PBK_D  = R;  R += 1
    R_PBK_E  = R;  R += 1

    YR0 = YR0_LTR

    result_row(ws, R_NPV_A, f"NPV — תרחיש A ({tech['name_baseline']})",
               f"=SUM({YR0}{R_A_DISC}:{MAX_COL_LTR}{R_A_DISC})", FMT_NIS, F_RESULT)
    result_row(ws, R_NPV_B, f"NPV — תרחיש B ({tech['name_efficient']}, ללא תמריץ)",
               f"=SUM({YR0}{R_B_DISC}:{MAX_COL_LTR}{R_B_DISC})", FMT_NIS, F_SCN_B)
    result_row(ws, R_NPV_C, f"NPV — תרחיש C ({tech['name_efficient']}, עם תמריץ)",
               f"=SUM({YR0}{R_C_DISC}:{MAX_COL_LTR}{R_C_DISC})", FMT_NIS, F_SCN_C)

    result_row(ws, R_DINPV, "NPV מצטבר — הצדקת השדרוג (B−A)",
               f"=F{R_NPV_B}-F{R_NPV_A}", FMT_NIS, F_INPUT)
    result_row(ws, R_DPNPV, "ערך התמריץ — תועלת הפחת המואץ (C−B)",
               f"=F{R_NPV_C}-F{R_NPV_B}", FMT_NIS, F_POLICY)

    for rr, row_disc, lbl, fill, cx in [
        (R_ROI_A, R_A_DISC, f"ROI — תרחיש A", F_RESULT, capex_b),
        (R_ROI_B, R_B_DISC, f"ROI — תרחיש B", F_SCN_B, capex_e),
        (R_ROI_C, R_C_DISC, f"ROI — תרחיש C", F_SCN_C, capex_e),
    ]:
        result_row(ws, rr, lbl,
                   f"=IF(ISNUMBER({cx}),"
                   f"(SUM({YR0}{row_disc}:{MAX_COL_LTR}{row_disc})+{cx})/{cx},"
                   f"\"PENDING\")", FMT_PCT, fill)

    sc(ws, R_PBK_D, 2, "תקופת החזר ΔCapEx ללא תמריץ (B−A)",
       fill=F_INPUT, font=Font(name="Arial", bold=True, size=10), align=A_R)
    sc(ws, R_PBK_D, 3, "שנים", align=A_R)
    sc(ws, R_PBK_D, 5, "זמן להחזר ההשקעה הנוספת | מחושב על ΔCF = CF_B − CF_A", font=FONT_SM, align=A_R)
    sc(ws, R_PBK_D, 6, "ראו שורת ΔNPVמצטבר מטה", fill=F_INPUT, font=FONT_SM, align=A_C)

    sc(ws, R_PBK_E, 2, "תקופת החזר ΔCapEx עם תמריץ (C−A)",
       fill=F_POLICY, font=Font(name="Arial", bold=True, size=10, color="FFFFFF"), align=A_R)
    sc(ws, R_PBK_E, 3, "שנים", font=Font(name="Arial", color="FFFFFF"), align=A_R)
    sc(ws, R_PBK_E, 5, "כולל יתרון מגן המס המואץ | מחושב על ΔCF = CF_C − CF_A", font=FONT_SM, align=A_R)
    sc(ws, R_PBK_E, 6, "ראו שורת ΔNPVמצטבר מטה",
       fill=F_POLICY, font=Font(name="Arial", bold=True, size=10, color="FFFFFF"), align=A_C)

    R += 1
    R += 1
    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, "NPV מצטבר — מנקודת מבט ΔCapEx (לחישוב תקופת החזר על השקעה נוספת)",
       fill=F_SUBHEAD, font=FONT_SH, align=A_R)
    R += 1

    year_header_row(ws, R)
    R += 1

    R_DCB = R;  R += 1
    R_DCC = R;  R += 1

    sc(ws, R_DCB, 2, "NPV מצטבר ΔCF — שדרוג ללא תמריץ (B−A)", fill=F_INPUT, font=FONT_N, align=A_R)
    sc(ws, R_DCB, 3, "₪", align=A_R)
    sc(ws, R_DCB, 4, "חישוב", align=A_R)

    sc(ws, R_DCC, 2, "NPV מצטבר ΔCF — שדרוג עם תמריץ (C−A)", fill=F_POLICY, font=FONT_N, align=A_R)
    sc(ws, R_DCC, 3, "₪", align=A_R)
    sc(ws, R_DCC, 4, "חישוב", align=A_R)

    sc(ws, R_DCB, 6, f"=IF(ISNUMBER({dcapex}),-{dcapex},\"PENDING\")",
       fill=F_INPUT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_DCC, 6, f"=IF(ISNUMBER({dcapex}),-{dcapex},\"PENDING\")",
       fill=F_POLICY, fmt=FMT_NIS, align=A_C)

    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl  = get_column_letter(col)
        pcl = get_column_letter(col - 1)

        sc(ws, R_DCB, col, f"={pcl}{R_DCB}+({cl}{R_B_DISC}-{cl}{R_A_DISC})",
           fill=F_INPUT, fmt=FMT_NIS, align=A_C)
        sc(ws, R_DCC, col, f"={pcl}{R_DCC}+({cl}{R_C_DISC}-{cl}{R_A_DISC})",
           fill=F_POLICY, fmt=FMT_NIS, align=A_C)

    ws.cell(row=R_PBK_D, column=6).value = pbk_formula(R_DCB)
    ws.cell(row=R_PBK_D, column=6).fill  = F_INPUT
    ws.cell(row=R_PBK_D, column=6).number_format = FMT_YR

    ws.cell(row=R_PBK_E, column=6).value = pbk_formula(R_DCC)
    ws.cell(row=R_PBK_E, column=6).fill  = F_POLICY
    ws.cell(row=R_PBK_E, column=6).number_format = FMT_YR
    ws.cell(row=R_PBK_E, column=6).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")

    return R + 1, dict(
        npv_a=R_NPV_A, npv_b=R_NPV_B, npv_c=R_NPV_C,
        d_npv=R_DINPV, p_npv=R_DPNPV,
        roi_a=R_ROI_A, roi_b=R_ROI_B, roi_c=R_ROI_C,
        pbk_d=R_PBK_D, pbk_e=R_PBK_E,
    )


def _summary_block(ws, tech_results, R_start):
    R = R_start
    tech_end_col = 6 + MAX_YRS
    section_hdr(ws, R, "סיכום השוואתי לפי טכנולוגיה", number=4, col_end=14)
    R += 1

    sc(ws, R, 2,
       f"* מכפיל פחת = '{GLOBAL_SHEET}'!F22 — שנה שם ותוצאות מתעדכנות אוטומטית | "
       f"OPEX אחר ממתין לנתוני רפי | שאר הנתונים ראו גיליון '{GLOBAL_SHEET}' סעיף 5 למקורות",
       font=FONT_SM, align=A_R)
    R += 1

    hdrs = [
        (2, "טכנולוגיה"), (3, "NPV A (בסיסי)"), (4, "NPV B (יעיל, ללא)"),
        (5, "NPV C (יעיל, עם)"), (6, "ΔNPV B−A"), (7, "ערך תמריץ C−B"),
        (8, "ROI A"), (9, "ROI B"), (10, "ROI C"),
        (11, "החזר ΔCapEx B−A (שנים)"), (12, "החזר ΔCapEx C−A (שנים)"),
    ]
    for col, lbl in hdrs:
        sc(ws, R, col, lbl, fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    R += 1

    for tech, rmap in tech_results:
        sc(ws, R, 2, tech['name_efficient'], font=FONT_N, align=A_R)
        for col, key, fmt, fill in [
            (3,  'npv_a',  FMT_NIS, F_RESULT), (4,  'npv_b',  FMT_NIS, F_SCN_B),
            (5,  'npv_c',  FMT_NIS, F_SCN_C), (6,  'd_npv',  FMT_NIS, F_INPUT),
            (7,  'p_npv',  FMT_NIS, F_POLICY), (8,  'roi_a',  FMT_PCT, F_RESULT),
            (9,  'roi_b',  FMT_PCT, F_SCN_B), (10, 'roi_c',  FMT_PCT, F_SCN_C),
            (11, 'pbk_d',  FMT_YR,  F_INPUT), (12, 'pbk_e',  FMT_YR,  F_POLICY),
        ]:
            sc(ws, R, col, f"=$F${rmap[key]}", fill=fill, fmt=fmt, align=A_C)
        R += 1

    sc(ws, R + 1, 2,
       "** ערכי PENDING מתעדכנים אוטומטית עם קבלת הנתונים (שעות משאבות חום, OPEX אחר מרפי)",
       font=Font(name="Arial", italic=True, color="FF0000", size=9), align=A_R)


# ─── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    tech_refs = build_global_sheet(wb)
    build_analysis_sheet(wb, tech_refs)

    out = "/home/user/shmags-2/projects/energy-program/tax_incentive_model_v2.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print("Sheets:", [ws.title for ws in wb.worksheets])


if __name__ == "__main__":
    main()
