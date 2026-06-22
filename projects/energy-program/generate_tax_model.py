#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tax Incentive Model Generator — v3
National Energy Efficiency Program — EcoTraders

Three-scenario comparison per technology:
  A — Buy baseline (inefficient) technology
  B — Buy efficient technology, standard depreciation
  C — Buy efficient technology, accelerated depreciation (incentive)

Key metrics:
  Incremental NPV (B−A) = economic case for upgrading, regardless of policy
  Policy value  (C−B) = additional benefit from the tax incentive
  Policy payback       = time to recover ΔCapEx under each scenario
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─── COLORS ───────────────────────────────────────────────────────────────────
F_INPUT   = PatternFill("solid", fgColor="FFFF00")   # yellow  – user input
F_CONTROL = PatternFill("solid", fgColor="FFC000")   # orange  – verify/update
F_HEADING = PatternFill("solid", fgColor="4472C4")   # blue    – section header
F_SUBHEAD = PatternFill("solid", fgColor="BDD7EE")   # l.blue  – col headers
F_RESULT  = PatternFill("solid", fgColor="E2EFDA")   # l.green – Scenario A results
F_SCN_B   = PatternFill("solid", fgColor="DDEBF7")   # blue-tint – Scenario B (efficient, std)
F_SCN_C   = PatternFill("solid", fgColor="FFF2CC")   # cream   – Scenario C (efficient, accel)
F_POLICY  = PatternFill("solid", fgColor="FF6B6B")   # red     – key policy row
F_DELTA   = PatternFill("solid", fgColor="E2EFDA")   # green   – incremental / delta rows
F_NONE    = PatternFill("solid", fgColor="FFFFFF")

FONT_H   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
FONT_SH  = Font(name="Arial", bold=True, color="000000", size=10)
FONT_N   = Font(name="Arial", size=10)
FONT_P   = Font(name="Arial", italic=True, color="FF0000", size=10)
FONT_SM  = Font(name="Arial", italic=True, size=9, color="595959")
FONT_B   = Font(name="Arial", bold=True, size=10)

A_R = Alignment(horizontal="right",  vertical="center", readingOrder=2)
A_C = Alignment(horizontal="center", vertical="center", readingOrder=2)

FMT_NIS = '#,##0 ₪'
FMT_PCT = '0.0%'
FMT_NUM = '#,##0'
FMT_YR  = '0.0'
FMT_X   = '0.0"×"'

# ─── SHEET / CELL REFERENCES ──────────────────────────────────────────────────
GLOBAL_SHEET   = "נתונים והנחות"
ANALYSIS_SHEET = "ניתוח"

# Must match row layout in build_global_sheet()
TAX_REF  = f"'{GLOBAL_SHEET}'!$F$10"
DISC_REF = f"'{GLOBAL_SHEET}'!$F$11"
ELEC_REF = f"'{GLOBAL_SHEET}'!$F$12"   # agorot/kWh — divide by 100 for ₪/kWh
MULT_REF = f"'{GLOBAL_SHEET}'!$F$15"   # depreciation multiplier (key policy param)

# ─── TECHNOLOGIES ─────────────────────────────────────────────────────────────
TECHS = [
    dict(
        number="3.1",
        name_efficient="משאבות חום (חימום מים)",
        name_baseline="דוד חשמל קונבנציונלי",
        lifetime_efficient=10,
        lifetime_baseline=15,
        savings_pct=0.50,
        std_depr_pct=0.10,
        std_depr_yrs=10,
        output_unit='kW תרמי',
        notes="החלפת דוד חשמל במשאבת חום | חיסכון אנרגטי ~50% | אורך חיים: 10 vs 15 שנה",
    ),
    dict(
        number="3.2",
        name_efficient="צ'ילרים",
        name_baseline="מערכת קירור קונבנציונלית",
        lifetime_efficient=17,
        lifetime_baseline=15,
        savings_pct=0.20,
        std_depr_pct=0.10,
        std_depr_yrs=10,
        output_unit='TR',
        notes="שדרוג מערכת קירור תעשייתית | חיסכון ~20% | אורך חיים: 17 vs 15 שנה",
    ),
    dict(
        number="3.3",
        name_efficient="מדחסי VSD",
        name_baseline="מדחס מהירות קבועה",
        lifetime_efficient=12,
        lifetime_baseline=12,
        savings_pct=0.20,
        std_depr_pct=0.10,
        std_depr_yrs=10,
        output_unit='מ"ק/דק',
        notes="מדחסי אוויר עם בקרת מהירות משתנה | חיסכון ~20% | אורך חיים: 12 שנה",
    ),
    dict(
        number="3.4",
        name_efficient="מערכות קיטור חשמליות",
        name_baseline="קיטור מבוסס דלק / גז",
        lifetime_efficient=10,
        lifetime_baseline=15,
        savings_pct=0.50,
        std_depr_pct=0.10,
        std_depr_yrs=10,
        output_unit='ק"ג/שעה',
        notes="החלפת קיטור מבוסס דלק בקיטור חשמלי | חיסכון ~50% | אורך חיים: 10 vs 15 שנה",
    ),
]

MAX_YRS     = 20
YR0_COL     = 6                          # column index of "Year 0"
MAX_COL_IDX = YR0_COL + MAX_YRS
MAX_COL_LTR = get_column_letter(MAX_COL_IDX)
YR0_LTR     = get_column_letter(YR0_COL)  # "F"


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def sc(ws, row, col, value=None, fill=None, font=None, align=None, fmt=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if fill:  c.fill = fill
    if font:  c.font = font
    if align: c.alignment = align
    if fmt:   c.number_format = fmt
    return c


def section_hdr(ws, row, label, number=None, col_start=2, col_end=8):
    text = f"{number}. {label}" if number else label
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start)
    c.value = text
    c.fill  = F_HEADING
    c.font  = FONT_H
    c.alignment = A_R


def year_header_row(ws, row, yr0_label="שנה 0"):
    hdrs = [(2, "פרמטר"), (3, "יחידות"), (4, "מקור"),
            (5, "הערות"), (6, yr0_label)]
    for col, val in hdrs:
        sc(ws, row, col, val, fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    for i in range(1, MAX_YRS + 1):
        sc(ws, row, 6 + i, f"שנה {i}", fill=F_SUBHEAD, font=FONT_SH, align=A_C)


def scenario_label_row(ws, row, label, fill, col_end=None):
    end = col_end or (6 + MAX_YRS)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=end)
    c = ws.cell(row=row, column=2)
    c.value = label
    c.fill  = fill
    c.font  = Font(name="Arial", bold=True, size=10,
                   color="FFFFFF" if fill in (F_HEADING, F_POLICY) else "000000")
    c.alignment = A_R


def set_col_widths(ws, include_year_cols=False, extra_tech_cols=0):
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 18
    for i in range(1, extra_tech_cols + 1):
        ws.column_dimensions[get_column_letter(6 + i)].width = 18
    if include_year_cols:
        for i in range(1, MAX_YRS + 1):
            ws.column_dimensions[get_column_letter(6 + i)].width = 13


def color_legend(ws, row=1):
    ws.cell(row=row, column=2).value = "מקרא צבעים"
    ws.cell(row=row, column=2).font = FONT_SH
    items = [
        ("ערך להזנה",                    F_INPUT),
        ("נתון לעדכון / בקרה",           F_CONTROL),
        ("תרחיש A — טכנולוגיה קיימת",    F_RESULT),
        ("תרחיש B — יעיל, ללא תמריץ",    F_SCN_B),
        ("תרחיש C — יעיל, עם תמריץ",     F_SCN_C),
        ("פרמטר מדיניות מרכזי",           F_POLICY),
    ]
    for i, (label, fill) in enumerate(items):
        r = row + 1 + i
        sc(ws, r, 2, label, fill=fill, font=FONT_N, align=A_R)


def pbk_formula(row_cumul):
    """Payback via cumulative NPV crossover — linear interpolation."""
    rng = f"{YR0_LTR}{row_cumul}:{MAX_COL_LTR}{row_cumul}"
    return (
        f"=IFERROR("
        f"MATCH(1,({rng}>0)*1,0)-2+"
        f"ABS(INDEX({rng},MATCH(1,({rng}>0)*1,0)-1))/"
        f"(INDEX({rng},MATCH(1,({rng}>0)*1,0))-"
        f"INDEX({rng},MATCH(1,({rng}>0)*1,0)-1)),"
        f"\"לא מגיע לפירעון\")"
    )


# ─── SHEET 1: GLOBAL ASSUMPTIONS ──────────────────────────────────────────────

def build_global_sheet(wb):
    """Build the assumptions sheet. Returns tech_refs: list of per-tech cell ref dicts."""
    ws = wb.active
    ws.title = GLOBAL_SHEET
    ws.sheet_view.rightToLeft = True
    set_col_widths(ws, extra_tech_cols=len(TECHS))   # F + one col per tech
    color_legend(ws, row=1)

    # ── Section 1: Financial Parameters ──
    section_hdr(ws, 9, "פרמטרים פיננסיים", number=1, col_end=6 + len(TECHS))
    fin_rows = [
        (10, "שיעור מס חברות",           "%",          "רשות המסים",
             "", 0.23, F_CONTROL, FMT_PCT),
        (11, "שיעור היוון (יזמי / פרטי)", "%",          "הנחת עבודה",
             "6% לאומי / 10% תעשייתי — ממתין להחלטת דניאל",
             0.10, F_INPUT, FMT_PCT),
        (12, 'מחיר חשמל ממוצע לתעשייה', 'אג\'/קוט"ש', "חברת החשמל",
             "ממוצע SMP 2024", 14.0, F_INPUT, '#,##0.0'),
    ]
    for r, label, units, source, notes, val, fill, fmt in fin_rows:
        sc(ws, r, 2, label,  fill=fill, font=FONT_N, align=A_R)
        sc(ws, r, 3, units,  align=A_R)
        sc(ws, r, 4, source, align=A_R)
        sc(ws, r, 5, notes,  align=A_R)
        sc(ws, r, 6, val,    fill=fill, fmt=fmt, align=A_C)

    # ── Section 2: Depreciation Multiplier ──
    section_hdr(ws, 14, "פרמטר תמריץ — מכפיל פחת (פחת מואץ)", number=2,
                col_end=6 + len(TECHS))
    sc(ws, 15, 2,
       "מכפיל שיעור הפחת  ←  פרמטר המדיניות המרכזי",
       fill=F_POLICY,
       font=Font(name="Arial", bold=True, size=11, color="FFFFFF"), align=A_R)
    sc(ws, 15, 3, "מכפיל", fill=F_POLICY,
       font=Font(name="Arial", bold=True, size=11, color="FFFFFF"), align=A_C)
    sc(ws, 15, 4, "פרמטר ניתן לשינוי", align=A_R)
    sc(ws, 15, 5,
       "1.0 = ללא תמריץ | 2.0 = קפריסין (כפול, 5 שנים) | 5.0 = 2 שנים",
       fill=F_POLICY,
       font=Font(name="Arial", italic=True, size=9, color="FFFFFF"), align=A_R)
    sc(ws, 15, 6, 2.0,
       fill=F_INPUT,
       font=Font(name="Arial", bold=True, size=14, color="C00000"),
       fmt=FMT_X, align=A_C)

    for r, txt in [
        (16, ("מכפיל 2.0 + פחת סטנדרטי 10% = שיעור מואץ 20%/שנה למשך 5 שנים. "
              "סה\"כ ניכוי = 100% CapEx — היתרון הוא קדמות הניכוי (Time Value of Money).")),
        (17, ("דוגמאות: 1.0 = 10%/שנה × 10 שנים (סטנדרטי) | "
              "2.0 = 20%/שנה × 5 שנים | 3.33 = 33%/שנה × 3 שנים | "
              "5.0 = 50%/שנה × 2 שנים (מודל קפריסין/אירלנד)")),
    ]:
        ws.cell(row=r, column=2).value = txt
        ws.cell(row=r, column=2).font = FONT_SM
        ws.merge_cells(start_row=r, start_column=2,
                       end_row=r, end_column=6 + len(TECHS))

    # ── Section 3: Standard Depreciation Rates ──
    section_hdr(ws, 19, "שיעורי פחת סטנדרטיים — תקנות מס הכנסה", number=3,
                col_end=6 + len(TECHS))
    for r, label, val in [
        (20, "משאבות חום",           0.10),
        (21, "צ'ילרים",              0.10),
        (22, "מדחסי VSD",            0.10),
        (23, "מערכות קיטור",         0.10),
    ]:
        sc(ws, r, 2, label, font=FONT_N, align=A_R)
        sc(ws, r, 3, "% לשנה",  align=A_R)
        sc(ws, r, 4, "תקנות פחת", align=A_R)
        sc(ws, r, 5, "ציוד מכני", align=A_R)
        sc(ws, r, 6, val, fmt=FMT_PCT, align=A_C)

    sc(ws, 25, 2,
       "* שיעורי הפחת הם הנחות עבודה — יש לוודא מול יועץ מס לפני הגשה",
       font=FONT_SM)

    # ── Section 4: Technology Assumptions Table ──
    # One row per parameter, one column per technology (F, G, H, I for 4 techs)
    T_START_COL = 6   # column F = first tech column (same as the single-value cols above)
    T_END_COL   = T_START_COL + len(TECHS) - 1

    section_hdr(ws, 27, "הנחות לפי טכנולוגיה — כל הנתונים לעדכון בטבלה זו", number=4,
                col_end=T_END_COL)

    # Tech name headers
    sc(ws, 28, 2, "פרמטר",   fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    sc(ws, 28, 3, "יחידות",  fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    sc(ws, 28, 4, "מקור",    fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    sc(ws, 28, 5, "הערות",   fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    for i, tech in enumerate(TECHS):
        sc(ws, 28, T_START_COL + i, tech['name_efficient'],
           fill=PatternFill("solid", fgColor="1F4E79"),
           font=Font(name="Arial", bold=True, color="FFFFFF", size=10), align=A_C)
        sc(ws, 29, T_START_COL + i, f"vs. {tech['name_baseline']}",
           font=FONT_SM, align=A_C)

    # Parameter rows — row numbers must stay stable (refs used in analysis sheet)
    # Row 30: CapEx baseline
    # Row 31: CapEx efficient
    # Row 32: ΔCapEx (computed)
    # Row 33: Lifetime baseline
    # Row 34: Lifetime efficient
    # Row 35: kWh/yr baseline
    # Row 36: Energy savings %
    # Row 37: kWh/yr efficient (computed)
    # Row 38: Degradation rate
    # Row 39: Other OPEX baseline (maintenance)
    # Row 40: Other OPEX efficient
    # Row 41: OPEX energy baseline ₪/yr (computed from kWh × price)
    # Row 42: OPEX energy efficient ₪/yr (computed)

    param_meta = [
        (30, "CapEx — ציוד קיים (בסיסי)",              "₪",        "נתוני רפי",    F_INPUT),
        (31, "CapEx — ציוד יעיל",                       "₪",        "נתוני רפי",    F_INPUT),
        (32, "ΔCapEx = יעיל − בסיסי",                   "₪",        "חישוב",         F_NONE),
        (33, "אורך חיים — ציוד בסיסי",                  "שנים",     "ממצאי הנדסה",  F_CONTROL),
        (34, "אורך חיים — ציוד יעיל",                   "שנים",     "ממצאי הנדסה",  F_CONTROL),
        (35, "עלות אנרגיה שנתית — ציוד בסיסי (שנה 1)", "₪/שנה",   "נתוני רפי",    F_INPUT),
        (36, "חיסכון אנרגטי",                           "% מצריכה", "נתוני רפי",    F_INPUT),
        (37, "עלות אנרגיה שנתית — ציוד יעיל (שנה 1)",  "₪/שנה",   "חישוב",         F_NONE),
        (38, "גורם שחיקת ביצועים",                      "%/שנה",    "הנחת עבודה",   F_INPUT),
        (39, "OPEX אחר — בסיסי (תחזוקה, חלפים...)",    "₪/שנה",    "נתוני רפי",    F_INPUT),
        (40, "OPEX אחר — יעיל (תחזוקה, חלפים...)",     "₪/שנה",    "נתוני רפי",    F_INPUT),
        (41, "OPEX אנרגיה שנה 1 — בסיסי",              "₪/שנה",    "חישוב",         F_NONE),
        (42, "OPEX אנרגיה שנה 1 — יעיל",               "₪/שנה",    "חישוב",         F_NONE),
    ]
    for row, label, units, source, fill in param_meta:
        sc(ws, row, 2, label,  font=FONT_N, align=A_R)
        sc(ws, row, 3, units,  align=A_R)
        sc(ws, row, 4, source, align=A_R)

    # Per-technology values
    notes_row = {
        30: "⚠ ממתין לנתוני רפי",
        31: "⚠ ממתין לנתוני רפי",
        35: "⚠ ממתין לנתוני רפי | עלות חשמל שנתית לציוד הקיים (ישירות ב-₪, ללא kWh)",
        39: "⚠ ממתין | תחזוקה, חלפים, ביטוח",
        40: "⚠ ממתין | תחזוקה, חלפים, ביטוח",
    }
    for row, _, _, _, _ in param_meta:
        if row in notes_row:
            sc(ws, row, 5, notes_row[row], font=FONT_P, align=A_R)

    for i, tech in enumerate(TECHS):
        col = T_START_COL + i
        cl  = get_column_letter(col)

        sc(ws, 30, col, "PENDING", fill=F_INPUT,  font=FONT_P, fmt=FMT_NIS, align=A_C)
        sc(ws, 31, col, "PENDING", fill=F_INPUT,  font=FONT_P, fmt=FMT_NIS, align=A_C)
        sc(ws, 32, col,
           f"=IF(ISNUMBER({cl}31)*ISNUMBER({cl}30),{cl}31-{cl}30,\"PENDING\")",
           fmt=FMT_NIS, align=A_C)
        sc(ws, 33, col, tech['lifetime_baseline'],   fill=F_CONTROL, fmt=FMT_YR,  align=A_C)
        sc(ws, 34, col, tech['lifetime_efficient'],  fill=F_CONTROL, fmt=FMT_YR,  align=A_C)
        sc(ws, 35, col, "PENDING", fill=F_INPUT,  font=FONT_P, fmt=FMT_NIS, align=A_C)
        sc(ws, 36, col, tech['savings_pct'],         fill=F_INPUT,   fmt=FMT_PCT, align=A_C)
        sc(ws, 37, col,
           f"=IF(ISNUMBER({cl}35),{cl}35*(1-{cl}36),\"PENDING\")",
           fmt=FMT_NIS, align=A_C)
        sc(ws, 38, col, 0.005, fill=F_INPUT, fmt=FMT_PCT, align=A_C)
        sc(ws, 39, col, "PENDING", fill=F_INPUT, font=FONT_P, fmt=FMT_NIS, align=A_C)
        sc(ws, 40, col, "PENDING", fill=F_INPUT, font=FONT_P, fmt=FMT_NIS, align=A_C)
        sc(ws, 41, col,
           f"=IF(ISNUMBER({cl}35),{cl}35,\"PENDING\")",
           fmt=FMT_NIS, align=A_C)
        sc(ws, 42, col,
           f"=IF(ISNUMBER({cl}35),{cl}35*(1-{cl}36),\"PENDING\")",
           fmt=FMT_NIS, align=A_C)

    # ── Section 5: CapEx per Output (normalization) ──
    ws.merge_cells(start_row=43, start_column=2, end_row=43, end_column=T_END_COL)
    c43 = ws.cell(row=43, column=2)
    c43.value = "5. נירמול CapEx לפי תפוקה — השוואה בין-מיזמית"
    c43.fill  = PatternFill("solid", fgColor="70AD47")
    c43.font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c43.alignment = A_R

    norm_meta = [
        (44, "הספק מותקן / תפוקה נומינלית", "יח' תפוקה", "נתוני רפי",
             "kW תרמי (משאבות) | TR (צ'ילרים) | מ\"ק/דק (מדחסים) | ק\"ג/שעה (קיטור)",
             F_INPUT),
        (45, "CapEx/תפוקה — ציוד בסיסי",    "₪/יח'",     "חישוב",
             "= CapEx בסיסי ÷ תפוקה מותקנת",
             F_NONE),
        (46, "CapEx/תפוקה — ציוד יעיל",     "₪/יח'",     "חישוב",
             "= CapEx יעיל ÷ תפוקה מותקנת",
             F_NONE),
        (47, "ΔCapEx/תפוקה (יעיל − בסיסי)", "₪/יח'",     "חישוב",
             "= ΔCapEx ÷ תפוקה | עלות נוספת לנירמול | מאפשר השוואה בין מיזמים בגדלים שונים",
             F_NONE),
    ]
    for row, label, units, source, notes, fill in norm_meta:
        sc(ws, row, 2, label,  font=FONT_N if fill == F_NONE else FONT_P, align=A_R)
        sc(ws, row, 3, units,  align=A_R)
        sc(ws, row, 4, source, align=A_R)
        sc(ws, row, 5, notes,  font=FONT_SM, align=A_R)
        if fill != F_NONE:
            ws.cell(row=row, column=2).fill = fill

    for i, tech in enumerate(TECHS):
        col = T_START_COL + i
        cl  = get_column_letter(col)
        sc(ws, 44, col, "PENDING", fill=F_INPUT, font=FONT_P, fmt=FMT_NUM, align=A_C)
        sc(ws, 44, col).comment = None   # placeholder; unit shown in col E note
        sc(ws, 45, col,
           f"=IF(ISNUMBER({cl}30)*ISNUMBER({cl}44),{cl}30/{cl}44,\"PENDING\")",
           fmt=FMT_NIS, align=A_C)
        sc(ws, 46, col,
           f"=IF(ISNUMBER({cl}31)*ISNUMBER({cl}44),{cl}31/{cl}44,\"PENDING\")",
           fmt=FMT_NIS, align=A_C)
        sc(ws, 47, col,
           f"=IF(ISNUMBER({cl}32)*ISNUMBER({cl}44),{cl}32/{cl}44,\"PENDING\")",
           fmt=FMT_NIS, align=A_C)

    sc(ws, 49, 2,
       "* CapEx, עלות אנרגיה שנתית בסיסית, תפוקה מותקנת ו-OPEX אחר — ממתינים לנתוני רפי. "
       "% חיסכון אנרגטי אושר על ידי רפי. שאר הנתונים הם הנחות עבודה לאישור.",
       font=FONT_SM)

    # Build and return cell reference dict for each tech (used by analysis sheet)
    tech_refs = []
    for i, tech in enumerate(TECHS):
        cl = get_column_letter(T_START_COL + i)
        s  = GLOBAL_SHEET
        tech_refs.append({
            'capex_b':         f"'{s}'!${cl}$30",
            'capex_e':         f"'{s}'!${cl}$31",
            'dcapex':          f"'{s}'!${cl}$32",
            'life_b':          f"'{s}'!${cl}$33",
            'life_e':          f"'{s}'!${cl}$34",
            'energy_b':        f"'{s}'!${cl}$35",   # annual energy cost, baseline (₪/yr)
            'svpct':           f"'{s}'!${cl}$36",
            'energy_e':        f"'{s}'!${cl}$37",   # annual energy cost, efficient (₪/yr)
            'degrad':          f"'{s}'!${cl}$38",
            'opex_o_b':        f"'{s}'!${cl}$39",
            'opex_o_e':        f"'{s}'!${cl}$40",
            'opex_e_b':        f"'{s}'!${cl}$41",
            'opex_e_e':        f"'{s}'!${cl}$42",
            'output_cap':      f"'{s}'!${cl}$44",   # rated output capacity
            'capex_b_per_out': f"'{s}'!${cl}$45",   # ₪ per output unit, baseline
            'capex_e_per_out': f"'{s}'!${cl}$46",   # ₪ per output unit, efficient
            'dcapex_per_out':  f"'{s}'!${cl}$47",   # ₪ per output unit, delta
        })
    return tech_refs


# ─── SHEET 2: ANALYSIS ────────────────────────────────────────────────────────

def build_analysis_sheet(wb, tech_refs):
    ws = wb.create_sheet(ANALYSIS_SHEET)
    ws.sheet_view.rightToLeft = True
    set_col_widths(ws, include_year_cols=True)
    color_legend(ws, row=1)

    tech_end_col = 6 + MAX_YRS

    # Title
    ws.merge_cells(start_row=8, start_column=2,
                   end_row=8, end_column=tech_end_col)
    sc(ws, 8, 2, "ניתוח כלכלי — פחת מואץ לפי טכנולוגיה | השוואת שלושה תרחישים",
       fill=F_HEADING,
       font=Font(name="Arial", bold=True, color="FFFFFF", size=13), align=A_R)

    # Section 1: Financial Parameters (display, linked from Sheet 1)
    section_hdr(ws, 10, "פרמטרים פיננסיים", number=1)
    for r, label, units, ref, fmt, fill in [
        (11, "שיעור מס חברות",     "%",          f"={TAX_REF}",        FMT_PCT,    F_CONTROL),
        (12, "שיעור היוון",         "%",          f"={DISC_REF}",       FMT_PCT,    F_INPUT),
        (13, 'מחיר חשמל (₪/קוט"ש)', '₪/קוט"ש',  f"={ELEC_REF}/100",   '#,##0.000', F_INPUT),
    ]:
        sc(ws, r, 2, label,        fill=fill, font=FONT_N, align=A_R)
        sc(ws, r, 3, units,        align=A_R)
        sc(ws, r, 4, GLOBAL_SHEET, align=A_R)
        sc(ws, r, 6, ref,          fill=fill, fmt=fmt, align=A_C)

    # Section 2: Policy Parameter (display, linked from Sheet 1)
    section_hdr(ws, 15, "פרמטר תמריץ — מכפיל פחת", number=2)
    sc(ws, 16, 2,
       "מכפיל שיעור הפחת  ←  שנה ב'נתונים והנחות'!F15",
       fill=F_POLICY,
       font=Font(name="Arial", bold=True, size=11, color="FFFFFF"), align=A_R)
    sc(ws, 16, 3, "מכפיל", fill=F_POLICY,
       font=Font(name="Arial", bold=True, size=11, color="FFFFFF"), align=A_C)
    sc(ws, 16, 4, GLOBAL_SHEET, align=A_R)
    sc(ws, 16, 6, f"={MULT_REF}",
       fill=F_INPUT,
       font=Font(name="Arial", bold=True, size=13, color="C00000"),
       fmt=FMT_X, align=A_C)
    sc(ws, 17, 2,
       "מכפיל × פחת סטנדרטי = שיעור מואץ | סה\"כ ניכוי = 100% CapEx | "
       "תקופת פחת = 1 ÷ שיעור מואץ",
       font=FONT_SM, align=A_R)

    # Section 3: Technology Analyses
    section_hdr(ws, 19, "ניתוחים לפי טכנולוגיה", number=3,
                col_end=tech_end_col)

    R = 21
    tech_results = []
    for idx, tech in enumerate(TECHS):
        R, rmap = _tech_block(ws, tech, R, tech_refs[idx])
        tech_results.append((tech, rmap))
        R += 2  # separator

    # Section 4: Summary
    _summary_block(ws, tech_results, R)


def _tech_block(ws, tech, R_start, refs):
    """refs: dict of cell references pointing into the global assumptions table (Sheet 1)."""
    R = R_start
    tech_end_col = 6 + MAX_YRS

    # ── Header ────────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=R, start_column=2,
                   end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, f"{tech['number']}  {tech['name_efficient']}",
       fill=PatternFill("solid", fgColor="1F4E79"),
       font=Font(name="Arial", bold=True, color="FFFFFF", size=11), align=A_R)
    R += 1
    ws.merge_cells(start_row=R, start_column=2,
                   end_row=R, end_column=tech_end_col)
    sc(ws, R, 2,
       f"* {tech['notes']} | כל הנתונים: גיליון '{GLOBAL_SHEET}' — סעיף 4",
       font=FONT_SM, align=A_R)
    R += 1

    std_r   = tech['std_depr_pct']
    std_yrs = tech['std_depr_yrs']
    life_e  = tech['lifetime_efficient']
    life_b  = tech['lifetime_baseline']

    # All inputs reference the consolidated assumptions table on Sheet 1
    capex_b  = refs['capex_b']
    capex_e  = refs['capex_e']
    opex_e_b = refs['opex_e_b']   # energy OPEX baseline (₪/yr, year 1)
    opex_e_e = refs['opex_e_e']   # energy OPEX efficient (₪/yr, year 1)
    opex_o_b = refs['opex_o_b']   # other OPEX baseline
    opex_o_e = refs['opex_o_e']   # other OPEX efficient
    degrad   = refs['degrad']
    dcapex   = refs['dcapex']

    # ── B: Depreciation Schedules ──────────────────────────────────────────────
    ws.merge_cells(start_row=R, start_column=2,
                   end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, "ב. לוחות פחת (ציוד בסיסי vs. יעיל)",
       fill=F_SUBHEAD, font=FONT_SH, align=A_R)
    R += 1

    year_header_row(ws, R, yr0_label="שנה 0 (השקעה)")
    R += 1

    R_SD_B = R;  R += 1   # baseline std depreciation
    R_ST_B = R;  R += 1   # baseline std tax shield
    R_SD_E = R;  R += 1   # efficient std depreciation
    R_ST_E = R;  R += 1   # efficient std tax shield  (Scenario B)
    R_AD_E = R;  R += 1   # efficient accel depreciation
    R_AT_E = R;  R += 1   # efficient accel tax shield (Scenario C)

    depr_rows = [
        (R_SD_B, f"פחת סטנדרטי — {tech['name_baseline']}",   F_RESULT),
        (R_ST_B, f"מגן מס — {tech['name_baseline']}",         F_RESULT),
        (R_SD_E, f"פחת סטנדרטי — {tech['name_efficient']}",   F_SCN_B),
        (R_ST_E, f"מגן מס סטנדרטי — {tech['name_efficient']}", F_SCN_B),
        (R_AD_E, f"פחת מואץ — {tech['name_efficient']}",       F_SCN_C),
        (R_AT_E, f"מגן מס מואץ — {tech['name_efficient']}",    F_SCN_C),
    ]
    for rr, lbl, fill in depr_rows:
        sc(ws, rr, 2, lbl, fill=fill, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪",   align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)
        sc(ws, rr, 6, 0, fill=fill, fmt=FMT_NIS, align=A_C)

    sc(ws, R_SD_B, 5, f"קו ישר | {std_r:.0%}/שנה | עד {std_yrs} שנים | CapEx בסיסי",  align=A_R)
    sc(ws, R_ST_B, 5, "= פחת × שיעור מס", align=A_R)
    sc(ws, R_SD_E, 5, f"קו ישר | {std_r:.0%}/שנה | עד {std_yrs} שנים | CapEx יעיל",   align=A_R)
    sc(ws, R_ST_E, 5, "= פחת × שיעור מס", align=A_R)
    sc(ws, R_AD_E, 5, "= CapEx יעיל × (פחת סטנדרטי × מכפיל) | תקופה = 1 ÷ שיעור מואץ", align=A_R)
    sc(ws, R_AT_E, 5, "= פחת מואץ × שיעור מס", align=A_R)

    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl  = get_column_letter(col)

        # Baseline std depreciation (on baseline CapEx)
        sd_b = (f"=IF(ISNUMBER({capex_b}),{capex_b}*{std_r},0)"
                if i <= std_yrs else "=0")
        sc(ws, R_SD_B, col, sd_b, fill=F_RESULT, fmt=FMT_NIS, align=A_C)
        sc(ws, R_ST_B, col, f"={cl}{R_SD_B}*{TAX_REF}",
           fill=F_RESULT, fmt=FMT_NIS, align=A_C)

        # Efficient std depreciation (on efficient CapEx)
        sd_e = (f"=IF(ISNUMBER({capex_e}),{capex_e}*{std_r},0)"
                if i <= std_yrs else "=0")
        sc(ws, R_SD_E, col, sd_e, fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
        sc(ws, R_ST_E, col, f"={cl}{R_SD_E}*{TAX_REF}",
           fill=F_SCN_B, fmt=FMT_NIS, align=A_C)

        # Efficient accel depreciation — multiplier model
        ad_e = (
            f"=IF(ISNUMBER({capex_e}),"
            f"IF({i}<=ROUND(1/({std_r}*{MULT_REF}),0),"
            f"{capex_e}*{std_r}*{MULT_REF},0),0)"
        )
        sc(ws, R_AD_E, col, ad_e, fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
        sc(ws, R_AT_E, col, f"={cl}{R_AD_E}*{TAX_REF}",
           fill=F_SCN_C, fmt=FMT_NIS, align=A_C)

    R += 1  # blank row

    # ── C: Cash Flow Analysis — 3 Scenarios ───────────────────────────────────
    ws.merge_cells(start_row=R, start_column=2,
                   end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, "ג. ניתוח תזרים מזומנים — שלושה תרחישים",
       fill=F_SUBHEAD, font=FONT_SH, align=A_R)
    R += 1

    year_header_row(ws, R, yr0_label="שנה 0 (השקעה)")
    R += 1

    # ── Scenario A: Baseline technology ──
    scenario_label_row(ws, R,
       f"תרחיש A — רכישת {tech['name_baseline']} (ציוד קיים / לא יעיל)",
       PatternFill("solid", fgColor="538135"))
    R += 1

    R_A_INV  = R;  R += 1   # investment
    R_A_EOPC = R;  R += 1   # energy OPEX (baseline, flat no degradation on inefficient)
    R_A_OOPC = R;  R += 1   # other OPEX
    R_A_NET  = R;  R += 1   # net CF (investment + OPEX + tax shield)
    R_A_DISC = R;  R += 1   # discounted CF
    R_A_CUM  = R;  R += 1   # cumulative NPV

    for rr, lbl, fill in [
        (R_A_INV,  "השקעה ראשונית",            F_RESULT),
        (R_A_EOPC, "OPEX אנרגיה",              F_RESULT),
        (R_A_OOPC, "OPEX אחר (תחזוקה וכו')",   F_RESULT),
        (R_A_NET,  "תזרים נקי — תרחיש A",      F_RESULT),
        (R_A_DISC, "תזרים מהוון — תרחיש A",    F_RESULT),
        (R_A_CUM,  "NPV מצטבר — תרחיש A",      F_RESULT),
    ]:
        sc(ws, rr, 2, lbl, fill=fill, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪",  align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)

    sc(ws, R_A_INV,  6, f"=IF(ISNUMBER({capex_b}),-{capex_b},\"PENDING\")",
       fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_EOPC, 6, 0, fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_OOPC, 6, 0, fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_NET,  6, f"=F{R_A_INV}",   fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_DISC, 6, f"=F{R_A_NET}",   fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_CUM,  6, f"=F{R_A_DISC}",  fill=F_RESULT, fmt=FMT_NIS, align=A_C)

    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl  = get_column_letter(col)
        pcl = get_column_letter(col - 1)

        # Scenario A: energy OPEX on baseline (no degradation — inefficient tech is flat)
        eopc_a = (f"=IF(ISNUMBER({opex_e_b}),-{opex_e_b},0)"
                  if i <= life_b else "=0")
        sc(ws, R_A_EOPC, col, eopc_a, fill=F_RESULT, fmt=FMT_NIS, align=A_C)

        # Other OPEX baseline
        oopc_a = (f"=IF(ISNUMBER({opex_o_b}),-{opex_o_b},0)"
                  if i <= life_b else "=0")
        sc(ws, R_A_OOPC, col, oopc_a, fill=F_RESULT, fmt=FMT_NIS, align=A_C)

        # Net CF A = OPEX energy + OPEX other + tax shield (income from depreciation shield)
        net_a = f"={cl}{R_A_EOPC}+{cl}{R_A_OOPC}+{cl}{R_ST_B}"
        sc(ws, R_A_NET, col, net_a, fill=F_RESULT, fmt=FMT_NIS, align=A_C)

        sc(ws, R_A_DISC, col,
           f"={cl}{R_A_NET}/(1+{DISC_REF})^{i}",
           fill=F_RESULT, fmt=FMT_NIS, align=A_C)
        sc(ws, R_A_CUM, col,
           f"={pcl}{R_A_CUM}+{cl}{R_A_DISC}",
           fill=F_RESULT, fmt=FMT_NIS, align=A_C)

    R += 1

    # ── Scenario B: Efficient, standard depreciation ──
    scenario_label_row(ws, R,
       f"תרחיש B — רכישת {tech['name_efficient']} | פחת סטנדרטי (ללא תמריץ)",
       PatternFill("solid", fgColor="2E75B6"))
    R += 1

    R_B_INV  = R;  R += 1
    R_B_EOPC = R;  R += 1
    R_B_OOPC = R;  R += 1
    R_B_NET  = R;  R += 1
    R_B_DISC = R;  R += 1
    R_B_CUM  = R;  R += 1

    for rr, lbl, fill in [
        (R_B_INV,  "השקעה ראשונית",                  F_SCN_B),
        (R_B_EOPC, "OPEX אנרגיה (עם שחיקה שנתית)",   F_SCN_B),
        (R_B_OOPC, "OPEX אחר (תחזוקה וכו')",          F_SCN_B),
        (R_B_NET,  "תזרים נקי — תרחיש B",             F_SCN_B),
        (R_B_DISC, "תזרים מהוון — תרחיש B",           F_SCN_B),
        (R_B_CUM,  "NPV מצטבר — תרחיש B",             F_SCN_B),
    ]:
        sc(ws, rr, 2, lbl, fill=fill, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪", align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)

    sc(ws, R_B_INV,  6, f"=IF(ISNUMBER({capex_e}),-{capex_e},\"PENDING\")",
       fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_EOPC, 6, 0, fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_OOPC, 6, 0, fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_NET,  6, f"=F{R_B_INV}",  fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_DISC, 6, f"=F{R_B_NET}",  fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_CUM,  6, f"=F{R_B_DISC}", fill=F_SCN_B, fmt=FMT_NIS, align=A_C)

    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl  = get_column_letter(col)
        pcl = get_column_letter(col - 1)

        # Energy OPEX efficient — degraded: base × (1−degrad)^(yr−1)
        eopc_b = (f"=IF(ISNUMBER({opex_e_e}),-{opex_e_e}*(1-{degrad})^{i-1},0)"
                  if i <= life_e else "=0")
        sc(ws, R_B_EOPC, col, eopc_b, fill=F_SCN_B, fmt=FMT_NIS, align=A_C)

        oopc_b = (f"=IF(ISNUMBER({opex_o_e}),-{opex_o_e},0)"
                  if i <= life_e else "=0")
        sc(ws, R_B_OOPC, col, oopc_b, fill=F_SCN_B, fmt=FMT_NIS, align=A_C)

        sc(ws, R_B_NET, col,
           f"={cl}{R_B_EOPC}+{cl}{R_B_OOPC}+{cl}{R_ST_E}",
           fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
        sc(ws, R_B_DISC, col,
           f"={cl}{R_B_NET}/(1+{DISC_REF})^{i}",
           fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
        sc(ws, R_B_CUM, col,
           f"={pcl}{R_B_CUM}+{cl}{R_B_DISC}",
           fill=F_SCN_B, fmt=FMT_NIS, align=A_C)

    R += 1

    # ── Scenario C: Efficient + accelerated depreciation ──
    scenario_label_row(ws, R,
       f"תרחיש C — רכישת {tech['name_efficient']} | פחת מואץ (עם תמריץ)",
       PatternFill("solid", fgColor="7B6000"))
    R += 1

    sc(ws, R, 2, "* OPEX זהה לתרחיש B — ההבדל הוא מגן המס המואץ בלבד",
       font=FONT_SM, align=A_R)
    R += 1

    R_C_NET  = R;  R += 1
    R_C_DISC = R;  R += 1
    R_C_CUM  = R;  R += 1

    for rr, lbl in [
        (R_C_NET,  "תזרים נקי — תרחיש C"),
        (R_C_DISC, "תזרים מהוון — תרחיש C"),
        (R_C_CUM,  "NPV מצטבר — תרחיש C"),
    ]:
        sc(ws, rr, 2, lbl, fill=F_SCN_C, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪", align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)

    # Year 0: same investment as B, same OPEX rows as B, different tax shield
    sc(ws, R_C_NET,  6, f"=F{R_B_INV}",  fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
    sc(ws, R_C_DISC, 6, f"=F{R_C_NET}",  fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
    sc(ws, R_C_CUM,  6, f"=F{R_C_DISC}", fill=F_SCN_C, fmt=FMT_NIS, align=A_C)

    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl  = get_column_letter(col)
        pcl = get_column_letter(col - 1)

        # Net CF C = same OPEX as B + accel tax shield instead of std
        sc(ws, R_C_NET, col,
           f"={cl}{R_B_EOPC}+{cl}{R_B_OOPC}+{cl}{R_AT_E}",
           fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
        sc(ws, R_C_DISC, col,
           f"={cl}{R_C_NET}/(1+{DISC_REF})^{i}",
           fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
        sc(ws, R_C_CUM, col,
           f"={pcl}{R_C_CUM}+{cl}{R_C_DISC}",
           fill=F_SCN_C, fmt=FMT_NIS, align=A_C)

    R += 1

    # ── D: Results ─────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=8)
    sc(ws, R, 2, "ד. תוצאות", fill=F_SUBHEAD, font=FONT_SH, align=A_R)
    R += 1

    def result_row(ws, row, label, formula, fmt, fill, unit=None):
        sc(ws, row, 2, label, fill=fill,
           font=Font(name="Arial", bold=True, size=10), align=A_R)
        if unit:
            sc(ws, row, 3, unit, align=A_R)
        elif "₪" in fmt:
            sc(ws, row, 3, 'ש"ח', align=A_R)
        elif "%" in fmt:
            sc(ws, row, 3, "%", align=A_R)
        else:
            sc(ws, row, 3, "שנים", align=A_R)
        sc(ws, row, 6, formula, fill=fill, fmt=fmt, align=A_C)

    R_NPV_A  = R;  R += 1   # NPV Scenario A
    R_NPV_B  = R;  R += 1   # NPV Scenario B
    R_NPV_C  = R;  R += 1   # NPV Scenario C
    R += 1
    R_DINPV  = R;  R += 1   # Incremental NPV (B−A): case for upgrading
    R_DPNPV  = R;  R += 1   # Policy NPV (C−B): value of incentive
    R += 1
    R_ROI_A  = R;  R += 1
    R_ROI_B  = R;  R += 1
    R_ROI_C  = R;  R += 1
    R += 1
    R_PBK_D  = R;  R += 1   # payback ΔCapEx under B (incremental investment payback)
    R_PBK_E  = R;  R += 1   # payback ΔCapEx under C

    YR0 = YR0_LTR  # "F"

    # NPV = sum of discounted CFs (year 0 through year MAX_YRS)
    result_row(ws, R_NPV_A, f"NPV — תרחיש A ({tech['name_baseline']})",
               f"=SUM({YR0}{R_A_DISC}:{MAX_COL_LTR}{R_A_DISC})",
               FMT_NIS, F_RESULT)
    result_row(ws, R_NPV_B, f"NPV — תרחיש B ({tech['name_efficient']}, ללא תמריץ)",
               f"=SUM({YR0}{R_B_DISC}:{MAX_COL_LTR}{R_B_DISC})",
               FMT_NIS, F_SCN_B)
    result_row(ws, R_NPV_C, f"NPV — תרחיש C ({tech['name_efficient']}, עם תמריץ)",
               f"=SUM({YR0}{R_C_DISC}:{MAX_COL_LTR}{R_C_DISC})",
               FMT_NIS, F_SCN_C)

    # Incremental NPV (B−A): is it worth upgrading?
    result_row(ws, R_DINPV, "NPV מצטבר — הצדקת השדרוג (B−A)",
               f"=F{R_NPV_B}-F{R_NPV_A}",
               FMT_NIS, F_INPUT)
    # Policy value (C−B): what does the incentive add?
    result_row(ws, R_DPNPV, "ערך התמריץ — תועלת הפחת המואץ (C−B)",
               f"=F{R_NPV_C}-F{R_NPV_B}",
               FMT_NIS, F_POLICY)

    for rr, row_disc, lbl, fill, cx in [
        (R_ROI_A, R_A_DISC, f"ROI — תרחיש A", F_RESULT,  capex_b),
        (R_ROI_B, R_B_DISC, f"ROI — תרחיש B", F_SCN_B, capex_e),
        (R_ROI_C, R_C_DISC, f"ROI — תרחיש C", F_SCN_C, capex_e),
    ]:
        result_row(ws, rr, lbl,
                   f"=IF(ISNUMBER({cx}),"
                   f"(SUM({YR0}{row_disc}:{MAX_COL_LTR}{row_disc})+{cx})/{cx},"
                   f"\"PENDING\")",
                   FMT_PCT, fill)

    # Incremental payback: time to recover ΔCapEx from the differential cash flows (B−A)
    # We need a ΔCF cumulative row — build it inline using a note
    sc(ws, R_PBK_D, 2, "תקופת החזר ΔCapEx ללא תמריץ (B−A)",
       fill=F_INPUT, font=Font(name="Arial", bold=True, size=10), align=A_R)
    sc(ws, R_PBK_D, 3, "שנים", align=A_R)
    sc(ws, R_PBK_D, 5,
       "זמן להחזר ההשקעה הנוספת | מחושב על ΔCF = CF_B − CF_A",
       font=FONT_SM, align=A_R)
    sc(ws, R_PBK_D, 6, "ראו שורת ΔNPVמצטבר מטה",
       fill=F_INPUT, font=FONT_SM, align=A_C)

    sc(ws, R_PBK_E, 2, "תקופת החזר ΔCapEx עם תמריץ (C−A)",
       fill=F_POLICY, font=Font(name="Arial", bold=True, size=10, color="FFFFFF"), align=A_R)
    sc(ws, R_PBK_E, 3, "שנים",
       font=Font(name="Arial", color="FFFFFF"), align=A_R)
    sc(ws, R_PBK_E, 5,
       "כולל יתרון מגן המס המואץ | מחושב על ΔCF = CF_C − CF_A",
       font=FONT_SM, align=A_R)
    sc(ws, R_PBK_E, 6, "ראו שורת ΔNPVמצטבר מטה",
       fill=F_POLICY, font=Font(name="Arial", bold=True, size=10, color="FFFFFF"), align=A_C)

    # ΔCumulative NPV rows (B−A and C−A) for the incremental payback
    R += 1
    R += 1   # blank
    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, "NPV מצטבר — מנקודת מבט ΔCapEx (לחישוב תקופת החזר על השקעה נוספת)",
       fill=F_SUBHEAD, font=FONT_SH, align=A_R)
    R += 1

    year_header_row(ws, R)
    R += 1

    R_DCB = R;  R += 1   # ΔCumulative B−A
    R_DCC = R;  R += 1   # ΔCumulative C−A

    sc(ws, R_DCB, 2, "NPV מצטבר ΔCF — שדרוג ללא תמריץ (B−A)",
       fill=F_INPUT, font=FONT_N, align=A_R)
    sc(ws, R_DCB, 3, "₪", align=A_R)
    sc(ws, R_DCB, 4, "חישוב", align=A_R)

    sc(ws, R_DCC, 2, "NPV מצטבר ΔCF — שדרוג עם תמריץ (C−A)",
       fill=F_POLICY, font=FONT_N, align=A_R)
    sc(ws, R_DCC, 3, "₪", align=A_R)
    sc(ws, R_DCC, 4, "חישוב", align=A_R)

    # Year 0: −ΔCapEx (the extra investment for upgrading)
    sc(ws, R_DCB, 6,
       f"=IF(ISNUMBER({dcapex}),-{dcapex},\"PENDING\")",
       fill=F_INPUT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_DCC, 6,
       f"=IF(ISNUMBER({dcapex}),-{dcapex},\"PENDING\")",
       fill=F_POLICY, fmt=FMT_NIS, align=A_C)

    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl  = get_column_letter(col)
        pcl = get_column_letter(col - 1)

        # ΔCF_B = discounted B − discounted A
        sc(ws, R_DCB, col,
           f"={pcl}{R_DCB}+({cl}{R_B_DISC}-{cl}{R_A_DISC})",
           fill=F_INPUT, fmt=FMT_NIS, align=A_C)

        # ΔCF_C = discounted C − discounted A
        sc(ws, R_DCC, col,
           f"={pcl}{R_DCC}+({cl}{R_C_DISC}-{cl}{R_A_DISC})",
           fill=F_POLICY, fmt=FMT_NIS, align=A_C)

    # Now update the incremental payback cells to use the ΔCumulative rows
    ws.cell(row=R_PBK_D, column=6).value = pbk_formula(R_DCB)
    ws.cell(row=R_PBK_D, column=6).fill  = F_INPUT
    ws.cell(row=R_PBK_D, column=6).number_format = FMT_YR

    ws.cell(row=R_PBK_E, column=6).value = pbk_formula(R_DCC)
    ws.cell(row=R_PBK_E, column=6).fill  = F_POLICY
    ws.cell(row=R_PBK_E, column=6).number_format = FMT_YR
    ws.cell(row=R_PBK_E, column=6).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")

    return R + 1, dict(
        npv_a=R_NPV_A, npv_b=R_NPV_B, npv_c=R_NPV_C,
        d_npv=R_DINPV, p_npv=R_DPNPV,
        roi_a=R_ROI_A, roi_b=R_ROI_B, roi_c=R_ROI_C,
        pbk_d=R_PBK_D, pbk_e=R_PBK_E,
    )


def _summary_block(ws, tech_results, R_start):
    R = R_start
    tech_end_col = 6 + MAX_YRS
    section_hdr(ws, R, "סיכום השוואתי לפי טכנולוגיה", number=4, col_end=17)
    R += 1

    sc(ws, R, 2,
       f"* מכפיל פחת = '{GLOBAL_SHEET}'!F15 — שנה שם ותוצאות מתעדכנות אוטומטית | "
       f"CapEx, עלות אנרגיה ותפוקה ממתינים לנתוני רפי",
       font=FONT_SM, align=A_R)
    R += 1

    hdrs = [
        (2,  "טכנולוגיה"),
        (3,  "NPV A (בסיסי)"),
        (4,  "NPV B (יעיל, ללא)"),
        (5,  "NPV C (יעיל, עם)"),
        (6,  "ΔNPV B−A"),
        (7,  "ערך תמריץ C−B"),
        (8,  "ROI A"),
        (9,  "ROI B"),
        (10, "ROI C"),
        (11, "החזר ΔCapEx B−A (שנים)"),
        (12, "החזר ΔCapEx C−A (שנים)"),
        (13, "CapEx בסיסי/תפוקה (₪/יח')"),
        (14, "CapEx יעיל/תפוקה (₪/יח')"),
        (15, "ΔCapEx/תפוקה (₪/יח')"),
    ]
    for col, lbl in hdrs:
        sc(ws, R, col, lbl, fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    R += 1

    T_START_COL = 6   # mirrors global sheet layout
    for idx, (tech, rmap) in enumerate(tech_results):
        sc(ws, R, 2, tech['name_efficient'], font=FONT_N, align=A_R)
        for col, key, fmt, fill in [
            (3,  'npv_a',  FMT_NIS, F_RESULT),
            (4,  'npv_b',  FMT_NIS, F_SCN_B),
            (5,  'npv_c',  FMT_NIS, F_SCN_C),
            (6,  'd_npv',  FMT_NIS, F_INPUT),
            (7,  'p_npv',  FMT_NIS, F_POLICY),
            (8,  'roi_a',  FMT_PCT, F_RESULT),
            (9,  'roi_b',  FMT_PCT, F_SCN_B),
            (10, 'roi_c',  FMT_PCT, F_SCN_C),
            (11, 'pbk_d',  FMT_YR,  F_INPUT),
            (12, 'pbk_e',  FMT_YR,  F_POLICY),
        ]:
            sc(ws, R, col, f"=$F${rmap[key]}",
               fill=fill, fmt=fmt, align=A_C)
        # CapEx per output — reference global sheet rows 45-47 directly
        gl_cl = get_column_letter(T_START_COL + idx)
        for col, row_num, fill in [
            (13, 45, F_RESULT),
            (14, 46, F_SCN_B),
            (15, 47, F_INPUT),
        ]:
            sc(ws, R, col,
               f"=IF(ISNUMBER('{GLOBAL_SHEET}'!${gl_cl}${row_num}),"
               f"'{GLOBAL_SHEET}'!${gl_cl}${row_num},\"PENDING\")",
               fill=fill, fmt=FMT_NIS, align=A_C)
        R += 1

    sc(ws, R + 1, 2,
       "** כל ערכי PENDING מתעדכנים אוטומטית עם קבלת נתוני רפי (CapEx, עלות אנרגיה, תפוקה מותקנת, OPEX אחר)",
       font=Font(name="Arial", italic=True, color="FF0000", size=9), align=A_R)


# ─── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    tech_refs = build_global_sheet(wb)
    build_analysis_sheet(wb, tech_refs)

    out = "/home/user/shmags-2/projects/energy-program/tax_incentive_model.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print("Sheets:", [ws.title for ws in wb.worksheets])


if __name__ == "__main__":
    main()
