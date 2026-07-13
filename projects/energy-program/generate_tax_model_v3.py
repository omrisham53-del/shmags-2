#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Tax Incentive Model Generator — v5 (2026-07-13)
National Energy Efficiency Program — EcoTraders

Rebuilt fresh from Omri's manually-edited v2 file (NOT from the old v2
script), incorporating his manual changes + new structural requests.

Manual changes preserved from Omri's edited file:
  - Compact global-sheet row layout
  - Diesel (סולר) as the default baseline fuel type
  - Mazut price ₪2,344.72/ton (corrected back from a decimal-shifted value)
  - "(בתוספת מרווח שיווק)" fuel-price labels + margin note (Omri applies the
    0.5 margin himself)

New structural changes this version:
  1. Two capacity points per technology (low + high) — 6 tech columns total,
     not 3. Each drives its own full analysis block.
  2. Real per-technology units shown on every value cell (via number-format
     suffixes: kW / RT / kW/ton / kW/100cfm), so units are unambiguous.
  3. Sources written directly in the מקור column, not as [n] citations.
     A short מתודולוגיה section at the bottom holds derivation math only.
  4. Analysis trimmed to NPV results only: NPV A/B/C, ΔNPV (B−A), and the
     tax-incentive benefit (C−B). ROI and payback rows removed.

Data locked in this version:
  - Heat pump CapEx (efficient): ₪1,050/kW (grant data). Baseline (fuel
    oven) CapEx: PENDING — no sourced way to estimate it.
  - Chiller CapEx: efficient ₪4,186/ton (grant median), baseline ₪3,562/ton
    (efficient ÷ 1.175, a 17.5% efficiency-premium midpoint).
  - VSD CapEx (efficient): ₪1,500/kW (grant median of 3 units; the mean was
    outlier-skewed to ~₪2,030). Baseline (fixed-speed): efficient ÷ 1.225
    (22.5% VSD-premium midpoint).
  - Chiller efficiency now differs by capacity: 100 RT = 0.95/0.80 kW/ton,
    500 RT = 0.60/0.48 kW/ton (this was flattened to one point before).
  - Heat pump hours: PENDING (engineer consult), both capacity points.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─── COLORS ───────────────────────────────────────────────────────────────────
F_INPUT   = PatternFill("solid", fgColor="FFFF00")   # yellow – user input / PENDING
F_CONTROL = PatternFill("solid", fgColor="FFC000")   # orange – verify/update
F_HEADING = PatternFill("solid", fgColor="4472C4")   # blue   – section header
F_SUBHEAD = PatternFill("solid", fgColor="BDD7EE")   # l.blue – col headers
F_RESULT  = PatternFill("solid", fgColor="E2EFDA")   # l.green– Scenario A
F_SCN_B   = PatternFill("solid", fgColor="DDEBF7")   # blue-tint – Scenario B
F_SCN_C   = PatternFill("solid", fgColor="FFF2CC")   # cream  – Scenario C
F_POLICY  = PatternFill("solid", fgColor="FF6B6B")   # red    – key policy row
F_SOURCED = PatternFill("solid", fgColor="C6E0B4")   # green  – real sourced data
F_ESTIM   = PatternFill("solid", fgColor="FCE4D6")   # peach  – derived/estimated
F_NONE    = PatternFill("solid", fgColor="FFFFFF")

FONT_H   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
FONT_SH  = Font(name="Arial", bold=True, color="000000", size=10)
FONT_N   = Font(name="Arial", size=10)
FONT_P   = Font(name="Arial", italic=True, color="FF0000", size=10)
FONT_SM  = Font(name="Arial", italic=True, size=9, color="595959")
FONT_B   = Font(name="Arial", bold=True, size=10)

A_R  = Alignment(horizontal="right",  vertical="center", readingOrder=2)
A_C  = Alignment(horizontal="center", vertical="center", readingOrder=2)
A_RW = Alignment(horizontal="right",  vertical="center", readingOrder=2, wrap_text=True)

FMT_NIS = '#,##0 ₪'
FMT_PCT = '0.0%'
FMT_NUM = '#,##0'
FMT_YR  = '0.0'
FMT_X   = '0.0"×"'
FMT_KW      = '#,##0" kW"'
FMT_RT      = '#,##0" RT"'
FMT_KWTON   = '0.00" kW/ton"'
FMT_KWCFM   = '0.0" kW/100cfm"'
FMT_COP     = '0.00'
FMT_TONS    = '#,##0.0" טון"'
FMT_KWH     = '#,##0" קוט""ש"'
FMT_PRICE3  = '#,##0.000'

GLOBAL_SHEET   = "נתונים והנחות"
ANALYSIS_SHEET = "ניתוח"

# ─── Global-sheet cell references (single-value rows) ──────────────────────────
TAX_REF        = f"'{GLOBAL_SHEET}'!$F$10"
DISC_REF       = f"'{GLOBAL_SHEET}'!$F$11"
ELEC_REF       = f"'{GLOBAL_SHEET}'!$F$12"   # agorot/kWh
DIESEL_PRICE   = f"'{GLOBAL_SHEET}'!$F$14"   # ₪/liter
MAZUT_PRICE    = f"'{GLOBAL_SHEET}'!$F$15"   # ₪/ton
DIESEL_CAL     = f"'{GLOBAL_SHEET}'!$F$16"   # ton fuel/MWh
MAZUT_CAL      = f"'{GLOBAL_SHEET}'!$F$17"   # ton fuel/MWh
DIESEL_DENS    = f"'{GLOBAL_SHEET}'!$F$18"   # kg/liter
MULT_REF       = f"'{GLOBAL_SHEET}'!$F$20"   # depreciation multiplier

# ─── Global-sheet tech-table row numbers (match Omri's compact layout) ─────────
R_CAP   = 28   # capacity
R_HRS   = 29   # annual hours
R_COP   = 30   # heat-pump COP (efficient side)
R_COMB  = 31   # baseline oven combustion efficiency (heat pump only)
R_FUEL  = 32   # baseline fuel type (heat pump only): "מזוט" / "סולר"
R_EFFB  = 33   # native-unit efficiency, baseline (chiller/VSD)
R_EFFE  = 34   # native-unit efficiency, efficient (chiller/VSD)
R_SAV   = 35   # computed savings %
R_CONSB = 36   # computed annual consumption, baseline
R_CONSE = 37   # computed annual consumption, efficient (always kWh)
R_CAPXB = 39   # CapEx baseline
R_CAPXE = 40   # CapEx efficient
R_DCAPX = 41   # ΔCapEx
R_LIFEB = 42   # lifetime baseline
R_LIFEE = 43   # lifetime efficient
R_DEGR  = 44   # performance degradation
R_OPXOB = 45   # other OPEX baseline (Rafi, PENDING)
R_OPXOE = 46   # other OPEX efficient (Rafi, PENDING)
R_OPXEB = 47   # computed energy/fuel OPEX baseline
R_OPXEE = 48   # computed energy OPEX efficient

T_START_COL = 6   # column F = first tech column

# ─── TECHNOLOGIES (6: low + high capacity per tech) ───────────────────────────
TECHS = [
    dict(kind="heat_pump", num="3.1א",
         name_eff="משאבות חום — 40 kW", name_base="תנור מזוט/סולר — 40 kW",
         cap=40, cap_fmt=FMT_KW, hours=None,
         cop=4.13, comb=0.835, fuel="סולר",
         capex_b=None, capex_e=42000, life_b=15, life_e=10),
    dict(kind="heat_pump", num="3.1ב",
         name_eff="משאבות חום — 70 kW", name_base="תנור מזוט/סולר — 70 kW",
         cap=70, cap_fmt=FMT_KW, hours=None,
         cop=3.235, comb=0.835, fuel="סולר",
         capex_b=None, capex_e=73500, life_b=15, life_e=10),
    dict(kind="chiller", num="3.2א",
         name_eff="צ'ילרים — 100 RT", name_base="מערכת קירור קונבנציונלית — 100 RT",
         cap=100, cap_fmt=FMT_RT, hours=3000,
         eff_b=0.95, eff_e=0.80, eff_fmt=FMT_KWTON,
         capex_b=356200, capex_e=418600, life_b=15, life_e=17),
    dict(kind="chiller", num="3.2ב",
         name_eff="צ'ילרים — 500 RT", name_base="מערכת קירור קונבנציונלית — 500 RT",
         cap=500, cap_fmt=FMT_RT, hours=3000,
         eff_b=0.60, eff_e=0.48, eff_fmt=FMT_KWTON,
         capex_b=1781000, capex_e=2093000, life_b=15, life_e=17),
    dict(kind="vsd", num="3.3א",
         name_eff="מדחסי VSD — 45 kW", name_base="מדחס מהירות קבועה — 45 kW",
         cap=45, cap_fmt=FMT_KW, hours=6400,
         eff_b=21.5, eff_e=16.5, eff_fmt=FMT_KWCFM,
         capex_b=55102, capex_e=67500, life_b=12, life_e=12),
    dict(kind="vsd", num="3.3ב",
         name_eff="מדחסי VSD — 150 kW", name_base="מדחס מהירות קבועה — 150 kW",
         cap=150, cap_fmt=FMT_KW, hours=6400,
         eff_b=21.5, eff_e=16.5, eff_fmt=FMT_KWCFM,
         capex_b=183673, capex_e=225000, life_b=12, life_e=12),
]

STD_DEPR_PCT = 0.10
STD_DEPR_YRS = 10

MAX_YRS     = 20
YR0_COL     = 6
MAX_COL_IDX = YR0_COL + MAX_YRS
MAX_COL_LTR = get_column_letter(MAX_COL_IDX)
YR0_LTR     = get_column_letter(YR0_COL)


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def sc(ws, row, col, value=None, fill=None, font=None, align=None, fmt=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if fill:  c.fill = fill
    if font:  c.font = font
    if align: c.alignment = align
    if fmt:   c.number_format = fmt
    return c


def section_hdr(ws, row, label, number=None, col_start=2, col_end=8):
    text = f"{number}. {label}" if number else label
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start)
    c.value, c.fill, c.font, c.alignment = text, F_HEADING, FONT_H, A_R


def year_header_row(ws, row, yr0_label="שנה 0"):
    for col, val in [(2, "פרמטר"), (3, "יחידות"), (4, "מקור"), (5, "הערות"), (6, yr0_label)]:
        sc(ws, row, col, val, fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    for i in range(1, MAX_YRS + 1):
        sc(ws, row, 6 + i, f"שנה {i}", fill=F_SUBHEAD, font=FONT_SH, align=A_C)


def scenario_label_row(ws, row, label, fill, col_end=None):
    end = col_end or (6 + MAX_YRS)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=end)
    c = ws.cell(row=row, column=2)
    c.value, c.fill = label, fill
    c.font = Font(name="Arial", bold=True, size=10,
                  color="FFFFFF" if fill in (F_HEADING, F_POLICY) else "000000")
    c.alignment = A_R


def set_col_widths(ws, include_year_cols=False, n_tech=0):
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 44
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 32
    ws.column_dimensions['E'].width = 34
    ws.column_dimensions['F'].width = 17
    for i in range(1, n_tech + 1):
        ws.column_dimensions[get_column_letter(6 + i)].width = 17
    if include_year_cols:
        for i in range(1, MAX_YRS + 1):
            ws.column_dimensions[get_column_letter(6 + i)].width = 12


def color_legend(ws, row=1):
    ws.cell(row=row, column=2).value = "מקרא צבעים"
    ws.cell(row=row, column=2).font = FONT_SH
    items = [
        ("ערך להזנה / ממתין לנתונים", F_INPUT),
        ("נתון לעדכון / בקרה",        F_CONTROL),
        ("נתון ממקור אמיתי",          F_SOURCED),
        ("נתון מוערך/נגזר (עם מקור)", F_ESTIM),
        ("תרחיש A — טכנולוגיה קיימת", F_RESULT),
        ("תרחיש B — יעיל, ללא תמריץ", F_SCN_B),
        ("תרחיש C — יעיל, עם תמריץ",  F_SCN_C),
    ]
    for i, (label, fill) in enumerate(items):
        sc(ws, row + 1 + i, 2, label, fill=fill, font=FONT_N, align=A_R)


def col_of(i):
    return get_column_letter(T_START_COL + i)


# ─── SHEET 1: GLOBAL ASSUMPTIONS ──────────────────────────────────────────────

def build_global_sheet(wb):
    ws = wb.active
    ws.title = GLOBAL_SHEET
    ws.sheet_view.rightToLeft = True
    set_col_widths(ws, n_tech=len(TECHS))
    color_legend(ws, row=1)

    # ── Section 1: Financial parameters ──
    section_hdr(ws, 9, "פרמטרים פיננסיים", number=1, col_end=6 + len(TECHS))
    fin = [
        (10, "שיעור מס חברות", "%", "רשות המסים", "", 0.23, F_CONTROL, FMT_PCT),
        (11, "שיעור היוון (יזמי / פרטי)", "%", "סוכם עם דניאל (2026-07-13)",
             "6% — הוחלט עם דניאל", 0.06, F_SOURCED, FMT_PCT),
        (12, 'מחיר חשמל ממוצע לתעשייה', 'אג\'/קוט"ש', "חברת החשמל — ממוצע SMP 2024",
             'נתון כללי, ייתכן שיעודכן לממוצע משוקלל לפי שעות תעו"ז', 14.0, F_INPUT, '#,##0.0'),
    ]
    for r, lbl, u, src, note, val, fill, fmt in fin:
        sc(ws, r, 2, lbl, fill=fill, font=FONT_N, align=A_R)
        sc(ws, r, 3, u, align=A_R)
        sc(ws, r, 4, src, align=A_RW)
        sc(ws, r, 5, note, align=A_RW)
        sc(ws, r, 6, val, fill=fill, fmt=fmt, align=A_C)

    # ── Section 1b: Fuel prices & conversions ──
    section_hdr(ws, 13, "מחירי דלקים ומקדמי המרה (עבור בסיס משאבות חום)",
                number="1ב", col_end=6 + len(TECHS))
    fuel = [
        (14, "מחיר סולר בתעשייה (בתוספת מרווח שיווק)", "₪/ליטר",
             "משרד האנרגיה — מחירים תיאורטיים 2024", "מרווח שיווק - 0.5 ₪ (מוסיף עמרי)",
             2.59368, F_SOURCED, FMT_PRICE3),
        (15, "מחיר מזוט בתעשייה (בתוספת מרווח שיווק)", "₪/טון",
             "משרד האנרגיה — מחירים תיאורטיים 2024", "מרווח שיווק - 0.5 ₪ (מוסיף עמרי)",
             2344.72, F_SOURCED, FMT_NIS),
        (16, "יחס קלורי — סולר", "טון דלק/MWh", "אקסל של עמרי (MRV)", "", 0.085, F_SOURCED, FMT_PRICE3),
        (17, "יחס קלורי — מזוט", "טון דלק/MWh", "אקסל של עמרי (MRV)", "", 0.088, F_SOURCED, FMT_PRICE3),
        (18, "צפיפות סולר", 'ק"ג/ליטר', "אקסל של עמרי (MRV)",
             "להמרת טון סולר לליטר אם נבחר סולר כדלק בסיס", 0.82, F_SOURCED, FMT_PRICE3),
    ]
    for r, lbl, u, src, note, val, fill, fmt in fuel:
        sc(ws, r, 2, lbl, fill=fill, font=FONT_N, align=A_R)
        sc(ws, r, 3, u, align=A_R)
        sc(ws, r, 4, src, align=A_RW)
        sc(ws, r, 5, note, align=A_RW)
        sc(ws, r, 6, val, fill=fill, fmt=fmt, align=A_C)

    # ── Section 2: Depreciation multiplier ──
    section_hdr(ws, 19, "פרמטר תמריץ — מכפיל פחת (פחת מואץ)", number=2, col_end=6 + len(TECHS))
    sc(ws, 20, 2, "מכפיל שיעור הפחת  ←  פרמטר המדיניות המרכזי", fill=F_POLICY,
       font=Font(name="Arial", bold=True, size=11, color="FFFFFF"), align=A_R)
    sc(ws, 20, 3, "מכפיל", fill=F_POLICY,
       font=Font(name="Arial", bold=True, size=11, color="FFFFFF"), align=A_C)
    sc(ws, 20, 4, "פרמטר ניתן לשינוי", align=A_R)
    sc(ws, 20, 5, "1.0 = ללא תמריץ | 2.0 = כפול, 5 שנים | 5.0 = 2 שנים",
       fill=F_POLICY, font=Font(name="Arial", italic=True, size=9, color="FFFFFF"), align=A_R)
    sc(ws, 20, 6, 2.0, fill=F_INPUT,
       font=Font(name="Arial", bold=True, size=14, color="C00000"), fmt=FMT_X, align=A_C)

    # ── Section 3: Standard depreciation rates ──
    section_hdr(ws, 21, "שיעורי פחת סטנדרטיים — תקנות מס הכנסה", number=3, col_end=6 + len(TECHS))
    for r, lbl in [(22, "משאבות חום"), (23, "צ'ילרים"), (24, "מדחסי VSD")]:
        sc(ws, r, 2, lbl, font=FONT_N, align=A_R)
        sc(ws, r, 3, "% לשנה", align=A_R)
        sc(ws, r, 4, "תקנות פחת — ציוד מכני", align=A_R)
        sc(ws, r, 6, 0.10, fmt=FMT_PCT, align=A_C)

    # ── Section 4: Technology assumptions (6 columns) ──
    T_END_COL = T_START_COL + len(TECHS) - 1
    section_hdr(ws, 25, "הנחות לפי טכנולוגיה — שתי נקודות קיבולת לכל טכנולוגיה",
                number=4, col_end=T_END_COL)

    for col, lbl in [(2, "פרמטר"), (3, "יחידות"), (4, "מקור"), (5, "הערות")]:
        sc(ws, 26, col, lbl, fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    for i, t in enumerate(TECHS):
        sc(ws, 26, T_START_COL + i, t['name_eff'],
           fill=PatternFill("solid", fgColor="1F4E79"),
           font=Font(name="Arial", bold=True, color="FFFFFF", size=9), align=A_C)
        sc(ws, 27, T_START_COL + i, f"vs. {t['name_base']}", font=FONT_SM, align=A_C)

    # Row labels: (row, param, units-col, source-col, notes-col)
    rowmeta = [
        (R_CAP,   "קיבולת מותקנת",               "kW / RT",       "baseline-technology-data.md",        ""),
        (R_HRS,   "שעות פעילות שנתיות",           "שעות/שנה",      "ראו הערות לפי טכנולוגיה",            ""),
        (R_COP,   "COP — משאבת חום (יעיל)",        "יחס",           "Sprsun — עמוד מפרט מוצר",            "משאבות חום בלבד"),
        (R_COMB,  "יעילות בעירה — תנור בסיס",      "%",             "ASME PTC 4 (אמצע 80-88%)",           "משאבות חום בלבד"),
        (R_FUEL,  "סוג דלק בסיס",                  "טקסט",          "הנחת עבודה — ניתן לשנות",            'משאבות חום בלבד — "מזוט"/"סולר"'),
        (R_EFFB,  "יעילות אנרגטית — בסיסי",        "kW/ton | kW/100cfm", "צ'ילר: ASHRAE 90.1 | VSD: CAGI", "צ'ילרים/VSD בלבד"),
        (R_EFFE,  "יעילות אנרגטית — יעיל",         "kW/ton | kW/100cfm", "צ'ילר: DOE FEMP | VSD: CAGI",    "צ'ילרים/VSD בלבד"),
        (R_SAV,   "חיסכון אנרגטי (מחושב)",         "% מצריכה",      "חישוב",                              "צ'ילרים/VSD — נגזר משורות 33-34"),
        (R_CONSB, "צריכת אנרגיה/דלק שנתית — בסיסי", 'טון דלק / קוט"ש', "חישוב — קיבולת × שעות × יעילות",   "משאבות חום: טון דלק | אחר: קוט\"ש"),
        (R_CONSE, "צריכת חשמל שנתית — יעיל",       'קוט"ש/שנה',     "חישוב",                              "הצד היעיל תמיד חשמלי"),
        (R_CAPXB, "CapEx — ציוד קיים (בסיסי)",     "₪",             "ראו הערות לפי טכנולוגיה",            ""),
        (R_CAPXE, "CapEx — ציוד יעיל",             "₪",             "ראו הערות לפי טכנולוגיה",            ""),
        (R_DCAPX, "ΔCapEx = יעיל − בסיסי",         "₪",             "חישוב",                              ""),
        (R_LIFEB, "אורך חיים — ציוד בסיסי",        "שנים",          "ממצאי הנדסה / הנחת עבודה",           ""),
        (R_LIFEE, "אורך חיים — ציוד יעיל",         "שנים",          "ממצאי הנדסה / הנחת עבודה",           ""),
        (R_DEGR,  "גורם שחיקת ביצועים",            "%/שנה",         "הנחת עבודה (0.5-1% טיפוסי)",         ""),
        (R_OPXOB, "OPEX אחר — בסיסי (תחזוקה...)",  "₪/שנה",         "נתוני רפי (טרם התקבל)",              "⚠ ממתין לרפי"),
        (R_OPXOE, "OPEX אחר — יעיל (תחזוקה...)",   "₪/שנה",         "נתוני רפי (טרם התקבל)",              "⚠ ממתין לרפי"),
        (R_OPXEB, "OPEX אנרגיה/דלק שנה 1 — בסיסי", "₪/שנה",         "חישוב",                              ""),
        (R_OPXEE, "OPEX אנרגיה שנה 1 — יעיל",      "₪/שנה",         "חישוב",                              ""),
    ]
    for row, param, units, src, note in rowmeta:
        sc(ws, row, 2, param, font=FONT_N, align=A_R)
        sc(ws, row, 3, units, align=A_R)
        sc(ws, row, 4, src, align=A_RW)
        if note:
            sc(ws, row, 5, note, font=FONT_SM, align=A_RW)

    # Heat-pump hours note
    sc(ws, R_HRS, 5,
       'משאבות חום: ממתין לייעוץ מהנדס | צ\'ילרים: 3,000 (מספר עבודה) | VSD: 6,400 (אמצע 6,000-6,800)',
       font=FONT_SM, align=A_RW)

    # Per-technology values
    for i, t in enumerate(TECHS):
        X = col_of(i)
        col = T_START_COL + i

        # Capacity
        sc(ws, R_CAP, col, t['cap'], fill=F_SOURCED, fmt=t['cap_fmt'], align=A_C)

        # Hours
        if t['hours'] is None:
            sc(ws, R_HRS, col, "PENDING", fill=F_INPUT, font=FONT_P, fmt=FMT_NUM, align=A_C)
        else:
            sc(ws, R_HRS, col, t['hours'], fill=F_ESTIM, fmt=FMT_NUM, align=A_C)

        if t['kind'] == "heat_pump":
            sc(ws, R_COP, col, t['cop'], fill=F_SOURCED, fmt=FMT_COP, align=A_C)
            sc(ws, R_COMB, col, t['comb'], fill=F_ESTIM, fmt=FMT_PCT, align=A_C)
            sc(ws, R_FUEL, col, t['fuel'], fill=F_INPUT, font=FONT_B, align=A_C)
            sc(ws, R_SAV, col, "-", align=A_C)
            # baseline consumption = fuel tons/yr
            sc(ws, R_CONSB, col,
               f'=IF(AND(ISNUMBER({X}{R_CAP}),ISNUMBER({X}{R_HRS}),ISNUMBER({X}{R_COMB})),'
               f'({X}{R_CAP}*{X}{R_HRS}/1000/{X}{R_COMB})*'
               f'IF({X}{R_FUEL}="מזוט",{MAZUT_CAL},{DIESEL_CAL}),"PENDING")',
               fmt=FMT_TONS, align=A_C)
            # efficient consumption = kWh via COP
            sc(ws, R_CONSE, col,
               f'=IF(AND(ISNUMBER({X}{R_CAP}),ISNUMBER({X}{R_HRS}),ISNUMBER({X}{R_COP})),'
               f'{X}{R_CAP}*{X}{R_HRS}/{X}{R_COP},"PENDING")',
               fmt=FMT_KWH, align=A_C)
            # baseline OPEX = fuel cost
            sc(ws, R_OPXEB, col,
               f'=IF(ISNUMBER({X}{R_CONSB}),'
               f'IF({X}{R_FUEL}="מזוט",{X}{R_CONSB}*{MAZUT_PRICE},'
               f'{X}{R_CONSB}*1000/{DIESEL_DENS}*{DIESEL_PRICE}),"PENDING")',
               fmt=FMT_NIS, align=A_C)
        else:
            sc(ws, R_EFFB, col, t['eff_b'], fill=F_SOURCED, fmt=t['eff_fmt'], align=A_C)
            sc(ws, R_EFFE, col, t['eff_e'], fill=F_SOURCED, fmt=t['eff_fmt'], align=A_C)
            sc(ws, R_SAV, col,
               f'=IF(AND(ISNUMBER({X}{R_EFFB}),ISNUMBER({X}{R_EFFE})),'
               f'({X}{R_EFFB}-{X}{R_EFFE})/{X}{R_EFFB},"PENDING")',
               fmt=FMT_PCT, align=A_C)
            if t['kind'] == "chiller":
                # kWh = RT × kW/ton × hours
                sc(ws, R_CONSB, col,
                   f'=IF(AND(ISNUMBER({X}{R_CAP}),ISNUMBER({X}{R_HRS}),ISNUMBER({X}{R_EFFB})),'
                   f'{X}{R_CAP}*{X}{R_EFFB}*{X}{R_HRS},"PENDING")',
                   fmt=FMT_KWH, align=A_C)
            else:  # vsd: kWh = nameplate kW × hours
                sc(ws, R_CONSB, col,
                   f'=IF(AND(ISNUMBER({X}{R_CAP}),ISNUMBER({X}{R_HRS})),'
                   f'{X}{R_CAP}*{X}{R_HRS},"PENDING")',
                   fmt=FMT_KWH, align=A_C)
            sc(ws, R_CONSE, col,
               f'=IF(AND(ISNUMBER({X}{R_CONSB}),ISNUMBER({X}{R_SAV})),'
               f'{X}{R_CONSB}*(1-{X}{R_SAV}),"PENDING")',
               fmt=FMT_KWH, align=A_C)
            sc(ws, R_OPXEB, col,
               f'=IF(ISNUMBER({X}{R_CONSB}),{X}{R_CONSB}*{ELEC_REF}/100,"PENDING")',
               fmt=FMT_NIS, align=A_C)

        # Efficient-side OPEX is always electric
        sc(ws, R_OPXEE, col,
           f'=IF(ISNUMBER({X}{R_CONSE}),{X}{R_CONSE}*{ELEC_REF}/100,"PENDING")',
           fmt=FMT_NIS, align=A_C)

        # CapEx
        if t['capex_b'] is None:
            sc(ws, R_CAPXB, col, "PENDING", fill=F_INPUT, font=FONT_P, fmt=FMT_NIS, align=A_C)
        else:
            sc(ws, R_CAPXB, col, t['capex_b'], fill=F_ESTIM, fmt=FMT_NIS, align=A_C)
        sc(ws, R_CAPXE, col, t['capex_e'],
           fill=(F_ESTIM if t['kind'] == "heat_pump" else F_SOURCED), fmt=FMT_NIS, align=A_C)
        sc(ws, R_DCAPX, col,
           f'=IF(ISNUMBER({X}{R_CAPXE})*ISNUMBER({X}{R_CAPXB}),{X}{R_CAPXE}-{X}{R_CAPXB},"PENDING")',
           fmt=FMT_NIS, align=A_C)

        sc(ws, R_LIFEB, col, t['life_b'], fill=F_CONTROL, fmt=FMT_YR, align=A_C)
        sc(ws, R_LIFEE, col, t['life_e'], fill=F_CONTROL, fmt=FMT_YR, align=A_C)
        sc(ws, R_DEGR, col, 0.005, fill=F_INPUT, fmt=FMT_PCT, align=A_C)
        sc(ws, R_OPXOB, col, "PENDING", fill=F_INPUT, font=FONT_P, fmt=FMT_NIS, align=A_C)
        sc(ws, R_OPXOE, col, "PENDING", fill=F_INPUT, font=FONT_P, fmt=FMT_NIS, align=A_C)

    # Per-tech CapEx source detail (goes in the notes column of the CapEx rows —
    # combined because the three technologies derive it differently)
    sc(ws, R_CAPXE, 5,
       "משאבות חום: מענקים ₪1,050/kW | צ'ילר: מענקים חציון ₪4,186/טון | VSD: מענקים חציון ₪1,500/kW",
       font=FONT_SM, align=A_RW)
    sc(ws, R_CAPXB, 5,
       "צ'ילר: יעיל ÷ 1.175 (פרמיית יעילות 10-25%) | VSD: יעיל ÷ 1.225 (פרמיית VSD 15-30%) | "
       "משאבות חום: אין מקור לתנור בסיס — ממתין",
       font=FONT_SM, align=A_RW)

    # ── Methodology (derivations only, not [n] citations) ──
    R = 51
    section_hdr(ws, R, "מתודולוגיה — גזירות והבהרות", number=5, col_end=6 + len(TECHS))
    R += 1
    notes = [
        "CapEx צ'ילר בסיסי — נגזר מהיעיל: נתוני המענקים כוללים רק את הציוד היעיל שמומן, "
        "ולכן אין מחיר לציוד בסיסי. הבסיסי נאמד כ- ₪4,186 ÷ 1.175 = ₪3,562/טון, "
        "לפי פרמיית יעילות של 10-25% (DOE FEMP + מקורות שוק, אמצע 17.5%). הערכה, לא ציטוט ישיר.",
        "CapEx VSD בסיסי — נגזר בדומה: ₪1,500 ÷ 1.225 = ₪1,224/kW, לפי פרמיית VSD מול "
        "מהירות קבועה של 15-30% (אמצע 22.5%). ₪1,500/kW היעיל = חציון של 3 יחידות מהמענקים "
        "(הממוצע ~₪2,030 מוטה ע\"י יחידה חריגה יקרה).",
        "CapEx משאבת חום בסיסי (תנור מזוט/סולר) — נותר PENDING: בניגוד ל-VSD, תנור בעירה אינו "
        "\"אותה טכנולוגיה ברמת יעילות אחרת\", ולכן שיטת הפרמיה לא חלה. אין מקור נאות עדיין.",
        "יעילות צ'ילר משתנה לפי קיבולת: 100 RT = 0.95/0.80 kW/ton, 500 RT = 0.60/0.48 kW/ton "
        "(ASHRAE 90.1 מינימום תקן מול DOE FEMP יעיל). יעילות VSD (specific power) אינה משתנה "
        "לפי קיבולת — היא השוואת טכנולוגיה (מהירות קבועה מול VSD).",
        "מחירי דלק — משרד האנרגיה, מינהל הדלק והגז, \"מחירים תיאורטיים של מוצרי דלק שאינם "
        "בפיקוח\", 2024 (לפני מע\"מ ובלו). מרווח שיווק של 0.5 ₪ מתווסף ידנית ע\"י עמרי.",
        "OPEX אחר (תחזוקה, חלפים) — ממתין לנתוני רפי לכל 3 הטכנולוגיות.",
    ]
    for txt in notes:
        ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=6 + len(TECHS))
        sc(ws, R, 2, "• " + txt, font=FONT_SM, align=A_RW)
        ws.row_dimensions[R].height = 28
        R += 1

    # Build refs dict per tech
    tech_refs = []
    for i, t in enumerate(TECHS):
        X = col_of(i)
        s = GLOBAL_SHEET
        tech_refs.append(dict(
            capex_b=f"'{s}'!${X}${R_CAPXB}", capex_e=f"'{s}'!${X}${R_CAPXE}",
            dcapex=f"'{s}'!${X}${R_DCAPX}",
            opex_o_b=f"'{s}'!${X}${R_OPXOB}", opex_o_e=f"'{s}'!${X}${R_OPXOE}",
            opex_e_b=f"'{s}'!${X}${R_OPXEB}", opex_e_e=f"'{s}'!${X}${R_OPXEE}",
        ))
    return tech_refs


# ─── SHEET 2: ANALYSIS ────────────────────────────────────────────────────────

def build_analysis_sheet(wb, tech_refs):
    ws = wb.create_sheet(ANALYSIS_SHEET)
    ws.sheet_view.rightToLeft = True
    set_col_widths(ws, include_year_cols=True)
    color_legend(ws, row=1)

    tech_end_col = 6 + MAX_YRS
    ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=tech_end_col)
    sc(ws, 9, 2, "ניתוח כלכלי — פחת מואץ לפי טכנולוגיה | NPV בלבד",
       fill=F_HEADING, font=Font(name="Arial", bold=True, color="FFFFFF", size=13), align=A_R)

    section_hdr(ws, 11, "פרמטרים פיננסיים", number=1)
    for r, lbl, u, ref, fmt, fill in [
        (12, "שיעור מס חברות", "%", f"={TAX_REF}", FMT_PCT, F_CONTROL),
        (13, "שיעור היוון", "%", f"={DISC_REF}", FMT_PCT, F_SOURCED),
        (14, 'מחיר חשמל (₪/קוט"ש)', '₪/קוט"ש', f"={ELEC_REF}/100", '#,##0.000', F_INPUT),
    ]:
        sc(ws, r, 2, lbl, fill=fill, font=FONT_N, align=A_R)
        sc(ws, r, 3, u, align=A_R)
        sc(ws, r, 4, GLOBAL_SHEET, align=A_R)
        sc(ws, r, 6, ref, fill=fill, fmt=fmt, align=A_C)

    section_hdr(ws, 16, "פרמטר תמריץ — מכפיל פחת", number=2)
    sc(ws, 17, 2, "מכפיל שיעור הפחת  ←  שנה ב'נתונים והנחות'!F20", fill=F_POLICY,
       font=Font(name="Arial", bold=True, size=11, color="FFFFFF"), align=A_R)
    sc(ws, 17, 3, "מכפיל", fill=F_POLICY,
       font=Font(name="Arial", bold=True, size=11, color="FFFFFF"), align=A_C)
    sc(ws, 17, 4, GLOBAL_SHEET, align=A_R)
    sc(ws, 17, 6, f"={MULT_REF}", fill=F_INPUT,
       font=Font(name="Arial", bold=True, size=13, color="C00000"), fmt=FMT_X, align=A_C)

    section_hdr(ws, 19, "ניתוחים לפי טכנולוגיה", number=3, col_end=tech_end_col)

    R = 21
    tech_results = []
    for i, t in enumerate(TECHS):
        R, rmap = _tech_block(ws, t, R, tech_refs[i])
        tech_results.append((t, rmap))
        R += 2

    _summary_block(ws, tech_results, R)


def _tech_block(ws, t, R_start, refs):
    R = R_start
    tech_end_col = 6 + MAX_YRS
    life_e, life_b = t['life_e'], t['life_b']
    std_r, std_yrs = STD_DEPR_PCT, STD_DEPR_YRS

    capex_b, capex_e = refs['capex_b'], refs['capex_e']
    opex_e_b, opex_e_e = refs['opex_e_b'], refs['opex_e_e']
    opex_o_b, opex_o_e = refs['opex_o_b'], refs['opex_o_e']
    degrad_v = 0.005

    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, f"{t['num']}  {t['name_eff']}",
       fill=PatternFill("solid", fgColor="1F4E79"),
       font=Font(name="Arial", bold=True, color="FFFFFF", size=11), align=A_R)
    R += 1
    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, f"* מול {t['name_base']} | נתונים: גיליון '{GLOBAL_SHEET}' סעיף 4",
       font=FONT_SM, align=A_R)
    R += 1

    # ── ב: Depreciation ──
    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, "ב. לוחות פחת (ציוד בסיסי vs. יעיל)", fill=F_SUBHEAD, font=FONT_SH, align=A_R)
    R += 1
    year_header_row(ws, R, yr0_label="שנה 0 (השקעה)")
    R += 1

    R_SD_B = R; R += 1
    R_ST_B = R; R += 1
    R_SD_E = R; R += 1
    R_ST_E = R; R += 1
    R_AD_E = R; R += 1
    R_AT_E = R; R += 1

    for rr, lbl, fill in [
        (R_SD_B, f"פחת סטנדרטי — {t['name_base']}", F_RESULT),
        (R_ST_B, f"מגן מס — {t['name_base']}", F_RESULT),
        (R_SD_E, f"פחת סטנדרטי — {t['name_eff']}", F_SCN_B),
        (R_ST_E, f"מגן מס סטנדרטי — {t['name_eff']}", F_SCN_B),
        (R_AD_E, f"פחת מואץ — {t['name_eff']}", F_SCN_C),
        (R_AT_E, f"מגן מס מואץ — {t['name_eff']}", F_SCN_C),
    ]:
        sc(ws, rr, 2, lbl, fill=fill, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪", align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)
        sc(ws, rr, 6, 0, fill=fill, fmt=FMT_NIS, align=A_C)

    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl = get_column_letter(col)
        sd_b = (f"=IF(ISNUMBER({capex_b}),{capex_b}*{std_r},0)" if i <= std_yrs else "=0")
        sc(ws, R_SD_B, col, sd_b, fill=F_RESULT, fmt=FMT_NIS, align=A_C)
        sc(ws, R_ST_B, col, f"={cl}{R_SD_B}*{TAX_REF}", fill=F_RESULT, fmt=FMT_NIS, align=A_C)
        sd_e = (f"=IF(ISNUMBER({capex_e}),{capex_e}*{std_r},0)" if i <= std_yrs else "=0")
        sc(ws, R_SD_E, col, sd_e, fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
        sc(ws, R_ST_E, col, f"={cl}{R_SD_E}*{TAX_REF}", fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
        ad_e = (f"=IF(ISNUMBER({capex_e}),IF({i}<=ROUND(1/({std_r}*{MULT_REF}),0),"
                f"{capex_e}*{std_r}*{MULT_REF},0),0)")
        sc(ws, R_AD_E, col, ad_e, fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
        sc(ws, R_AT_E, col, f"={cl}{R_AD_E}*{TAX_REF}", fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
    R += 1

    # ── ג: Cash flow ──
    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=tech_end_col)
    sc(ws, R, 2, "ג. ניתוח תזרים מזומנים — שלושה תרחישים", fill=F_SUBHEAD, font=FONT_SH, align=A_R)
    R += 1
    year_header_row(ws, R, yr0_label="שנה 0 (השקעה)")
    R += 1

    # Scenario A
    scenario_label_row(ws, R, f"תרחיש A — רכישת {t['name_base']} (ציוד קיים)",
                       PatternFill("solid", fgColor="538135"))
    R += 1
    R_A_INV = R; R += 1
    R_A_EOPC = R; R += 1
    R_A_OOPC = R; R += 1
    R_A_NET = R; R += 1
    R_A_DISC = R; R += 1
    for rr, lbl in [(R_A_INV, "השקעה ראשונית"), (R_A_EOPC, "OPEX אנרגיה/דלק"),
                    (R_A_OOPC, "OPEX אחר"), (R_A_NET, "תזרים נקי — A"),
                    (R_A_DISC, "תזרים מהוון — A")]:
        sc(ws, rr, 2, lbl, fill=F_RESULT, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪", align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)
    sc(ws, R_A_INV, 6, f'=IF(ISNUMBER({capex_b}),-{capex_b},"PENDING")', fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_EOPC, 6, 0, fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_OOPC, 6, 0, fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_NET, 6, f"=F{R_A_INV}", fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    sc(ws, R_A_DISC, 6, f"=F{R_A_NET}", fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl = get_column_letter(col)
        eopc = (f"=IF(ISNUMBER({opex_e_b}),-{opex_e_b},0)" if i <= life_b else "=0")
        sc(ws, R_A_EOPC, col, eopc, fill=F_RESULT, fmt=FMT_NIS, align=A_C)
        oopc = (f"=IF(ISNUMBER({opex_o_b}),-{opex_o_b},0)" if i <= life_b else "=0")
        sc(ws, R_A_OOPC, col, oopc, fill=F_RESULT, fmt=FMT_NIS, align=A_C)
        sc(ws, R_A_NET, col, f"={cl}{R_A_EOPC}+{cl}{R_A_OOPC}+{cl}{R_ST_B}", fill=F_RESULT, fmt=FMT_NIS, align=A_C)
        sc(ws, R_A_DISC, col, f"={cl}{R_A_NET}/(1+{DISC_REF})^{i}", fill=F_RESULT, fmt=FMT_NIS, align=A_C)
    R += 1

    # Scenario B
    scenario_label_row(ws, R, f"תרחיש B — רכישת {t['name_eff']} | פחת סטנדרטי (ללא תמריץ)",
                       PatternFill("solid", fgColor="2E75B6"))
    R += 1
    R_B_INV = R; R += 1
    R_B_EOPC = R; R += 1
    R_B_OOPC = R; R += 1
    R_B_NET = R; R += 1
    R_B_DISC = R; R += 1
    for rr, lbl in [(R_B_INV, "השקעה ראשונית"), (R_B_EOPC, "OPEX אנרגיה (עם שחיקה)"),
                    (R_B_OOPC, "OPEX אחר"), (R_B_NET, "תזרים נקי — B"),
                    (R_B_DISC, "תזרים מהוון — B")]:
        sc(ws, rr, 2, lbl, fill=F_SCN_B, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪", align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)
    sc(ws, R_B_INV, 6, f'=IF(ISNUMBER({capex_e}),-{capex_e},"PENDING")', fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_EOPC, 6, 0, fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_OOPC, 6, 0, fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_NET, 6, f"=F{R_B_INV}", fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    sc(ws, R_B_DISC, 6, f"=F{R_B_NET}", fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl = get_column_letter(col)
        eopc = (f"=IF(ISNUMBER({opex_e_e}),-{opex_e_e}*(1-{degrad_v})^{i-1},0)" if i <= life_e else "=0")
        sc(ws, R_B_EOPC, col, eopc, fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
        oopc = (f"=IF(ISNUMBER({opex_o_e}),-{opex_o_e},0)" if i <= life_e else "=0")
        sc(ws, R_B_OOPC, col, oopc, fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
        sc(ws, R_B_NET, col, f"={cl}{R_B_EOPC}+{cl}{R_B_OOPC}+{cl}{R_ST_E}", fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
        sc(ws, R_B_DISC, col, f"={cl}{R_B_NET}/(1+{DISC_REF})^{i}", fill=F_SCN_B, fmt=FMT_NIS, align=A_C)
    R += 1

    # Scenario C
    scenario_label_row(ws, R, f"תרחיש C — רכישת {t['name_eff']} | פחת מואץ (עם תמריץ)",
                       PatternFill("solid", fgColor="7B6000"))
    R += 1
    sc(ws, R, 2, "* OPEX זהה ל-B — ההבדל הוא מגן המס המואץ בלבד", font=FONT_SM, align=A_R)
    R += 1
    R_C_NET = R; R += 1
    R_C_DISC = R; R += 1
    for rr, lbl in [(R_C_NET, "תזרים נקי — C"), (R_C_DISC, "תזרים מהוון — C")]:
        sc(ws, rr, 2, lbl, fill=F_SCN_C, font=FONT_N, align=A_R)
        sc(ws, rr, 3, "₪", align=A_R)
        sc(ws, rr, 4, "חישוב", align=A_R)
    sc(ws, R_C_NET, 6, f"=F{R_B_INV}", fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
    sc(ws, R_C_DISC, 6, f"=F{R_C_NET}", fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
    for i in range(1, MAX_YRS + 1):
        col = 6 + i
        cl = get_column_letter(col)
        sc(ws, R_C_NET, col, f"={cl}{R_B_EOPC}+{cl}{R_B_OOPC}+{cl}{R_AT_E}", fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
        sc(ws, R_C_DISC, col, f"={cl}{R_C_NET}/(1+{DISC_REF})^{i}", fill=F_SCN_C, fmt=FMT_NIS, align=A_C)
    R += 1

    # ── ד: Results (NPV only) ──
    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=8)
    sc(ws, R, 2, "ד. תוצאות — NPV", fill=F_SUBHEAD, font=FONT_SH, align=A_R)
    R += 1

    def result_row(row, label, formula, fill):
        sc(ws, row, 2, label, fill=fill, font=Font(name="Arial", bold=True, size=10), align=A_R)
        sc(ws, row, 3, 'ש"ח', align=A_R)
        sc(ws, row, 6, formula, fill=fill, fmt=FMT_NIS, align=A_C)

    R_NPV_A = R; R += 1
    R_NPV_B = R; R += 1
    R_NPV_C = R; R += 1
    R += 1
    R_DNPV = R; R += 1
    R_INCV = R; R += 1

    YR0 = YR0_LTR
    result_row(R_NPV_A, f"NPV — תרחיש A ({t['name_base']})",
               f"=SUM({YR0}{R_A_DISC}:{MAX_COL_LTR}{R_A_DISC})", F_RESULT)
    result_row(R_NPV_B, f"NPV — תרחיש B ({t['name_eff']}, ללא תמריץ)",
               f"=SUM({YR0}{R_B_DISC}:{MAX_COL_LTR}{R_B_DISC})", F_SCN_B)
    result_row(R_NPV_C, f"NPV — תרחיש C ({t['name_eff']}, עם תמריץ)",
               f"=SUM({YR0}{R_C_DISC}:{MAX_COL_LTR}{R_C_DISC})", F_SCN_C)
    result_row(R_DNPV, "NPV מצטבר — הצדקת השדרוג (B−A)", f"=F{R_NPV_B}-F{R_NPV_A}", F_INPUT)
    result_row(R_INCV, "ערך התמריץ — תועלת הפחת המואץ (C−B)", f"=F{R_NPV_C}-F{R_NPV_B}", F_POLICY)

    return R + 1, dict(npv_a=R_NPV_A, npv_b=R_NPV_B, npv_c=R_NPV_C, d_npv=R_DNPV, inc=R_INCV)


def _summary_block(ws, tech_results, R_start):
    R = R_start
    section_hdr(ws, R, "סיכום השוואתי לפי טכנולוגיה — NPV", number=4, col_end=8)
    R += 1
    sc(ws, R, 2,
       f"* מכפיל פחת = '{GLOBAL_SHEET}'!F20 — שנה שם ותוצאות מתעדכנות | OPEX אחר וחלק מה-CapEx ממתינים",
       font=FONT_SM, align=A_R)
    R += 1
    hdrs = [(2, "טכנולוגיה"), (3, "NPV A (בסיסי)"), (4, "NPV B (יעיל, ללא)"),
            (5, "NPV C (יעיל, עם)"), (6, "ΔNPV B−A"), (7, "ערך תמריץ C−B")]
    for col, lbl in hdrs:
        sc(ws, R, col, lbl, fill=F_SUBHEAD, font=FONT_SH, align=A_C)
    R += 1
    for t, rmap in tech_results:
        sc(ws, R, 2, t['name_eff'], font=FONT_N, align=A_R)
        for col, key, fill in [(3, 'npv_a', F_RESULT), (4, 'npv_b', F_SCN_B),
                               (5, 'npv_c', F_SCN_C), (6, 'd_npv', F_INPUT), (7, 'inc', F_POLICY)]:
            sc(ws, R, col, f"=$F${rmap[key]}", fill=fill, fmt=FMT_NIS, align=A_C)
        R += 1
    sc(ws, R + 1, 2,
       "** ערכי PENDING מתעדכנים אוטומטית עם קבלת הנתונים (שעות משאבות חום, CapEx בסיס משאבות חום, OPEX אחר)",
       font=Font(name="Arial", italic=True, color="FF0000", size=9), align=A_R)


def main():
    wb = openpyxl.Workbook()
    tech_refs = build_global_sheet(wb)
    build_analysis_sheet(wb, tech_refs)
    out = "/home/user/shmags-2/projects/energy-program/tax_incentive_model_v3.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print("Sheets:", [ws.title for ws in wb.worksheets])


if __name__ == "__main__":
    main()
